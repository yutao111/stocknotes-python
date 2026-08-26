import io
import gc
import json
import os
import re
import sqlite3
import tempfile
import threading
import time
import unittest
from datetime import date, datetime, timedelta
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook


class StockNotesTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        os.environ["STOCKNOTES_DB"] = os.path.join(cls.temp_dir.name, "test.db")
        import app
        cls.module = app
        cls.client = app.app.test_client()

    @classmethod
    def tearDownClass(cls):
        gc.collect()
        cls.temp_dir.cleanup()

    def setUp(self):
        with self.module.db_connect() as db:
            db.execute("DELETE FROM alert_signal_event_outcomes")
            db.execute("DELETE FROM alert_signal_events")
            db.execute("DELETE FROM alert_signal_outcomes")
            db.execute("DELETE FROM alert_signal_samples")
            db.execute("DELETE FROM notifications")
            db.execute("DELETE FROM intraday_rebound_states")
            db.execute("DELETE FROM alert_rule_states")
            db.execute(
                "UPDATE alert_types SET params_json = ?, enabled = 1 WHERE code = ?",
                (json.dumps(self.module.THREE_DAY_DIP_DEFAULT_PARAMS), self.module.THREE_DAY_DIP_CODE),
            )
            db.execute(
                "UPDATE alert_types SET params_json = ?, enabled = 1 WHERE code = ?",
                (json.dumps(self.module.INTRADAY_REBOUND_DEFAULT_PARAMS), self.module.INTRADAY_REBOUND_CODE),
            )
            db.execute(
                "UPDATE alert_types SET params_json = ?, enabled = 1 WHERE code = ?",
                (
                    json.dumps(self.module.CONSOLIDATION_STABILIZATION_DEFAULT_PARAMS),
                    self.module.CONSOLIDATION_STABILIZATION_CODE,
                ),
            )
            for code, params in (
                (self.module.WATCHLIST_OPEN_GAIN_CODE, self.module.WATCHLIST_OPEN_GAIN_DEFAULT_PARAMS),
                (self.module.WATCHLIST_OPEN_LOSS_CODE, self.module.WATCHLIST_OPEN_LOSS_DEFAULT_PARAMS),
                (self.module.WATCHLIST_LIMIT_UP_CODE, {}),
                (self.module.WATCHLIST_LIMIT_DOWN_CODE, {}),
            ):
                db.execute(
                    "UPDATE alert_types SET params_json = ?, enabled = 1 WHERE code = ?",
                    (json.dumps(params), code),
                )
            db.execute("DELETE FROM account_snapshots")
            db.execute("DELETE FROM watchlist_stocks")
            db.execute("DELETE FROM trade_excursion_metrics")
            db.execute("DELETE FROM daily_prices")
            db.execute("DELETE FROM intraday_quotes")
            db.execute("DELETE FROM trade_review_reason_types")
            db.execute("DELETE FROM trade_reviews")
            db.execute("DELETE FROM trade_episode_executions")
            db.execute("DELETE FROM trade_episodes")
            db.execute("DELETE FROM fifo_matches")
            db.execute("DELETE FROM positions")
            db.execute("DELETE FROM current_positions")
            db.execute("DELETE FROM unmatched_sells")
            db.execute("DELETE FROM executions")
            db.execute("DELETE FROM import_jobs")
            db.execute("DELETE FROM users WHERE name <> 'yutaoGS'")
            user = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()
        with self.client.session_transaction() as session:
            session["user_id"] = user["id"]

    def make_statement(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格", "成交金额", "佣金", "印花税", "其他杂费", "其他费", "过户费", "交易市场"])
        sheet.append([20260801, "09:31:00", 2156, "通富微电", "证券买入", 100, "B1", 10, 1000, 1, 0, 0, 0, 0, "深圳A股"])
        sheet.append([20260802, "10:00:00", 2156, "通富微电", "证券买入", 100, "B2", 12, 1200, 1, 0, 0, 0, 0, "深圳A股"])
        sheet.append([20260803, "13:00:00", 2156, "通富微电", "证券卖出", -150, "S1", 14, 2100, 1, 2, 0, 0, 0, "深圳A股"])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def import_statement(self, days=1):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格", "成交金额", "佣金", "印花税", "其他杂费", "其他费", "过户费", "交易市场"])
        for day in range(1, days + 1):
            sheet.append([20260100 + day, f"{9 + day % 9}:00:00", 2156, "通富微电", "证券买入", 100, f"B{day}", 10, 1000, 1, 0, 0, 0, 0, "深圳A股"])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        response = self.client.post("/admin/preview", data={"statement": (output, "statement.xlsx")}, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200)
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
        response = self.client.post(f"/admin/import/{job['id']}", follow_redirects=True)
        self.assertEqual(response.status_code, 200)

    def test_import_fifo_and_duplicate(self):
        response = self.client.post("/admin/preview", data={"statement": (self.make_statement(), "statement.xlsx")}, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200)
        preview_text = response.get_data(as_text=True)
        self.assertIn("解析完成", preview_text)
        self.assertIn("有效记录", preview_text)
        self.assertIn("预计新增", preview_text)
        self.assertIn("确认导入 3 条记录", preview_text)
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs ORDER BY created_at DESC LIMIT 1").fetchone()
        response = self.client.post(f"/admin/import/{job['id']}", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions").fetchone()[0], 3)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM fifo_matches").fetchone()[0], 2)
            position = db.execute("SELECT * FROM positions").fetchone()
            self.assertEqual(position["stock_code"], "002156")
            self.assertAlmostEqual(position["quantity"], 50)
            profit = db.execute("SELECT SUM(profit) FROM fifo_matches").fetchone()[0]
            self.assertAlmostEqual(profit, 495.5, places=2)
            episode = db.execute("SELECT * FROM trade_episodes").fetchone()
            self.assertEqual(episode["status"], "OPEN")
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_episode_executions").fetchone()[0], 3)

        response = self.client.post("/admin/preview", data={"statement": (self.make_statement(), "statement.xlsx")}, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200)
        with self.module.db_connect() as db:
            duplicate_job = db.execute("SELECT * FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
            self.assertEqual(duplicate_job["duplicate_rows"], 3)
        self.client.post(f"/admin/import/{duplicate_job['id']}")
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions").fetchone()[0], 3)

    def test_manual_execution_is_reconciled_by_statement(self):
        response = self.client.post(
            "/admin/manual",
            data={
                "trade_date": "2026-08-01",
                "trade_time": "09:30:00",
                "stock_code": "2156",
                "stock_name": "通富微电",
                "action": "BUY",
                "quantity": "100",
                "deal_price": "10",
                "commission": "0",
                "stamp_tax": "0",
                "transfer_fee": "0",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("已补录当天成交", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            manual = db.execute("SELECT * FROM executions").fetchone()
            self.assertEqual(manual["source"], "MANUAL")
            self.assertEqual(db.execute("SELECT COUNT(*) FROM positions").fetchone()[0], 1)

        self.client.post(
            "/admin/preview",
            data={"statement": (self.make_statement(), "statement.xlsx")},
            content_type="multipart/form-data",
        )
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
        response = self.client.post(f"/admin/import/{job['id']}", follow_redirects=True)
        text = response.get_data(as_text=True)
        self.assertIn("转正补录 1 条", text)
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions").fetchone()[0], 3)
            reconciled = db.execute("SELECT * FROM executions WHERE deal_id = 'B1'").fetchone()
            self.assertEqual(reconciled["source"], "STATEMENT")
            self.assertEqual(reconciled["trade_time"], "09:31:00")
            self.assertAlmostEqual(reconciled["commission"], 1)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM fifo_matches").fetchone()[0], 2)

    def test_manual_execution_validation(self):
        response = self.client.post(
            "/admin/manual",
            data={
                "trade_date": "2026-08-01",
                "trade_time": "09:30:00",
                "stock_code": "2156",
                "stock_name": "通富微电",
                "action": "BUY",
                "quantity": "0",
                "deal_price": "10",
                "commission": "0",
                "stamp_tax": "0",
                "transfer_fee": "0",
            },
            follow_redirects=True,
        )
        self.assertIn("成交数量必须大于 0", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions").fetchone()[0], 0)

    def test_trade_episode_groups_multiple_buys_and_split_sells(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格", "成交金额"])
        sheet.append([20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10, 1000])
        sheet.append([20260802, "10:00:00", 2156, "通富微电", "证券买入", 100, "B2", 12, 1200])
        sheet.append([20260803, "11:00:00", 2156, "通富微电", "证券卖出", -50, "S1", 13, 650])
        sheet.append([20260804, "14:00:00", 2156, "通富微电", "证券卖出", -150, "S2", 14, 2100])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "episode.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{job['id']}")

        with self.module.db_connect() as db:
            episodes = db.execute("SELECT * FROM trade_episodes").fetchall()
            self.assertEqual(len(episodes), 1)
            episode_id = episodes[0]["id"]
            self.assertEqual(episodes[0]["status"], "CLOSED")
            links = db.execute("SELECT role FROM trade_episode_executions ORDER BY execution_id").fetchall()
            self.assertEqual([row["role"] for row in links], ["ENTRY", "ADD", "REDUCE", "EXIT"])

        detail = self.client.get(f"/analysis/trades/{episode_id}")
        self.assertEqual(detail.status_code, 200)
        text = detail.get_data(as_text=True)
        self.assertIn("交易复盘详情", text)
        self.assertIn("通富微电", text)
        self.assertIn("首次建仓", text)
        self.assertIn("完全退出", text)
        self.assertIn("我的交易复盘", text)

        self.client.post("/admin/rebuild")
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT id FROM trade_episodes").fetchone()["id"], episode_id)
        self.assertEqual(self.client.get(f"/analysis/trades/{episode_id}").status_code, 200)

    def test_full_exit_then_reentry_creates_new_episode(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格"])
        sheet.append([20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10])
        sheet.append([20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 11])
        sheet.append([20260803, "11:00:00", 2156, "通富微电", "证券买入", 200, "B2", 12])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "reentry.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{job['id']}")
        with self.module.db_connect() as db:
            episodes = db.execute("SELECT status FROM trade_episodes ORDER BY opened_at").fetchall()
            self.assertEqual([row["status"] for row in episodes], ["CLOSED", "OPEN"])

    def _insert_review(self, episode_id, user_id=None, **overrides):
        with self.module.db_connect() as db:
            user_id = user_id or db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            episode = db.execute(
                "SELECT * FROM trade_episodes WHERE id = ? AND user_id = ?", (episode_id, user_id)
            ).fetchone()
            fields = {
                "user_id": user_id,
                "trade_episode_id": episode_id,
                "review_status": "PENDING",
                "trade_reason": "回调买入",
                "expected_profit_percent": 20.0,
                "expected_target_price": 12.0,
                "stop_loss_price": 9.0,
                "expected_holding_days": 10,
                "confidence_level": 4,
                "sell_reason": "达到目标",
                "judgement_result": "CORRECT",
                "main_problem": "NO_OBVIOUS_PROBLEM",
                "review_note": "执行顺利",
                "next_action": "继续按计划执行",
                "original_opening_execution_id": episode["opening_execution_id"],
                "stock_code_snapshot": episode["stock_code"],
                "opened_at_snapshot": episode["opened_at"],
                "closed_at_snapshot": episode["closed_at"],
                "created_at": "2026-08-21T00:00:00",
                "updated_at": "2026-08-21T00:00:00",
                "completed_at": None,
            }
            fields.update(overrides)
            columns = ", ".join(fields)
            placeholders = ", ".join("?" for _ in fields)
            db.execute(
                f"INSERT INTO trade_reviews ({columns}) VALUES ({placeholders})",
                tuple(fields.values()),
            )

    def test_trade_review_schema_and_constraints(self):
        self.client.post("/admin/preview", data={"statement": (self.make_statement(), "s.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
            tables = {row["name"] for row in db.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
        self.client.post(f"/admin/import/{job['id']}")
        self.assertIn("trade_reviews", tables)
        self.assertIn("trade_review_reason_types", tables)

        with self.module.db_connect() as db:
            episode_id = db.execute("SELECT id FROM trade_episodes").fetchone()["id"]
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
        self._insert_review(episode_id)
        with self.module.db_connect() as db:
            review_id = db.execute("SELECT id FROM trade_reviews").fetchone()["id"]
        with self.module.db_connect() as db:
            db.execute("INSERT INTO trade_review_reason_types (trade_review_id, reason_type) VALUES (?, 'PULLBACK')", (review_id,))
            db.execute("INSERT INTO trade_review_reason_types (trade_review_id, reason_type) VALUES (?, 'TECHNICAL_PATTERN')", (review_id,))
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_review_reason_types").fetchone()[0], 2)
        with self.module.db_connect() as db:
            with self.assertRaises(sqlite3.IntegrityError):
                db.execute("INSERT INTO trade_review_reason_types (trade_review_id, reason_type) VALUES (?, 'PULLBACK')", (review_id,))
            with self.assertRaises(sqlite3.IntegrityError):
                db.execute("INSERT INTO trade_review_reason_types (trade_review_id, reason_type) VALUES (?, 'INVALID')", (review_id,))
        with self.assertRaises(sqlite3.IntegrityError):
            self._insert_review(episode_id, user_id=user_id)
        with self.assertRaises(sqlite3.IntegrityError):
            self._insert_review(episode_id, confidence_level=0)
        with self.assertRaises(sqlite3.IntegrityError):
            self._insert_review(episode_id, confidence_level=6)
        with self.assertRaises(sqlite3.IntegrityError):
            self._insert_review(episode_id, expected_target_price=0)
        with self.assertRaises(sqlite3.IntegrityError):
            self._insert_review(episode_id, main_problem="INVALID")
        with self.assertRaises(sqlite3.IntegrityError):
            self._insert_review(episode_id, review_status="INVALID")

        self.client.post("/users/create", data={"name": "second", "next": "/analysis/stocks"})
        with self.module.db_connect() as db:
            second_id = db.execute("SELECT id FROM users WHERE name = 'second'").fetchone()["id"]
        self.client.post("/users/switch", data={"user_id": second_id, "next": "/analysis/stocks"})
        self.client.post("/admin/preview", data={"statement": (self.make_statement(), "second.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            second_job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{second_job['id']}")
        with self.module.db_connect() as db:
            second_episode = db.execute("SELECT id FROM trade_episodes WHERE user_id = ?", (second_id,)).fetchone()["id"]
        self._insert_review(second_episode, user_id=second_id)
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_reviews").fetchone()[0], 2)
            db.execute("DELETE FROM users WHERE id = ?", (second_id,))
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_reviews").fetchone()[0], 1)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_review_reason_types").fetchone()[0], 2)

    def test_trade_review_survives_episode_replacement(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格"])
        sheet.append([20260810, "09:30:00", 2156, "通富微电", "证券买入", 100, "B10", 10])
        sheet.append([20260811, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S11", 11])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "first.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{job['id']}")
        with self.module.db_connect() as db:
            episode_id = db.execute("SELECT id FROM trade_episodes").fetchone()["id"]
            opening_id = db.execute("SELECT opening_execution_id FROM trade_episodes").fetchone()["opening_execution_id"]
            opened_at = db.execute("SELECT opened_at FROM trade_episodes").fetchone()["opened_at"]
        self._insert_review(episode_id)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格"])
        sheet.append([20260801, "09:30:00", 2156, "通富微电", "证券买入", 200, "B01", 8])
        sheet.append([20260803, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S03", 9])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "earlier.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{job['id']}")

        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_episodes").fetchone()[0], 1)
            review = db.execute("SELECT * FROM trade_reviews").fetchone()
            self.assertIsNotNone(review)
            self.assertIsNone(review["trade_episode_id"])
            self.assertEqual(review["original_opening_execution_id"], opening_id)
            self.assertEqual(review["stock_code_snapshot"], "002156")
            self.assertEqual(review["opened_at_snapshot"], opened_at)
            self.assertEqual(review["review_note"], "执行顺利")

    def _import_workbook_rows(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格"])
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "ep.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
        self.client.post(f"/admin/import/{job['id']}")
        with self.module.db_connect() as db:
            return db.execute("SELECT id FROM trade_episodes ORDER BY id DESC LIMIT 1").fetchone()["id"]

    def _add_current_position(self, stock_code="2156", stock_name="通富微电", quantity="100", avg_cost="10", first_buy_date="2026-08-01"):
        response = self.client.post(
            "/analysis/portfolio/position",
            data={
                "stock_code": stock_code,
                "stock_name": stock_name,
                "quantity": quantity,
                "avg_cost": avg_cost,
                "first_buy_date": first_buy_date,
            },
            follow_redirects=True,
        )
        return response

    def test_trade_review_form_renders_and_saves(self):
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 12],
        ])
        page = self.client.get(f"/analysis/trades/{episode_id}").get_data(as_text=True)
        self.assertIn("我的交易复盘", page)
        self.assertIn("未复盘", page)
        self.assertIn("为什么买？", page)
        self.assertIn("回调买入", page)
        self.assertIn("标记已复盘", page)

        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={
                "reason_type": ["PULLBACK", "TECHNICAL_PATTERN"],
                "trade_reason": "回调到支撑位",
                "sell_reason": "达到目标价",
                "expected_profit_percent": "20",
                "expected_target_price": "12",
                "stop_loss_price": "9",
                "expected_holding_days": "10",
                "confidence_level": "4",
                "judgement_result": "CORRECT",
                "main_problem": "NO_OBVIOUS_PROBLEM",
                "review_note": "按计划执行",
                "next_action": "保持节奏",
                "action": "draft",
                "sort": "recent",
            },
            follow_redirects=True,
        )
        self.assertIn("复盘草稿已保存", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            review = db.execute("SELECT * FROM trade_reviews").fetchone()
            self.assertEqual(review["review_status"], "PENDING")
            self.assertEqual(review["trade_reason"], "回调到支撑位")
            self.assertAlmostEqual(review["expected_target_price"], 12)
            self.assertEqual(review["confidence_level"], 4)
            reason_types = {row["reason_type"] for row in db.execute("SELECT reason_type FROM trade_review_reason_types")}
            self.assertEqual(reason_types, {"PULLBACK", "TECHNICAL_PATTERN"})

        page = self.client.get(f"/analysis/trades/{episode_id}").get_data(as_text=True)
        self.assertIn("草稿", page)
        self.assertIn('name="reason_type" value="PULLBACK" checked', page)
        self.assertIn('value="12.0"', page)

        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={"action": "complete", "trade_reason": "x", "judgement_result": "CORRECT", "sort": "recent"},
            follow_redirects=True,
        )
        self.assertIn("复盘已完成并保存", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT review_status FROM trade_reviews").fetchone()["review_status"], "COMPLETED")
            self.assertIsNotNone(db.execute("SELECT completed_at FROM trade_reviews").fetchone()["completed_at"])

    def test_trade_review_open_episode_only_draft_and_validation(self):
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
        ])
        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={"action": "complete", "trade_reason": "买入持有", "sort": "recent"},
            follow_redirects=True,
        )
        self.assertIn("复盘草稿已保存", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT review_status FROM trade_reviews").fetchone()["review_status"], "PENDING")
            self.assertIsNone(db.execute("SELECT completed_at FROM trade_reviews").fetchone()["completed_at"])
        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={"action": "complete", "trade_reason": "买入持有", "judgement_result": "CORRECT", "sort": "recent"},
            follow_redirects=True,
        )
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT review_status FROM trade_reviews").fetchone()["review_status"], "PENDING")

        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={"action": "complete", "trade_reason": "x", "judgement_result": "INVALID", "sort": "recent"},
            follow_redirects=True,
        )
        self.assertIn("判断结果无效", response.get_data(as_text=True))

    def test_trade_review_save_requires_judgement_for_complete(self):
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 12],
        ])
        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={"action": "complete", "trade_reason": "x", "sort": "recent"},
            follow_redirects=True,
        )
        self.assertIn("标记已复盘前请选择判断结果", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_reviews").fetchone()[0], 0)

    def test_trade_review_save_is_user_isolated(self):
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 12],
        ])
        self.client.post("/users/create", data={"name": "second", "next": "/analysis/stocks"})
        response = self.client.post(
            f"/analysis/trades/{episode_id}/review",
            data={"action": "draft", "trade_reason": "x", "sort": "recent"},
        )
        self.assertEqual(response.status_code, 404)
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM trade_reviews").fetchone()[0], 0)

    def test_review_progress_and_reviews_page(self):
        first = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 12],
        ])
        second = self._import_workbook_rows([
            [20260803, "09:30:00", 100002, "国机精工", "证券买入", 200, "B2", 8],
            [20260804, "10:00:00", 100002, "国机精工", "证券卖出", -200, "S2", 9],
        ])
        self._import_workbook_rows([
            [20260805, "09:30:00", 300540, "蜀道装备", "证券买入", 100, "B3", 15],
        ])
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            progress = self.module.review_progress_data(db, user_id)
            self.assertEqual(progress["total"], 2)
            self.assertEqual(progress["completed"], 0)
            self.assertEqual(progress["pending"], 2)
            self.assertAlmostEqual(progress["rate"], 0.0)

        pending = self.client.get("/analysis/reviews?status=pending&start_date=2026-08-01&end_date=2026-08-31").get_data(as_text=True)
        self.assertIn("待复盘 2", pending)
        self.assertIn("去复盘 →", pending)
        self.assertIn("通富微电", pending)
        self.assertIn("国机精工", pending)
        self.assertNotIn("蜀道装备", pending)
        completed_page = self.client.get("/analysis/reviews?status=completed&start_date=2026-08-01&end_date=2026-08-31").get_data(as_text=True)
        self.assertIn("还没有已完成复盘的交易", completed_page)

        self.client.post(
            f"/analysis/trades/{first}/review",
            data={
                "action": "complete",
                "trade_reason": "x",
                "judgement_result": "CORRECT",
                "reason_type": ["PULLBACK", "TECHNICAL_PATTERN"],
                "sort": "recent",
            },
            follow_redirects=True,
        )
        with self.module.db_connect() as db:
            progress = self.module.review_progress_data(db, user_id)
            self.assertEqual(progress["completed"], 1)
            self.assertEqual(progress["pending"], 1)
            self.assertAlmostEqual(progress["rate"], 0.5)

        pending = self.client.get("/analysis/reviews?status=pending&start_date=2026-08-01&end_date=2026-08-31").get_data(as_text=True)
        self.assertIn("待复盘 1", pending)
        self.assertIn("国机精工", pending)
        self.assertNotIn("通富微电", pending)
        completed_page = self.client.get("/analysis/reviews?status=completed&start_date=2026-08-01&end_date=2026-08-31").get_data(as_text=True)
        self.assertIn("已复盘 1", completed_page)
        self.assertIn("通富微电", completed_page)
        self.assertIn("去复盘 →", completed_page)
        self.assertIn('id="review-modal-backdrop"', completed_page)
        self.assertIn("review-popup-trigger", completed_page)
        self.assertIn("CORRECT", completed_page)
        match = re.search(r"data-review='([^']+)'", completed_page)
        self.assertIsNotNone(match)
        review_data = json.loads(match.group(1))
        self.assertEqual(review_data["code"], "002156")
        self.assertEqual(review_data["judgement"], "CORRECT")
        self.assertEqual(review_data["trade_reason"], "x")
        self.assertEqual(review_data["reason_types"], ["回调买入", "技术形态"])

        summary = self.client.get("/analysis/summary").get_data(as_text=True)
        self.assertIn("交易复盘完成度", summary)
        self.assertIn("待复盘交易", summary)

    def test_review_date_filter(self):
        today = date.today()
        self._import_workbook_rows([
            [(today - timedelta(days=1)).strftime("%Y%m%d"), "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [today.strftime("%Y%m%d"), "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 12],
        ])
        self._import_workbook_rows([
            ["20260102", "09:30:00", 100002, "国机精工", "证券买入", 100, "B2", 8],
            ["20260103", "10:00:00", 100002, "国机精工", "证券卖出", -100, "S2", 9],
        ])

        page = self.client.get("/analysis/reviews").get_data(as_text=True)
        self.assertIn("通富微电", page)
        self.assertNotIn("国机精工", page)
        self.assertIn("交易复盘", page)

        page = self.client.get("/analysis/reviews?start_date=2026-01-01&end_date=2026-12-31").get_data(as_text=True)
        self.assertIn("国机精工", page)
        self.assertIn("通富微电", page)

        page = self.client.get("/analysis/reviews?start_date=2020-01-01&end_date=2020-12-31").get_data(as_text=True)
        self.assertNotIn("通富微电", page)
        self.assertNotIn("国机精工", page)

        start30 = (today - timedelta(days=29)).strftime("%Y-%m-%d")
        page = self.client.get(f"/analysis/reviews?start_date={start30}&end_date={today.strftime('%Y-%m-%d')}").get_data(as_text=True)
        self.assertIn("通富微电", page)
        self.assertNotIn("国机精工", page)

        page = self.client.get("/analysis/reviews?status=all&start_date=2026-01-01&end_date=2026-12-31").get_data(as_text=True)
        self.assertIn("通富微电", page)
        self.assertIn("国机精工", page)

    def test_excursion_metrics_calculation(self):
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260805, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 14],
        ])
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            for d, h, l in [("2026-08-01", 15, 9), ("2026-08-02", 16, 10), ("2026-08-03", 14, 8), ("2026-08-04", 13, 9), ("2026-08-05", 14, 10)]:
                db.execute(
                    "INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', ?, ?, ?, ?, ?, 0, 'test', '2026-08-21T00:00:00')",
                    (d, l, h, l, l),
                )
        with self.module.db_connect() as db:
            episode = [r for r in self.module.trade_episode_rows(db, user_id) if r["id"] == episode_id][0]
            metrics = self.module.calculate_excursion_metrics(db, user_id, episode)
        self.assertAlmostEqual(metrics["mfe"], 0.6, places=4)
        self.assertAlmostEqual(metrics["mae"], -0.2, places=4)
        self.assertAlmostEqual(metrics["capture_rate"], 400 / 600, places=4)
        self.assertAlmostEqual(metrics["highest_price"], 16)
        self.assertAlmostEqual(metrics["lowest_price"], 8)

        page = self.client.get(f"/analysis/trades/{episode_id}").get_data(as_text=True)
        self.assertIn("MFE 最大浮盈", page)
        self.assertIn("60.0%", page)
        self.assertIn("-20.0%", page)
        self.assertIn("利润捕获率", page)

    def test_excursion_capture_rate_edge_cases(self):
        lose = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260803, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 9],
        ])
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', '2026-08-01', 10, 12, 9, 11, 0, 'test', 'x')")
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', '2026-08-02', 11, 11, 8, 9, 0, 'test', 'x')")
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', '2026-08-03', 9, 10, 8, 9, 0, 'test', 'x')")
        with self.module.db_connect() as db:
            episode = [r for r in self.module.trade_episode_rows(db, user_id) if r["id"] == lose][0]
            metrics = self.module.calculate_excursion_metrics(db, user_id, episode)
        self.assertEqual(metrics["capture_rate"], 0.0)

        flat = self._import_workbook_rows([
            [20260801, "09:30:00", 100002, "国机精工", "证券买入", 100, "B1", 10],
            [20260803, "10:00:00", 100002, "国机精工", "证券卖出", -100, "S1", 8],
        ])
        with self.module.db_connect() as db:
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('100002', '2026-08-01', 10, 9, 8, 9, 0, 'test', 'x')")
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('100002', '2026-08-03', 9, 9, 7, 8, 0, 'test', 'x')")
        with self.module.db_connect() as db:
            episode = [r for r in self.module.trade_episode_rows(db, user_id) if r["id"] == flat][0]
            metrics = self.module.calculate_excursion_metrics(db, user_id, episode)
        self.assertIsNone(metrics["capture_rate"])

        empty_page = self.client.get(f"/analysis/trades/{lose}").get_data(as_text=True)
        self.assertIn("同步行情并计算", empty_page)

    def test_trade_diagnosis_rules(self):
        episode = {"profit": 100.0, "status": "CLOSED", "sell_price": 14.0, "buy_price": 10.0, "stock_code": "002156", "sell_date": "2026-08-05", "id": 1}
        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, None, None, episode)]
        self.assertIn("NO_PRICE", codes)

        early = {"capture_rate": 0.3, "highest_price": 16.0, "lowest_price": 8.0}
        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, None, early, episode)]
        self.assertIn("EARLY_EXIT", codes)

        good = {"capture_rate": 0.9, "highest_price": 14.2, "lowest_price": 8.0}
        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, None, good, episode)]
        self.assertIn("GOOD_EXIT", codes)

        loss_episode = {"profit": -100.0, "status": "CLOSED", "sell_price": 14.0, "buy_price": 10.0, "stock_code": "002156", "sell_date": "2026-08-05", "id": 2}
        stop_review = {"stop_loss_price": 15.0, "expected_target_price": None}
        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, stop_review, early, loss_episode)]
        self.assertIn("STOP_VIOLATION", codes)

        target_review = {"stop_loss_price": None, "expected_target_price": 13.0}
        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, target_review, early, episode)]
        self.assertIn("TARGET_MET", codes)

        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, None, early, episode)]
        self.assertIn("NO_PLAN", codes)

        with self.module.db_connect() as db:
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', '2026-08-10', 14, 15, 13, 15, 0, 'test', 'x')")
            codes = [item["code"] for item in self.module.trade_diagnosis(db, None, early, loss_episode)]
        self.assertIn("MISSED_REBOUND", codes)

        full_review = {"stop_loss_price": 12.0, "expected_target_price": 13.0}
        with self.module.db_connect() as db:
            codes = [item["code"] for item in self.module.trade_diagnosis(db, full_review, {"capture_rate": 0.8, "highest_price": 14.2, "lowest_price": 11.0}, episode)]
        self.assertNotIn("NO_PLAN", codes)
        self.assertIn("TARGET_MET", codes)

    def test_trading_system_stats(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            stats = self.module.trading_system_stats(db, user_id)
            self.assertEqual(stats["total"], 0)
            self.assertIsNone(stats["avg_win"])
            self.assertIsNone(stats["profit_factor"])
            self.assertEqual(stats["max_win_streak"], 0)

        trades = [
            (2156, "通富微电", 10, 12, 20260802),
            (100002, "国机精工", 10, 9, 20260803),
            (300540, "蜀道装备", 10, 13, 20260804),
            (301697, "贝特利", 10, 14, 20260805),
            (600309, "万华化学", 10, 8, 20260806),
        ]
        for code, name, buy, sell, sell_day in trades:
            self._import_workbook_rows([
                [20260801, "09:30:00", code, name, "证券买入", 100, "B", buy],
                [sell_day, "10:00:00", code, name, "证券卖出", -100, "S", sell],
            ])
        with self.module.db_connect() as db:
            stats = self.module.trading_system_stats(db, user_id)
        self.assertEqual(stats["total"], 5)
        self.assertAlmostEqual(stats["avg_win"], 300.0, places=4)
        self.assertAlmostEqual(stats["avg_loss"], -150.0, places=4)
        self.assertAlmostEqual(stats["profit_factor"], 2.0, places=4)
        self.assertEqual(stats["max_win_streak"], 2)
        self.assertEqual(stats["max_loss_streak"], 1)

        summary = self.client.get("/analysis/summary").get_data(as_text=True)
        self.assertIn("我的交易系统", summary)
        self.assertIn("盈亏比", summary)
        self.assertIn("最大连续盈利", summary)
        self.assertIn("最大连续亏损", summary)

    def test_strategy_type_stats(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            self.assertEqual(self.module.strategy_type_stats(db, user_id), [])

        win = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 13],
        ])
        lose = self._import_workbook_rows([
            [20260803, "09:30:00", 100002, "国机精工", "证券买入", 100, "B2", 10],
            [20260804, "10:00:00", 100002, "国机精工", "证券卖出", -100, "S2", 9],
        ])
        self.client.post(f"/analysis/trades/{win}/review", data={"action": "complete", "trade_reason": "x", "judgement_result": "CORRECT", "reason_type": ["PULLBACK"], "sort": "recent"})
        self.client.post(f"/analysis/trades/{lose}/review", data={"action": "complete", "trade_reason": "x", "judgement_result": "WRONG", "reason_type": ["PULLBACK", "TECHNICAL_PATTERN"], "sort": "recent"})

        with self.module.db_connect() as db:
            stats = self.module.strategy_type_stats(db, user_id)
        by_type = {row["reason_type"]: row for row in stats}
        self.assertEqual(by_type["PULLBACK"]["trades"], 2)
        self.assertEqual(by_type["PULLBACK"]["wins"], 1)
        self.assertAlmostEqual(by_type["PULLBACK"]["total_profit"], 200.0)
        self.assertAlmostEqual(by_type["PULLBACK"]["win_rate"], 0.5)
        self.assertEqual(by_type["TECHNICAL_PATTERN"]["trades"], 1)
        self.assertAlmostEqual(by_type["TECHNICAL_PATTERN"]["total_profit"], -100.0)

        page = self.client.get("/analysis/strategy").get_data(as_text=True)
        self.assertIn("我的策略类型统计", page)
        self.assertIn("回调买入", page)
        self.assertIn("技术形态", page)
        self.assertIn("累计盈亏", page)

    def test_mistake_stats(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            self.assertEqual(self.module.mistake_stats(db, user_id)["items"], [])

        ep1 = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 13],
        ])
        ep2 = self._import_workbook_rows([
            [20260803, "09:30:00", 100002, "国机精工", "证券买入", 100, "B2", 10],
            [20260804, "10:00:00", 100002, "国机精工", "证券卖出", -100, "S2", 9],
        ])
        ep3 = self._import_workbook_rows([
            [20260805, "09:30:00", 300540, "蜀道装备", "证券买入", 100, "B3", 10],
            [20260806, "10:00:00", 300540, "蜀道装备", "证券卖出", -100, "S3", 8],
        ])
        self.client.post(f"/analysis/trades/{ep1}/review", data={"action": "complete", "trade_reason": "x", "judgement_result": "PARTIAL", "main_problem": "CHASE", "sort": "recent"})
        self.client.post(f"/analysis/trades/{ep2}/review", data={"action": "complete", "trade_reason": "x", "judgement_result": "WRONG", "main_problem": "CHASE", "sort": "recent"})
        self.client.post(f"/analysis/trades/{ep3}/review", data={"action": "complete", "trade_reason": "x", "judgement_result": "WRONG", "main_problem": "EXIT_TIMING", "sort": "recent"})

        with self.module.db_connect() as db:
            result = self.module.mistake_stats(db, user_id)
        self.assertEqual(result["total_reviewed"], 3)
        by_type = {row["main_problem"]: row for row in result["items"]}
        self.assertEqual(by_type["CHASE"]["count"], 2)
        self.assertAlmostEqual(by_type["CHASE"]["total_profit"], 200.0)
        self.assertAlmostEqual(by_type["CHASE"]["rate"], 2 / 3, places=4)
        self.assertEqual(by_type["EXIT_TIMING"]["count"], 1)
        self.assertAlmostEqual(by_type["EXIT_TIMING"]["total_profit"], -200.0)
        self.assertEqual(result["items"][0]["main_problem"], "CHASE")

        page = self.client.get("/analysis/discipline").get_data(as_text=True)
        self.assertIn("我的交易错误排行榜", page)
        self.assertIn("追涨", page)
        self.assertIn("卖点错误", page)

    def test_trade_episode_detail_is_user_isolated(self):
        self.client.post("/admin/preview", data={"statement": (self.make_statement(), "first.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{job['id']}")
        with self.module.db_connect() as db:
            episode_id = db.execute("SELECT id FROM trade_episodes").fetchone()["id"]
        self.client.post("/users/create", data={"name": "second", "next": "/analysis/stocks"})
        self.assertEqual(self.client.get(f"/analysis/trades/{episode_id}").status_code, 404)

    def test_paste_manual_executions(self):
        trades = "\n".join([
            "成交时间\t证券代码\t证券名称\t操作\t成交数量\t成交均价\t成交金额\t合同编号\t成交编号\t委托时间",
            "14:49:00\t002156\t通富微电\t证券卖出\t100\t63.810\t6381.000\t23921\t0105000070904721\t14:48:59",
            "09:35:29\t300540\t蜀道装备\t证券买入\t200\t29.890\t5978.000\t4322\t0101000007141777\t09:35:29",
        ])
        response = self.client.post(
            "/admin/manual/paste",
            data={"trade_date": "2026-08-21", "trades": trades},
            follow_redirects=True,
        )
        self.assertIn("新增 2 条", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions WHERE source = 'BROKER_TODAY'").fetchone()[0], 2)
            sell = db.execute("SELECT * FROM executions WHERE stock_code = '002156'").fetchone()
            self.assertEqual(sell["action"], "SELL")
            self.assertEqual(sell["deal_id"], "0105000070904721")
            buy = db.execute("SELECT * FROM executions WHERE stock_code = '300540'").fetchone()
            self.assertAlmostEqual(buy["deal_price"], 29.89)

        response = self.client.post(
            "/admin/manual/paste",
            data={"trade_date": "2026-08-21", "trades": trades},
            follow_redirects=True,
        )
        self.assertIn("跳过重复 2 条", response.get_data(as_text=True))

    def test_statements_pagination_and_order(self):
        self.import_statement(days=25)
        page1 = self.client.get("/admin/statements")
        self.assertEqual(page1.status_code, 200)
        page1_text = page1.get_data(as_text=True)
        self.assertIn("第 1 / 2 页 · 共 25 条", page1_text)
        with self.module.db_connect() as db:
            first = db.execute("SELECT trade_date FROM executions ORDER BY trade_date DESC, trade_time DESC, id DESC LIMIT 1").fetchone()
        self.assertIn(first["trade_date"], page1_text)
        page2 = self.client.get("/admin/statements?page=2")
        self.assertEqual(page2.status_code, 200)
        page2_text = page2.get_data(as_text=True)
        self.assertIn("第 2 / 2 页 · 共 25 条", page2_text)
        with self.module.db_connect() as db:
            last = db.execute("SELECT trade_date FROM executions ORDER BY trade_date DESC, trade_time DESC, id DESC LIMIT 1 OFFSET 24").fetchone()
        self.assertIn(last["trade_date"], page2_text)

    def test_label_placement_no_overlap(self):
        labels = []
        for index in range(8):
            self.module.place_label(labels, 120.0 + index * 6, 100.0, f"+{index + 1},000.00", "#000")
        ys = [label[1] for label in labels]
        self.assertEqual(len(set(ys)), len(ys), "密集标注应上下错开不重叠")

    def test_stock_detail_single_sell_profit(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格", "成交金额", "佣金", "印花税", "其他杂费", "其他费", "过户费", "交易市场"])
        sheet.append([20260801, "09:31:00", 2156, "通富微电", "证券买入", 100, "B1", 10, 1000, 1, 0, 0, 0, 0, "深圳A股"])
        sheet.append([20260803, "13:00:00", 2156, "通富微电", "证券卖出", -50, "S1", 14, 700, 1, 2, 0, 0, 0, "深圳A股"])
        sheet.append([20260803, "14:00:00", 2156, "通富微电", "证券卖出", -50, "S2", 11, 550, 1, 2, 0, 0, 0, "深圳A股"])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "statement.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
        self.client.post(f"/admin/import/{job['id']}")

        detail = self.client.get("/analysis/stocks?code=002156")
        self.assertEqual(detail.status_code, 200)
        text = detail.get_data(as_text=True)
        self.assertIn("总交易笔数", text)
        self.assertIn("+243.00", text)
        self.assertIn("24.28%", text)
        self.assertNotIn("chart-wrap", text)
        self.assertNotIn("<svg", text)
        self.assertNotIn("查看复盘", text)
        self.assertNotIn("查看 FIFO", text)
        self.assertIn("stock-kline-trigger", text)
        kline = self.client.get("/analysis/stocks/kline/002156")
        self.assertEqual(kline.status_code, 200)
        self.assertEqual(kline.get_json()["stock_code"], "002156")
        self.assertEqual([point["action"] for point in kline.get_json()["bs_points"]], ["BUY", "SELL", "SELL"])

        with self.module.db_connect() as db:
            db.execute("DELETE FROM daily_prices WHERE stock_code = '002156'")
        with patch.object(self.module, "sync_history_codes", return_value=(0, ["002156：测试同步失败"])) as sync:
            empty_kline = self.client.get("/analysis/stocks/kline/002156")
        sync.assert_called_once_with(["002156"], full=True)
        self.assertTrue(empty_kline.get_json()["sync_attempted"])
        self.assertIn("测试同步失败", empty_kline.get_json()["sync_error"])

    def test_submenu_and_import_button(self):
        page = self.client.get("/admin/statements")
        text = page.get_data(as_text=True)
        self.assertIn('id="submenu-admin"', text)
        self.assertIn("交割单", text)
        self.assertIn("导入交割单", text)
        self.assertIn('class="database-backup"', text)
        self.assertIn("下载包含本机全部用户及所有业务数据的完整 SQLite 数据库", text)
        self.assertIn('href="/admin/database/export"', text)
        self.assertIn("重新计算 FIFO", text)
        self.assertIn("共 0 条成交记录", text)

        import_page = self.client.get("/admin/import").get_data(as_text=True)
        self.assertIn("选择文件、解析预览、确认写入", import_page)
        self.assertIn('id="file-title"', import_page)
        self.assertIn('id="preview-button"', import_page)
        self.assertIn("正在解析...", import_page)

        account_page = self.client.get("/admin/account").get_data(as_text=True)
        self.assertIn("账户资金", account_page)
        self.assertIn("记录当前资金", account_page)
        self.assertIn('class="database-backup"', account_page)
        self.assertIn('href="/admin/database/export"', account_page)

    def test_export_database_downloads_complete_valid_snapshot(self):
        with self.module.db_connect() as db:
            other_user_id = db.execute(
                "INSERT INTO users (name, created_at) VALUES (?, ?)",
                ("备份测试用户", datetime.now().isoformat(timespec="seconds")),
            ).lastrowid
            db.execute(
                "INSERT INTO account_snapshots (user_id, snapshot_date, total_assets, available_cash, note, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (other_user_id, "2026-08-23", 123456.78, 23456.78, "完整备份验证", "2026-08-23T12:00:00", "2026-08-23T12:00:00"),
            )

        response = self.client.get("/admin/database/export")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/vnd.sqlite3")
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        self.assertRegex(
            response.headers["Content-Disposition"],
            r'attachment; filename=stocknotes-backup-\d{8}-\d{6}\.db',
        )
        self.assertTrue(response.data.startswith(b"SQLite format 3\x00"))

        with tempfile.TemporaryDirectory() as directory:
            backup_path = os.path.join(directory, "downloaded.db")
            with open(backup_path, "wb") as backup_file:
                backup_file.write(response.data)
            backup = sqlite3.connect(backup_path)
            try:
                self.assertEqual(backup.execute("PRAGMA user_version").fetchone()[0], self.module.SCHEMA_VERSION)
                self.assertEqual(backup.execute("PRAGMA foreign_key_check").fetchall(), [])
                users = {row[0] for row in backup.execute("SELECT name FROM users")}
                self.assertEqual(users, {"yutaoGS", "备份测试用户"})
                snapshot = backup.execute(
                    "SELECT total_assets, note FROM account_snapshots WHERE user_id = ?",
                    (other_user_id,),
                ).fetchone()
                self.assertEqual(snapshot, (123456.78, "完整备份验证"))
            finally:
                backup.close()

    def test_daily_volume_schema_migration_normalizes_lots_once(self):
        with self.module.db_connect() as db:
            rows = (
                ("900001", "tencent-realtime", 1234),
                ("900002", "akshare-eastmoney", 2345),
                ("900003", "akshare-tencent", 345600),
                ("900004", "akshare-sina", 456700),
            )
            for stock_code, source, volume in rows:
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES (?, '2026-08-24', 10, 11, 9, 10, ?, ?, '2026-08-24T15:00:00')""",
                    (stock_code, volume, source),
                )
            db.execute("PRAGMA user_version = 14")

        self.module.init_db()
        self.module.init_db()

        with self.module.db_connect() as db:
            volumes = dict(db.execute(
                "SELECT stock_code, volume FROM daily_prices WHERE stock_code LIKE '90000%'"
            ).fetchall())
            version = db.execute("PRAGMA user_version").fetchone()[0]
        self.assertEqual(version, self.module.SCHEMA_VERSION)
        self.assertEqual(volumes, {
            "900001": 123400, "900002": 234500,
            "900003": 345600, "900004": 456700,
        })

    def test_current_positions_schema_migration_allows_optional_fields(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute("DROP TABLE current_positions")
            db.execute(
                """CREATE TABLE current_positions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    stock_code TEXT NOT NULL,
                    stock_name TEXT NOT NULL,
                    quantity REAL NOT NULL CHECK(quantity > 0),
                    avg_cost REAL NOT NULL CHECK(avg_cost > 0),
                    first_buy_date TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(user_id, stock_code)
                )"""
            )
            db.execute(
                """INSERT INTO current_positions
                (user_id, stock_code, stock_name, quantity, avg_cost, first_buy_date, created_at, updated_at)
                VALUES (?, '002156', '通富微电', 100, 10, '2026-08-01', '2026-08-01', '2026-08-01')""",
                (user_id,),
            )
            db.execute("PRAGMA user_version = 19")

        self.module.init_db()
        self.module.init_db()

        with self.module.db_connect() as db:
            columns = {row["name"]: row for row in db.execute("PRAGMA table_info(current_positions)")}
            row = db.execute("SELECT * FROM current_positions WHERE stock_code = '002156'").fetchone()
            db.execute(
                """INSERT INTO current_positions
                (user_id, stock_code, stock_name, quantity, avg_cost, first_buy_date, created_at, updated_at)
                VALUES (?, '000001', '平安银行', NULL, NULL, NULL, '2026-08-01', '2026-08-01')""",
                (user_id,),
            )
            index_names = {item["name"] for item in db.execute("PRAGMA index_list(current_positions)")}
            foreign_key_errors = db.execute("PRAGMA foreign_key_check").fetchall()
        self.assertEqual(row["quantity"], 100)
        self.assertEqual(row["avg_cost"], 10)
        self.assertEqual(row["first_buy_date"], "2026-08-01")
        self.assertFalse(columns["quantity"]["notnull"])
        self.assertFalse(columns["avg_cost"]["notnull"])
        self.assertFalse(columns["first_buy_date"]["notnull"])
        self.assertIn("idx_current_positions_user", index_names)
        self.assertEqual(foreign_key_errors, [])

    def test_split_menus(self):
        root = self.client.get("/")
        self.assertEqual(root.status_code, 302)
        self.assertIn("/analysis/portfolio", root.headers["Location"])

        stocks = self.client.get("/analysis/stocks")
        self.assertEqual(stocks.status_code, 200)
        stocks_text = stocks.get_data(as_text=True)
        self.assertIn("个股分析", stocks_text)
        self.assertIn('id="submenu-review"', stocks_text)
        self.assertIn("交易记录", stocks_text)
        self.assertIn("情绪纪律", stocks_text)
        self.assertIn("还没有已平仓交易", stocks_text)

        summary = self.client.get("/analysis/summary")
        self.assertEqual(summary.status_code, 200)
        summary_text = summary.get_data(as_text=True)
        self.assertIn("汇总", summary_text)
        self.assertIn("月度盈亏", summary_text)
        self.assertIn("交易复盘", summary_text)

        portfolio = self.client.get("/analysis/portfolio")
        self.assertEqual(portfolio.status_code, 200)
        portfolio_text = portfolio.get_data(as_text=True)
        self.assertIn("持仓分析", portfolio_text)
        self.assertIn("不只看“持有什么”", portfolio_text)
        self.assertIn("当前没有持仓", portfolio_text)
        self.assertNotIn("仓位集中度", portfolio_text)
        self.assertNotIn("持仓效率", portfolio_text)

        watchlist = self.client.get("/analysis/watchlist")
        self.assertEqual(watchlist.status_code, 200)
        watchlist_text = watchlist.get_data(as_text=True)
        self.assertIn("重点观察", watchlist_text)
        self.assertIn('action="/analysis/watchlist/sync-prices"', watchlist_text)
        self.assertIn('action="/analysis/watchlist/sync-history"', watchlist_text)
        self.assertIn('id="watchlist-add-trigger"', watchlist_text)
        self.assertIn('id="watchlist-add-backdrop"', watchlist_text)
        self.assertNotIn('class="panel watchlist-add"', watchlist_text)

        reviews = self.client.get("/analysis/reviews")
        self.assertEqual(reviews.status_code, 200)
        reviews_text = reviews.get_data(as_text=True)
        self.assertIn("交易复盘", reviews_text)
        self.assertIn("时间范围", reviews_text)
        self.assertIn("7天", reviews_text)
        self.assertIn("15天", reviews_text)
        self.assertIn("30天", reviews_text)
        self.assertIn("查询", reviews_text)

        strategy = self.client.get("/analysis/strategy")
        self.assertEqual(strategy.status_code, 200)
        strategy_text = strategy.get_data(as_text=True)
        self.assertIn("策略验证", strategy_text)
        self.assertIn("没有符合当前条件的历史交易", strategy_text)

    def test_portfolio_analysis_uses_positions_and_latest_close(self):
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
        ])
        self._add_current_position(quantity="100", avg_cost="10")

        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertIn("投入资金", page)
        self.assertIn("1,000.00", page)
        self.assertIn("暂无行情数据", page)
        self.assertIn('/analysis/portfolio/kline/002156', page)
        self.assertIn('id="portfolio-kline-backdrop"', page)
        self.assertIn("开盘：", page)
        self.assertIn("收盘：", page)
        self.assertIn("涨幅：", page)
        self.assertIn("MA5：", page)
        self.assertIn("MA10：", page)
        self.assertIn("MA20：", page)
        self.assertNotIn("持仓成本口径", page)

        stock_page = self.client.get("/analysis/stocks?code=002156").get_data(as_text=True)
        self.assertIn("当前持仓", stock_page)

        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('002156', '2026-08-20', 10, 10.5, 9.8, 10, 10000, 'test', '2026-08-20T15:01:00')"""
            )
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('002156', '2026-08-21', 11, 12.5, 10.8, 12, 10000, 'test', '2026-08-21T15:01:00')"""
            )

        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertIn("总市值", page)
        self.assertIn("1,200.00", page)
        self.assertIn("+200.00", page)
        self.assertIn("20.0%", page)
        self.assertIn("涨跌幅", page)
        self.assertIn("+20.00%", page)
        self.assertIn("2026-08-21 · 15:01:00", page)
        self.assertNotIn("市值口径", page)
        self.assertIn("行情日期 2026-08-21", page)

        quotes = self.client.get("/analysis/portfolio/quotes").get_json()
        self.assertAlmostEqual(quotes["positions"][0]["change_rate"], 0.2)
        self.assertEqual(quotes["summary"]["price_fetched_at"], "2026-08-21T15:01:00")
        self.assertEqual(self.client.get("/analysis/portfolio/quotes").headers["Cache-Control"], "no-store")

        kline_response = self.client.get("/analysis/portfolio/kline/002156")
        self.assertEqual(kline_response.status_code, 200)
        kline = kline_response.get_json()
        self.assertEqual(kline["stock_code"], "002156")
        self.assertEqual(kline["stock_name"], "通富微电")
        self.assertEqual(kline["avg_cost"], 10)
        self.assertEqual(kline["data"][-1]["date"], "2026-08-21")
        self.assertEqual(kline["data"][-1]["close"], 12)
        self.assertEqual([point["action"] for point in kline["bs_points"]], ["BUY"])
        self.assertEqual(kline["bs_points"][0]["price"], 10)

        response = self.client.post(
            "/admin/account",
            data={
                "snapshot_date": "2026-08-21", "total_assets": "20000",
                "available_cash": "8000", "note": "收盘快照",
            },
            follow_redirects=True,
        )
        self.assertIn("已保存 2026-08-21 的账户资金快照", response.get_data(as_text=True))
        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertNotIn("20,000.00", page)
        self.assertNotIn("股票总仓位", page)
        self.assertNotIn("非股票资产", page)

    def test_account_snapshot_upsert_validation_and_user_isolation(self):
        response = self.client.post(
            "/admin/account",
            data={"snapshot_date": "2026-08-21", "total_assets": "10000", "available_cash": "12000"},
            follow_redirects=True,
        )
        self.assertIn("可用资金不能大于账户总资产", response.get_data(as_text=True))

        self.client.post(
            "/admin/account",
            data={"snapshot_date": "2026-08-21", "total_assets": "10000", "available_cash": "2000"},
        )
        self.client.post(
            "/admin/account",
            data={"snapshot_date": "2026-08-21", "total_assets": "12000", "available_cash": "3000"},
        )
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM account_snapshots").fetchone()[0], 1)
            self.assertEqual(db.execute("SELECT total_assets FROM account_snapshots").fetchone()[0], 12000)

        self.client.post("/users/create", data={"name": "account-second", "next": "/admin/account"})
        second_page = self.client.get("/admin/account").get_data(as_text=True)
        self.assertIn("还没有账户资金快照", second_page)
        self.assertEqual(self.client.get("/analysis/portfolio/kline/002156").status_code, 404)

    def test_portfolio_sync_with_empty_positions(self):
        response = self.client.post("/analysis/portfolio/sync-prices", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("当前没有持仓，无需同步行情", response.get_data(as_text=True))

        self._import_workbook_rows([
            [20260820, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
        ])
        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertNotIn("持仓时间", page)
        self.assertIn("当前没有持仓。在下方维护独立持仓后，系统会据此计算市值与浮动盈亏。", page)

        self._add_current_position(quantity="100", avg_cost="10")
        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertIn("当前持仓", page)
        self.assertNotIn("持仓时间", page)

    def test_portfolio_realtime_sync_refreshes_indexes_in_background(self):
        self._add_current_position()
        with patch.object(self.module, "sync_realtime_codes", return_value=(1, [])) as sync, \
             patch.object(self.module, "refresh_market_indexes_in_background") as refresh, \
             patch.object(self.module, "sync_market_indexes") as blocking_sync:
            response = self.client.post(
                "/analysis/portfolio/sync-realtime",
                headers={"X-Requested-With": "XMLHttpRequest"},
            )

        self.assertEqual(response.get_json(), {"message": "已同步 1 条实时行情。", "category": "success"})
        sync.assert_called_once_with(["002156"])
        refresh.assert_called_once_with()
        blocking_sync.assert_not_called()

    def test_realtime_pages_use_short_polling_without_event_source(self):
        portfolio = self.client.get("/analysis/portfolio").get_data(as_text=True)
        watchlist = self.client.get("/analysis/watchlist").get_data(as_text=True)

        self.assertNotIn("new EventSource", portfolio)
        self.assertNotIn("new EventSource", watchlist)
        self.assertIn("/analysis/portfolio/quotes", portfolio)
        self.assertIn("/analysis/watchlist/quotes", watchlist)
        self.assertIn("window.setInterval(refresh, 5000)", portfolio)
        self.assertIn("window.setInterval(refresh,5000)", watchlist)

    def test_current_position_crud_and_validation(self):
        response = self._add_current_position()
        self.assertIn("已保存当前持仓", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            row = db.execute("SELECT * FROM current_positions").fetchone()
            self.assertEqual(row["stock_code"], "002156")
            self.assertEqual(row["stock_name"], "通富微电")
            self.assertAlmostEqual(row["quantity"], 100)
            self.assertAlmostEqual(row["avg_cost"], 10)
            self.assertEqual(row["first_buy_date"], "2026-08-01")

        response = self._add_current_position(stock_code="002156", quantity="200", avg_cost="12", first_buy_date="2026-08-05")
        self.assertIn("已保存当前持仓", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions").fetchone()[0], 1)
            row = db.execute("SELECT * FROM current_positions").fetchone()
            self.assertAlmostEqual(row["quantity"], 200)
            self.assertAlmostEqual(row["avg_cost"], 12)
            self.assertEqual(row["first_buy_date"], "2026-08-05")

        response = self.client.post(
            "/analysis/portfolio/position/delete",
            data={"stock_code": "002156"},
            follow_redirects=True,
        )
        self.assertIn("已删除当前持仓", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions").fetchone()[0], 0)

        invalid_cases = [
            {"stock_code": "", "stock_name": "通富微电", "quantity": "100", "avg_cost": "10", "first_buy_date": "2026-08-01"},
            {"stock_code": "2156", "stock_name": "", "quantity": "100", "avg_cost": "10", "first_buy_date": "2026-08-01"},
            {"stock_code": "2156", "stock_name": "通富微电", "quantity": "0", "avg_cost": "10", "first_buy_date": "2026-08-01"},
            {"stock_code": "2156", "stock_name": "通富微电", "quantity": "100", "avg_cost": "-1", "first_buy_date": "2026-08-01"},
            {"stock_code": "2156", "stock_name": "通富微电", "quantity": "100", "avg_cost": "10", "first_buy_date": "not-a-date"},
        ]
        for payload in invalid_cases:
            response = self.client.post("/analysis/portfolio/position", data=payload, follow_redirects=True)
            self.assertIn("保存持仓失败", response.get_data(as_text=True))
            with self.module.db_connect() as db:
                self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions").fetchone()[0], 0)

    def test_incomplete_current_position_is_saved_but_excluded_from_calculations(self):
        response = self._add_current_position(quantity="", avg_cost="", first_buy_date="")
        page = response.get_data(as_text=True)
        self.assertIn("已保存当前持仓", page)
        self.assertIn("1 只资料不完整，不参与汇总", page)
        self.assertIn("暂无完整持仓", page)
        self.assertIn('data-quantity=""', page)
        self.assertIn('data-avg-cost=""', page)
        self.assertIn('data-date=""', page)

        with self.module.db_connect() as db:
            row = db.execute("SELECT * FROM current_positions").fetchone()
            self.assertIsNone(row["quantity"])
            self.assertIsNone(row["avg_cost"])
            self.assertIsNone(row["first_buy_date"])
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('002156', '2026-08-21', 11, 12.5, 10.8, 12, 10000, 'test', '2026-08-21T15:01:00')"""
            )

        quotes = self.client.get("/analysis/portfolio/quotes").get_json()
        self.assertEqual(quotes["summary"]["count"], 1)
        self.assertEqual(quotes["summary"]["calculated_count"], 0)
        self.assertEqual(quotes["summary"]["incomplete_count"], 1)
        self.assertEqual(quotes["summary"]["total_cost"], 0)
        self.assertIsNone(quotes["summary"]["total_market_value"])
        self.assertIsNone(quotes["positions"][0]["unrealized_profit"])

        stock_page = self.client.get("/analysis/stocks?code=002156").get_data(as_text=True)
        self.assertIn("当前持仓", stock_page)
        self.assertIn("通富微电 · 002156", stock_page)
        kline = self.client.get("/analysis/portfolio/kline/002156").get_json()
        self.assertIsNone(kline["avg_cost"])

        self._add_current_position(quantity="100", avg_cost="10", first_buy_date="2026-08-01")
        with self.module.db_connect() as db:
            row = db.execute("SELECT quantity, avg_cost, first_buy_date FROM current_positions").fetchone()
            self.assertEqual(tuple(row), (100, 10, "2026-08-01"))
        self._add_current_position(quantity="", avg_cost="", first_buy_date="")
        with self.module.db_connect() as db:
            row = db.execute("SELECT quantity, avg_cost, first_buy_date FROM current_positions").fetchone()
            self.assertEqual(tuple(row), (None, None, None))

    def test_incomplete_current_position_sync_uses_default_history_window(self):
        self._add_current_position(quantity="", avg_cost="", first_buy_date="")
        expected_start = (date.today() - timedelta(days=100)).isoformat()
        with patch.object(self.module, "sync_daily_prices", return_value=5) as sync:
            response = self.client.post("/analysis/portfolio/sync-prices", follow_redirects=True)
        self.assertIn("已为 1 只持仓同步 5 条日线行情", response.get_data(as_text=True))
        sync.assert_called_once_with(
            unittest.mock.ANY, "002156", expected_start, date.today().isoformat(),
        )

    def test_current_position_user_isolation(self):
        self._add_current_position()
        self.client.post("/users/create", data={"name": "portfolio-second", "next": "/analysis/portfolio"})
        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertIn("当前没有持仓", page)
        with self.module.db_connect() as db:
            first_user = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_user = db.execute("SELECT id FROM users WHERE name = 'portfolio-second'").fetchone()["id"]
            self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions WHERE user_id = ?", (first_user,)).fetchone()[0], 1)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions WHERE user_id = ?", (second_user,)).fetchone()[0], 0)
        self.assertEqual(self.client.get("/analysis/portfolio/kline/002156").status_code, 404)

    def test_statement_import_and_rebuild_do_not_touch_current_positions(self):
        self._add_current_position(quantity="100", avg_cost="10")
        episode_id = self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260802, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 12],
        ])
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions").fetchone()[0], 1)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM fifo_matches").fetchone()[0], 1)

        self.client.post("/admin/rebuild")
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM current_positions").fetchone()[0], 1)
            row = db.execute("SELECT * FROM current_positions").fetchone()
            self.assertAlmostEqual(row["quantity"], 100)
            self.assertAlmostEqual(row["avg_cost"], 10)

        page = self.client.get("/analysis/portfolio").get_data(as_text=True)
        self.assertIn("1,000.00", page)

    def test_watchlist_held_status_uses_current_positions(self):
        self.client.post("/analysis/watchlist", data={"stock_code": "2156", "stock_name": "通富微电", "priority": "3", "note": ""})
        page = self.client.get("/analysis/watchlist").get_data(as_text=True)
        self.assertIn("仅观察", page)
        self.assertNotIn("当前持仓", page)

        self._add_current_position()
        page = self.client.get("/analysis/watchlist").get_data(as_text=True)
        self.assertIn("当前持仓", page)

    def test_watchlist_add_update_change_and_user_isolation(self):
        with patch.object(self.module, "sync_history_codes", return_value=(167, [])) as sync:
            response = self.client.post("/analysis/watchlist", data={"stock_code": "2156", "stock_name": "通富微电", "priority": "3", "note": "等待突破"}, follow_redirects=True)
        sync.assert_called_once_with(["002156"], full=True)
        self.assertIn("已加入重点观察：通富微电", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', '2026-08-20', 10, 10, 9, 10, 100, 'test', '2026-08-21T15:00:00')")
            db.execute("INSERT INTO daily_prices (stock_code, trade_date, open, high, low, close, volume, source, fetched_at) VALUES ('002156', '2026-08-21', 10, 11, 10, 11, 100, 'test', '2026-08-22T15:00:00')")
        page = self.client.get("/analysis/watchlist").get_data(as_text=True)
        self.assertIn("+10.0%", page)
        self.assertIn("等待突破", page)
        self.assertIn("2026-08-21", page)
        quotes = self.client.get("/analysis/watchlist/quotes").get_json()
        self.assertAlmostEqual(quotes["items"][0]["change_rate"], 0.1)
        self.assertEqual(quotes["items"][0]["price_fetched_at"], "2026-08-22T15:00:00")
        self.assertEqual(self.client.get("/analysis/watchlist/quotes").headers["Cache-Control"], "no-store")
        self.assertIn('href="/analysis/watchlist?q=&amp;priority=all&amp;sort=change&amp;direction=desc"', page)
        self.assertNotIn("涨幅排名", page)
        self.assertIn('href="/analysis/watchlist"', page)
        self.assertIn("重置", page)
        self.assertIn("/analysis/watchlist/kline/002156", page)
        kline = self.client.get("/analysis/watchlist/kline/002156")
        self.assertEqual(kline.status_code, 200)
        self.assertEqual(kline.get_json()["data"][-1]["close"], 11)
        self.assertEqual(kline.get_json()["bs_points"], [])
        response = self.client.post("/analysis/watchlist", data={"stock_code": "002156", "stock_name": "通富微电", "priority": "1", "note": "更新理由"}, follow_redirects=True)
        self.assertIn("已更新重点观察：通富微电", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            item = db.execute("SELECT * FROM watchlist_stocks").fetchone()
            self.assertEqual(item["priority"], 1)
            self.assertEqual(item["note"], "更新理由")
        page = self.client.get("/analysis/watchlist").get_data(as_text=True)
        self.assertIn("watch-edit-trigger", page)
        self.assertIn('data-priority="1"', page)
        self.client.post("/users/create", data={"name": "watch-second", "next": "/analysis/watchlist"})
        self.assertEqual(self.client.get("/analysis/watchlist").get_data(as_text=True).count("通富微电"), 0)
        self.assertEqual(self.client.get("/analysis/watchlist/kline/002156").status_code, 404)

    def test_watchlist_sync_fetches_recent_two_years(self):
        with patch.object(self.module, "sync_history_codes", return_value=(0, [])):
            self.client.post(
                "/analysis/watchlist",
                data={"stock_code": "2156", "stock_name": "通富微电", "priority": "2"},
            )
        today = date.today()
        expected_start = (today - timedelta(days=730)).isoformat()
        latest_date = today - timedelta(days=1)

        with patch.object(self.module, "fetch_daily_prices", return_value=(None, None)) as fetch:
            response = self.client.post("/analysis/watchlist/sync-history")
            self.assertEqual(response.status_code, 302)
            fetch.assert_called_once_with("002156", expected_start, today.isoformat())

        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, close, source, fetched_at)
                VALUES (?, ?, 12, 'test', '2026-08-21T15:00:00')""",
                ("002156", latest_date.isoformat()),
            )

        with patch.object(self.module, "fetch_daily_prices", return_value=(None, None)) as fetch:
            response = self.client.post("/analysis/watchlist/sync-history")
            self.assertEqual(response.status_code, 302)
            fetch.assert_called_once_with("002156", expected_start, today.isoformat())

    def test_watchlist_sync_fetches_stocks_in_parallel(self):
        with patch.object(self.module, "sync_history_codes", return_value=(0, [])):
            for stock_code in ("000001", "000002", "000003", "000004"):
                self.client.post(
                    "/analysis/watchlist",
                    data={"stock_code": stock_code, "stock_name": stock_code, "priority": "2"},
                )

        lock = threading.Lock()
        active = 0
        max_active = 0

        def fetch(*args):
            nonlocal active, max_active
            with lock:
                active += 1
                max_active = max(max_active, active)
            time.sleep(0.05)
            with lock:
                active -= 1
            return None, None

        with patch.object(self.module, "fetch_daily_prices", side_effect=fetch) as mocked_fetch:
            response = self.client.post("/analysis/watchlist/sync-history")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(mocked_fetch.call_count, 4)
        self.assertGreater(max_active, 1)

    def test_watchlist_sync_supports_ajax_response(self):
        with patch.object(self.module, "sync_history_codes", return_value=(0, [])):
            self.client.post(
                "/analysis/watchlist",
                data={"stock_code": "002156", "stock_name": "通富微电", "priority": "2"},
            )
        with patch.object(self.module, "sync_realtime_codes", return_value=(1, [])):
            response = self.client.post(
                "/analysis/watchlist/sync-prices",
                headers={"X-Requested-With": "XMLHttpRequest"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"message": "已同步 1 条实时行情。", "category": "success"})

    def test_watchlist_sync_prefers_batch_realtime_prices(self):
        quote = {
            "002156": {
                "trade_date": "2026-08-21", "open": 10.0, "high": 12.0, "low": 9.5,
                "close": 11.5, "volume": 1000.0, "amount": 11500.0,
                "fetched_at": "2026-08-21T15:00:00",
            }
        }
        with patch.object(self.module, "fetch_realtime_prices", return_value=quote) as realtime, \
             patch.object(self.module, "fetch_daily_prices") as history:
            synced, failed = self.module.sync_realtime_codes(["002156"])

        realtime.assert_called_once_with(["002156"])
        history.assert_not_called()
        self.assertEqual((synced, failed), (1, []))
        with self.module.db_connect() as db:
            price = db.execute("SELECT * FROM daily_prices WHERE stock_code = '002156'").fetchone()
        self.assertEqual(price["close"], 11.5)
        self.assertEqual(price["volume"], 1000.0)
        self.assertEqual(price["source"], "tencent-realtime")

    def _intraday_rebound_sample(self):
        prices = [
            ("13:50", 0.723, 1000), ("13:51", 0.709, 2000), ("13:52", 0.707, 3000),
            ("13:53", 0.705, 4000), ("13:54", 0.707, 5000), ("13:55", 0.709, 6000),
            ("13:56", 0.711, 7000), ("13:57", 0.709, 8000), ("13:58", 0.708, 9000),
            ("13:59", 0.707, 10000), ("14:00", 0.709, 11000), ("14:01", 0.710, 12000),
            ("14:02", 0.711, 13000), ("14:03", 0.710, 14000), ("14:04", 0.711, 15000),
            ("14:05", 0.713, 18000),
        ]
        return [
            {"quote_minute": f"2026-08-24T{minute}", "price": price, "volume": volume, "previous_close": 0.729}
            for minute, price, volume in prices
        ]

    def test_intraday_rebound_requires_breakout_after_higher_low(self):
        samples = self._intraday_rebound_sample()
        before_breakout = self.module.evaluate_intraday_rebound(samples[:-1])
        result = self.module.evaluate_intraday_rebound(samples)

        self.assertEqual(before_breakout["stage"], "CANDIDATE")
        self.assertTrue(result["matched"])
        self.assertEqual(result["stage"], "CONFIRMED")
        self.assertEqual(result["trough_price"], 0.705)
        self.assertEqual(result["breakout_price"], 0.711)
        self.assertGreater(result["higher_low"], result["trough_price"])
        self.assertGreaterEqual(result["volume_multiple"], 1.5)

    def test_intraday_rebound_alert_persists_minutes_and_deduplicates(self):
        now = "2026-08-24T14:05:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '159516', '半导体设备ETF', 3, '', ?, ?)""",
                (user_id, now, now),
            )
            for sample in self._intraday_rebound_sample():
                db.execute(
                    """INSERT INTO intraday_quotes
                    (stock_code, trade_date, quote_minute, price, previous_close, volume, source, fetched_at)
                    VALUES ('159516', '2026-08-24', ?, ?, ?, ?, 'test', ?)""",
                    (sample["quote_minute"], sample["price"], sample["previous_close"], sample["volume"], sample["quote_minute"]),
                )
        quote = {
            "159516": {
                "trade_date": "2026-08-24", "quote_time": now, "open": 0.73, "high": 0.738,
                "low": 0.705, "close": 0.713, "previous_close": 0.729, "volume": 18000,
                "amount": 12834, "fetched_at": now,
            }
        }
        with self.module.db_connect() as db:
            self.module.evaluate_intraday_rebound_alerts(db, quote)
            self.module.evaluate_intraday_rebound_alerts(db, quote)
            notifications = db.execute(
                "SELECT stage, title, content FROM notifications WHERE stock_code = '159516' ORDER BY id"
            ).fetchall()
            events = db.execute(
                "SELECT stage, signal_price, rule_version FROM alert_signal_events ORDER BY id"
            ).fetchall()
        self.assertEqual(len(notifications), 2)
        self.assertEqual([row["stage"] for row in notifications], ["CANDIDATE", "CONFIRMED"])
        self.assertIn("日内反弹", notifications[0]["title"])
        self.assertIn("固定突破位 0.711", notifications[1]["content"])
        self.assertEqual([row["stage"] for row in events], ["CANDIDATE", "CONFIRMED"])
        self.assertTrue(all(row["rule_version"] == self.module.INTRADAY_REBOUND_RULE_VERSION for row in events))

    def test_alert_signal_event_outcomes_require_complete_windows(self):
        now = "2026-08-24T10:00:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            alert_type_id = db.execute(
                "SELECT id FROM alert_types WHERE code = ?", (self.module.INTRADAY_REBOUND_CODE,)
            ).fetchone()["id"]
            for minute, price in (("10:00", 10.0), ("10:05", 10.5), ("10:10", 11.0), ("10:20", 9.5)):
                db.execute(
                    """INSERT INTO intraday_quotes
                    (stock_code, trade_date, quote_minute, price, previous_close, volume, source, fetched_at)
                    VALUES ('000001', '2026-08-24', ?, ?, 10.5, 100, 'test', ?)""",
                    (f"2026-08-24T{minute}", price, f"2026-08-24T{minute}:00"),
                )
            event_id = self.module.upsert_alert_signal_event(
                db, user_id, alert_type_id, "000001", "平安银行", "CANDIDATE", now, 10.0,
                self.module.INTRADAY_REBOUND_RULE_VERSION, {}, {},
            )
            outcome = db.execute(
                "SELECT * FROM alert_signal_event_outcomes WHERE event_id = ?", (event_id,)
            ).fetchone()
        self.assertAlmostEqual(outcome["minute_5_return"], 0.05)
        self.assertAlmostEqual(outcome["minute_10_return"], 0.1)
        self.assertAlmostEqual(outcome["minute_20_return"], -0.05)
        self.assertIsNone(outcome["session_close_return"])

    def test_alert_signal_event_daily_outcomes_wait_for_full_horizon(self):
        now = "2026-08-20T15:00:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            alert_type_id = db.execute(
                "SELECT id FROM alert_types WHERE code = ?", (self.module.THREE_DAY_DIP_CODE,)
            ).fetchone()["id"]
            for trade_date, high, low, close in (
                ("2026-08-21", 11.0, 9.5, 10.5), ("2026-08-24", 12.0, 10.0, 11.0),
            ):
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, high, low, close, source, fetched_at)
                    VALUES ('000001', ?, ?, ?, ?, 'test', ?)""",
                    (trade_date, high, low, close, now),
                )
            event_id = self.module.upsert_alert_signal_event(
                db, user_id, alert_type_id, "000001", "平安银行", "CONFIRMED", now, 10.0,
                self.module.THREE_DAY_DIP_RULE_VERSION, {}, {},
            )
            outcome = db.execute(
                "SELECT * FROM alert_signal_event_outcomes WHERE event_id = ?", (event_id,)
            ).fetchone()
        self.assertAlmostEqual(outcome["day_1_close_return"], 0.05)
        self.assertAlmostEqual(outcome["day_1_max_return"], 0.1)
        self.assertIsNone(outcome["day_3_close_return"])
        self.assertIsNone(outcome["day_3_max_return"])

    def test_alert_signal_event_backfill_is_idempotent(self):
        now = "2026-08-20T14:30:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            alert_type_id = db.execute(
                "SELECT id FROM alert_types WHERE code = ?", (self.module.THREE_DAY_DIP_CODE,)
            ).fetchone()["id"]
            db.execute(
                """INSERT INTO notifications
                (user_id, alert_type_id, stock_code, stock_name, stage, title, content,
                 details_json, quote_time, created_at, dedupe_key)
                VALUES (?, ?, '000001', '平安银行', 'CANDIDATE', '历史提醒', '历史内容',
                ?, ?, ?, 'backfill-test')""",
                (user_id, alert_type_id, json.dumps({"price": 10.5}), now, now),
            )
            self.assertEqual(self.module.backfill_alert_signal_events(db), 1)
            self.assertEqual(self.module.backfill_alert_signal_events(db), 0)
            event = db.execute("SELECT * FROM alert_signal_events").fetchone()
        self.assertEqual(event["signal_price"], 10.5)
        self.assertTrue(json.loads(event["params_json"])["backfilled"])

    def test_alert_signal_event_backfill_reuses_same_day_three_day_event(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            alert_type_id = db.execute(
                "SELECT id FROM alert_types WHERE code = ?", (self.module.THREE_DAY_DIP_CODE,)
            ).fetchone()["id"]
            self.module.upsert_alert_signal_event(
                db, user_id, alert_type_id, "000001", "平安银行", "CANDIDATE", "2026-08-20T14:29:57",
                10.5, self.module.THREE_DAY_DIP_RULE_VERSION, {}, {},
            )
            db.execute(
                """INSERT INTO notifications
                (user_id, alert_type_id, stock_code, stock_name, stage, title, content,
                 details_json, quote_time, created_at, dedupe_key)
                VALUES (?, ?, '000001', '平安银行', 'CANDIDATE', '历史提醒', '历史内容',
                ?, '2026-08-20T14:30:00', '2026-08-20T14:30:00', 'backfill-same-day-test')""",
                (user_id, alert_type_id, json.dumps({"price": 10.5})),
            )
            self.assertEqual(self.module.backfill_alert_signal_events(db), 0)
            events = db.execute(
                """SELECT * FROM alert_signal_events
                WHERE stock_code = '000001' AND signal_date = '2026-08-20'"""
            ).fetchall()
        self.assertEqual(len(events), 1)
        self.assertEqual(events[0]["signal_time"], "2026-08-20T14:29:57")

    def test_intraday_rebound_does_not_cross_lunch_break(self):
        morning = self._intraday_rebound_sample()
        afternoon = [
            {"quote_minute": f"2026-08-24T13:{minute:02d}", "price": 0.716 + minute * 0.001, "volume": 20000 + minute * 1000}
            for minute in range(7)
        ]
        result = self.module.evaluate_intraday_rebound(morning + afternoon)
        self.assertFalse(result["matched"])
        self.assertIn(result["status"], ("等待分钟行情", "等待止跌确认", "未形成急跌结构"))

    def test_intraday_rebound_159516_replay_is_early(self):
        prices = [
            ("09:33", 0.710, 1000), ("09:34", 0.705, 2000), ("09:35", 0.700, 3000),
            ("09:36", 0.695, 4000), ("09:37", 0.690, 5000), ("09:38", 0.687, 6000),
            ("09:39", 0.685, 7000), ("09:40", 0.690, 8000), ("09:41", 0.692, 9000),
            ("09:42", 0.690, 10000), ("09:43", 0.693, 11000), ("09:44", 0.691, 12000),
            ("09:45", 0.687, 13000), ("09:46", 0.683, 14000), ("09:47", 0.685, 14800),
            ("09:48", 0.686, 15400), ("09:49", 0.686, 16100), ("09:50", 0.685, 17000),
            ("09:51", 0.690, 18000),
        ]
        samples = [
            {"quote_minute": f"2026-08-25T{minute}", "price": price, "volume": volume, "previous_close": 34.8}
            for minute, price, volume in prices
        ]
        result = self.module.evaluate_intraday_rebound(samples)
        self.assertIn(result["stage"], ("CANDIDATE", "CONFIRMED"))
        self.assertEqual(result["trough_minute"], "2026-08-25T09:46")
        self.assertEqual(result["trough_price"], 0.683)

    def test_intraday_rebound_requires_waterline_drop(self):
        samples = [
            {"quote_minute": f"2026-08-25T13:{minute:02d}", "price": price, "volume": 1000 + minute * 100, "previous_close": 41.4}
            for minute, price in enumerate((42.8, 42.5, 42.3, 42.27, 42.33, 42.45, 42.57, 42.6))
        ]
        result = self.module.evaluate_intraday_rebound(samples)
        self.assertFalse(result["matched"])
        self.assertEqual(result["status"], "未形成急跌结构")

    def _exhaustion_sample(self):
        return [
            {"trade_date": "2026-08-17", "open": 133.50, "high": 135.79, "low": 129.50, "close": 135.50, "volume": 23473900},
            {"trade_date": "2026-08-18", "open": 135.75, "high": 136.55, "low": 128.21, "close": 131.70, "volume": 30619800},
            {"trade_date": "2026-08-19", "open": 126.00, "high": 127.15, "low": 118.53, "close": 118.53, "volume": 31212500},
            {"trade_date": "2026-08-20", "open": 118.53, "high": 121.84, "low": 115.17, "close": 118.28, "volume": 29231900},
        ]

    def test_three_day_dip_exhaustion_and_strong_reversal_samples(self):
        result = self.module.evaluate_three_day_dip(self._exhaustion_sample(), enforce_volume=True)
        self.assertTrue(result["matched"])
        self.assertEqual(result["pattern_type"], "EXHAUSTION")
        reversal = [
            {"trade_date": "2026-06-10", "open": 145.08, "high": 158.80, "low": 141.22, "close": 154.96, "volume": 34759500},
            {"trade_date": "2026-06-11", "open": 150.01, "high": 157.88, "low": 139.46, "close": 139.46, "volume": 29529300},
            {"trade_date": "2026-06-12", "open": 137.07, "high": 149.87, "low": 125.51, "close": 129.80, "volume": 54606600},
            {"trade_date": "2026-06-15", "open": 128.51, "high": 138.88, "low": 119.00, "close": 137.30, "volume": 37989800},
        ]
        result = self.module.evaluate_three_day_dip(reversal, enforce_volume=True)
        self.assertTrue(result["matched"])
        self.assertEqual(result["pattern_type"], "STRONG_REVERSAL")

    def test_three_day_dip_requires_bullish_baseline(self):
        prices = self._exhaustion_sample()
        result = self.module.evaluate_three_day_dip(prices, enforce_volume=True)
        self.assertTrue(result["baseline_bullish_ok"])
        self.assertTrue(result["matched"])

        bearish = [dict(row) for row in prices]
        bearish[0]["open"] = bearish[0]["close"] + 1
        result = self.module.evaluate_three_day_dip(bearish, enforce_volume=True)
        self.assertFalse(result["baseline_bullish_ok"])
        self.assertFalse(result["common_ok"])
        self.assertFalse(result["matched"])

        doji = [dict(row) for row in prices]
        doji[0]["open"] = doji[0]["close"]
        result = self.module.evaluate_three_day_dip(doji, enforce_volume=True)
        self.assertFalse(result["baseline_bullish_ok"])
        self.assertFalse(result["matched"])

        result = self.module.evaluate_three_day_dip(
            bearish, {"require_bullish_baseline": False}, enforce_volume=True,
        )
        self.assertTrue(result["baseline_bullish_ok"])
        self.assertTrue(result["matched"])

    def test_three_day_dip_shadow_stop_is_candidate_only(self):
        prices = [
            {"trade_date": "2026-08-17", "open": 40.99, "high": 45.09, "low": 40.98, "close": 45.09, "volume": 428071},
            {"trade_date": "2026-08-18", "open": 45.54, "high": 45.99, "low": 42.65, "close": 43.67, "volume": 476475},
            {"trade_date": "2026-08-19", "open": 41.92, "high": 42.44, "low": 39.30, "close": 39.36, "volume": 455993},
            {"trade_date": "2026-08-20", "open": 39.74, "high": 41.13, "low": 38.60, "close": 39.25, "volume": 426727},
        ]
        result = self.module.evaluate_three_day_dip(prices, enforce_volume=True)
        self.assertTrue(result["matched"])
        self.assertEqual(result["pattern_type"], "SHADOW_STOP")
        self.assertGreater(result["lower_shadow_body_ratio"], 1)

    def test_three_day_dip_allows_signal_low_one_percent_above_prior_low(self):
        prices = [
            {"trade_date": "2026-08-18", "open": 72.01, "high": 75.50, "low": 71.11, "close": 72.94, "volume": 1184806},
            {"trade_date": "2026-08-19", "open": 68.80, "high": 70.60, "low": 65.68, "close": 66.16, "volume": 1048449},
            {"trade_date": "2026-08-20", "open": 68.00, "high": 68.24, "low": 63.58, "close": 64.45, "volume": 843824},
            {"trade_date": "2026-08-21", "open": 64.51, "high": 68.60, "low": 63.91, "close": 68.14, "volume": 835712},
        ]
        result = self.module.evaluate_three_day_dip(prices, enforce_volume=True)
        self.assertFalse(result["new_low_ok"])
        self.assertTrue(result["low_position_ok"])
        self.assertTrue(result["common_ok"])
        self.assertTrue(result["matched"])
        self.assertEqual(result["pattern_type"], "STRONG_REVERSAL")
        self.assertAlmostEqual(result["signal_low_above_prior_ratio"], 63.91 / 63.58 - 1)

        prices[-1]["low"] = 66.20
        result = self.module.evaluate_three_day_dip(prices, enforce_volume=True)
        self.assertFalse(result["low_position_ok"])
        self.assertFalse(result["common_ok"])

    def test_three_day_dip_strong_reversal_allows_four_percent_above_prior_low(self):
        prices = [
            {"trade_date": "2025-11-17", "open": 16.55, "high": 18.24, "low": 16.10, "close": 18.24, "volume": 181425400},
            {"trade_date": "2025-11-21", "open": 15.05, "high": 17.97, "low": 14.59, "close": 14.59, "volume": 227157400},
            {"trade_date": "2025-11-24", "open": 14.00, "high": 14.75, "low": 11.92, "close": 12.58, "volume": 210811600},
            {"trade_date": "2025-11-25", "open": 12.58, "high": 14.09, "low": 12.34, "close": 13.43, "volume": 192976400},
        ]
        result = self.module.evaluate_three_day_dip(prices, enforce_volume=True)
        self.assertTrue(result["low_position_ok"])
        self.assertTrue(result["matched"])
        self.assertEqual(result["pattern_type"], "STRONG_REVERSAL")
        self.assertAlmostEqual(result["signal_low_above_prior_ratio"], 12.34 / 11.92 - 1)

    def test_three_day_dip_realtime_alerts_deduplicate_candidate_and_confirmation(self):
        now = "2026-08-20T14:30:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '600482', '中国动力', 3, '', ?, ?)""",
                (user_id, now, now),
            )
            for row in self._exhaustion_sample()[:-1]:
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('600482', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )

        candidate_quote = dict(self._exhaustion_sample()[-1], fetched_at=now, amount=1)
        candidate_quote["volume"] /= 100
        candidate = {"600482": candidate_quote}
        with patch.object(self.module, "fetch_realtime_prices", return_value=candidate):
            self.module.sync_realtime_codes(["600482"])
            self.module.sync_realtime_codes(["600482"])
        with self.module.db_connect() as db:
            stages = [row["stage"] for row in db.execute("SELECT stage FROM notifications ORDER BY id")]
            event_stages = [row["stage"] for row in db.execute("SELECT stage FROM alert_signal_events ORDER BY id")]
        self.assertEqual(stages, ["CANDIDATE"])
        self.assertEqual(event_stages, ["CANDIDATE"])

        confirmed = {"600482": dict(candidate["600482"], fetched_at="2026-08-20T15:00:00")}
        with patch.object(self.module, "fetch_realtime_prices", return_value=confirmed):
            self.module.sync_realtime_codes(["600482"])
            self.module.sync_realtime_codes(["600482"])
        with self.module.db_connect() as db:
            stages = [row["stage"] for row in db.execute("SELECT stage FROM notifications ORDER BY id")]
            event_stages = [row["stage"] for row in db.execute("SELECT stage FROM alert_signal_events ORDER BY id")]
        self.assertEqual(stages, ["CANDIDATE", "CONFIRMED"])
        self.assertEqual(event_stages, ["CANDIDATE", "CONFIRMED"])

    def _consolidation_stabilization_sample(self):
        return [
            {"trade_date": "2026-05-15", "open": 26.20, "high": 26.60, "low": 25.80, "close": 26.39, "volume": 24000000},
            {"trade_date": "2026-05-18", "open": 26.30, "high": 26.70, "low": 26.00, "close": 26.35, "volume": 25000000},
            {"trade_date": "2026-05-19", "open": 26.20, "high": 26.55, "low": 25.90, "close": 26.10, "volume": 23500000},
            {"trade_date": "2026-05-20", "open": 26.07, "high": 26.72, "low": 25.91, "close": 26.47, "volume": 24648000},
            {"trade_date": "2026-05-21", "open": 26.40, "high": 27.17, "low": 26.14, "close": 26.48, "volume": 33710600},
            {"trade_date": "2026-05-22", "open": 26.53, "high": 26.60, "low": 25.53, "close": 25.73, "volume": 33067700},
            {"trade_date": "2026-05-25", "open": 25.52, "high": 25.88, "low": 25.20, "close": 25.59, "volume": 20576500},
            {"trade_date": "2026-05-26", "open": 25.46, "high": 25.61, "low": 25.02, "close": 25.30, "volume": 20339000},
            {"trade_date": "2026-05-27", "open": 25.30, "high": 25.52, "low": 24.82, "close": 25.11, "volume": 26173800},
            {"trade_date": "2026-05-28", "open": 25.39, "high": 25.42, "low": 24.65, "close": 24.82, "volume": 22240500},
            {"trade_date": "2026-05-29", "open": 24.85, "high": 25.46, "low": 24.53, "close": 24.90, "volume": 31425500},
            {"trade_date": "2026-06-01", "open": 24.46, "high": 24.98, "low": 23.90, "close": 24.34, "volume": 24089000},
            {"trade_date": "2026-06-02", "open": 24.28, "high": 24.38, "low": 23.50, "close": 23.84, "volume": 22939900},
            {"trade_date": "2026-06-03", "open": 23.73, "high": 23.73, "low": 23.14, "close": 23.29, "volume": 30451400},
            {"trade_date": "2026-06-04", "open": 23.16, "high": 23.28, "low": 22.61, "close": 22.78, "volume": 24355900},
            {"trade_date": "2026-06-05", "open": 23.22, "high": 23.60, "low": 22.80, "close": 22.87, "volume": 27941800},
            {"trade_date": "2026-06-08", "open": 22.51, "high": 22.89, "low": 21.60, "close": 21.91, "volume": 24121600},
            {"trade_date": "2026-06-09", "open": 21.92, "high": 22.05, "low": 21.38, "close": 22.00, "volume": 20537300},
            {"trade_date": "2026-06-10", "open": 21.79, "high": 22.35, "low": 21.76, "close": 22.08, "volume": 20901200},
            {"trade_date": "2026-06-11", "open": 21.80, "high": 21.92, "low": 21.25, "close": 21.52, "volume": 29624800},
            {"trade_date": "2026-06-12", "open": 21.69, "high": 22.18, "low": 20.81, "close": 22.18, "volume": 70001500},
            {"trade_date": "2026-06-15", "open": 21.91, "high": 22.28, "low": 21.70, "close": 22.27, "volume": 42825200},
            {"trade_date": "2026-06-16", "open": 21.98, "high": 22.04, "low": 21.61, "close": 21.79, "volume": 25921000},
            {"trade_date": "2026-06-17", "open": 21.73, "high": 22.17, "low": 21.60, "close": 21.97, "volume": 24893000},
            {"trade_date": "2026-06-18", "open": 21.82, "high": 22.74, "low": 21.66, "close": 22.51, "volume": 41863700},
        ]

    def test_consolidation_stabilization_kanglong_sample_confirms_on_june_12(self):
        prices = self._consolidation_stabilization_sample()
        signal_index = next(index for index, row in enumerate(prices) if row["trade_date"] == "2026-06-12")
        result = self.module.evaluate_consolidation_stabilization(prices[:signal_index + 1])
        self.assertTrue(result["matched"])
        self.assertEqual(result["pattern_type"], "ANOMALY")
        self.assertEqual(result["signal_type"], "DOWNTREND_REVERSAL")
        self.assertEqual(result["signal"]["trade_date"], "2026-06-12")
        self.assertAlmostEqual(result["change_ratio"], 22.18 / 21.52 - 1)
        self.assertEqual(result["volume_window"], 20)
        expected_median = sorted(row["volume"] for row in prices[signal_index - 20:signal_index])
        self.assertAlmostEqual(result["volume_baseline"], (expected_median[9] + expected_median[10]) / 2)
        self.assertGreater(result["volume_ratio"], 2.8)
        self.assertLess(result["volume_ratio"], 3.0)
        self.assertTrue(result["short_consolidation"])
        self.assertAlmostEqual(result["short_close_range_ratio"], (22.08 - 21.52) / 21.52)
        self.assertEqual(result["background_type"], "DOWNTREND_REBOUND")
        self.assertTrue(result["downtrend_observation_ok"])
        self.assertTrue(result["downtrend_reversal_ok"])
        self.assertFalse(result["breakout_up"])

        realtime_volume = [dict(row) for row in prices[:signal_index + 1]]
        realtime_volume[-1]["source"] = "tencent-realtime"
        self.assertTrue(self.module.evaluate_consolidation_stabilization(realtime_volume)["matched"])
        self.assertAlmostEqual(
            self.module.evaluate_consolidation_stabilization(realtime_volume)["volume_ratio"],
            result["volume_ratio"],
        )

        no_volume = [dict(row) for row in prices[:signal_index + 1]]
        no_volume[-1]["volume"] = 30000000
        self.assertFalse(self.module.evaluate_consolidation_stabilization(no_volume)["matched"])
        low_change = [dict(row) for row in prices[:signal_index + 1]]
        low_change[-1]["close"] = 21.90
        self.assertFalse(self.module.evaluate_consolidation_stabilization(low_change)["matched"])

    def test_anomaly_boundaries_strength_and_no_future_data(self):
        prices = self._consolidation_stabilization_sample()
        signal_index = next(index for index, row in enumerate(prices) if row["trade_date"] == "2026-06-12")
        signal_prices = [dict(row) for row in prices[:signal_index + 1]]
        baseline_result = self.module.evaluate_consolidation_stabilization(signal_prices)

        future_changed = signal_prices + [
            {"trade_date": "2026-06-15", "open": 1, "high": 100, "low": 1, "close": 100, "volume": 9999999999},
        ]
        historical_result = self.module.evaluate_consolidation_stabilization(future_changed[:-1])
        self.assertEqual(historical_result["signal"]["trade_date"], "2026-06-12")
        self.assertAlmostEqual(historical_result["volume_ratio"], baseline_result["volume_ratio"])

        exactly_two_percent = [dict(row) for row in signal_prices]
        exactly_two_percent[-1]["close"] = exactly_two_percent[-2]["close"] * 1.02
        self.assertFalse(self.module.evaluate_consolidation_stabilization(exactly_two_percent)["matched"])

        strong_by_change = [dict(row) for row in signal_prices]
        strong_by_change[-1]["close"] = strong_by_change[-2]["close"] * 1.04
        self.assertEqual(
            self.module.evaluate_consolidation_stabilization(strong_by_change)["pattern_type"],
            "STRONG_ANOMALY",
        )
        strong_by_volume = [dict(row) for row in signal_prices]
        strong_by_volume[-1]["volume"] = baseline_result["volume_baseline"] * 3
        self.assertEqual(
            self.module.evaluate_consolidation_stabilization(strong_by_volume)["pattern_type"],
            "STRONG_ANOMALY",
        )

    def test_zhongke_downtrend_anomaly_is_observation_only(self):
        raw = [
            ("2026-05-13", 95.62, 100.62, 94.64, 100.33, 82187000),
            ("2026-05-14", 104.62, 104.62, 97.12, 97.21, 97816700),
            ("2026-05-15", 97.14, 98.98, 93.62, 94.08, 80334000),
            ("2026-05-18", 93.91, 101.22, 93.72, 98.32, 92563800),
            ("2026-05-19", 97.62, 98.17, 94.62, 97.88, 64716700),
            ("2026-05-20", 96.85, 99.11, 96.62, 98.00, 55467800),
            ("2026-05-21", 100.48, 100.62, 92.60, 92.87, 83566200),
            ("2026-05-22", 92.88, 95.06, 92.64, 94.42, 55590400),
            ("2026-05-25", 94.92, 96.07, 93.22, 96.07, 66580000),
            ("2026-05-26", 95.52, 95.60, 91.53, 93.95, 68746400),
            ("2026-05-27", 93.71, 95.58, 89.91, 90.61, 59121000),
            ("2026-05-28", 90.54, 91.78, 88.62, 90.87, 48132700),
            ("2026-05-29", 92.92, 93.62, 87.51, 87.84, 53148500),
            ("2026-06-01", 88.50, 89.54, 85.36, 85.40, 40355800),
            ("2026-06-02", 86.03, 86.54, 83.45, 86.02, 38634400),
            ("2026-06-03", 85.56, 88.78, 85.06, 86.32, 44761300),
            ("2026-06-04", 85.00, 86.21, 84.12, 84.40, 33937500),
            ("2026-06-05", 83.67, 84.82, 82.14, 82.37, 38642200),
            ("2026-06-08", 79.90, 80.69, 77.77, 78.51, 42722300),
            ("2026-06-09", 79.60, 80.49, 78.90, 80.40, 30629600),
            ("2026-06-10", 81.87, 86.33, 81.87, 82.70, 64650600),
        ]
        prices = [
            {"trade_date": day, "open": open_, "high": high, "low": low, "close": close, "volume": volume}
            for day, open_, high, low, close, volume in raw
        ]
        result = self.module.evaluate_consolidation_stabilization(prices)
        self.assertTrue(result["downtrend_observation_ok"])
        self.assertFalse(result["downtrend_reversal_ok"])
        self.assertTrue(result["candidate_ok"])
        self.assertFalse(result["confirmed_ok"])
        self.assertEqual(result["signal_type"], "DOWNTREND_ANOMALY")
        self.assertAlmostEqual(result["downtrend_volume_ratio"], 1.5563812845984681)
        self.assertAlmostEqual(result["close_position"], 0.18609865470852005)

        now = "2026-06-10T15:01:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '603019', '中科曙光', 3, '', ?, ?)""",
                (user_id, now, now),
            )
            for row in prices:
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('603019', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )
            quote = {"603019": dict(prices[-1], fetched_at=now, quote_time=now)}
            self.assertEqual(self.module.evaluate_consolidation_stabilization_alerts(db, quote), 1)
            self.assertEqual(self.module.evaluate_consolidation_stabilization_alerts(db, quote), 0)
            notification = db.execute("SELECT stage, content FROM notifications").fetchone()
            event = db.execute("SELECT stage, metrics_json FROM alert_signal_events").fetchone()
        self.assertEqual(notification["stage"], "CANDIDATE")
        self.assertIn("下跌异动观察", notification["content"])
        self.assertIn("前 10 日中位量的 1.56 倍", notification["content"])
        self.assertEqual(event["stage"], "CANDIDATE")
        self.assertEqual(json.loads(event["metrics_json"])["signal_type"], "DOWNTREND_ANOMALY")

    def test_consolidation_stabilization_realtime_is_confirmed_user_scoped_and_deduplicated(self):
        prices = self._consolidation_stabilization_sample()
        signal_index = next(index for index, row in enumerate(prices) if row["trade_date"] == "2026-06-12")
        signal_prices = prices[:signal_index + 1]
        now = "2026-06-12T14:30:00"
        with self.module.db_connect() as db:
            first_user = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_user = db.execute(
                "INSERT INTO users (name, created_at) VALUES ('second', ?) RETURNING id",
                (now,),
            ).fetchone()["id"]
            for user_id in (first_user, second_user):
                db.execute(
                    """INSERT INTO watchlist_stocks
                    (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                    VALUES (?, '300759', '康龙化成', 3, '', ?, ?)""",
                    (user_id, now, now),
                )
            for row in signal_prices:
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('300759', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )
            quote = {"300759": dict(signal_prices[-1], fetched_at=now, quote_time=now)}
            self.assertEqual(self.module.evaluate_consolidation_stabilization_alerts(db, quote), 2)
            self.assertEqual(self.module.evaluate_consolidation_stabilization_alerts(db, quote), 0)
            closing_time = "2026-06-12T15:01:00"
            quote["300759"].update({"fetched_at": closing_time, "quote_time": closing_time})
            self.assertEqual(self.module.evaluate_consolidation_stabilization_alerts(db, quote), 2)
            self.assertEqual(self.module.evaluate_consolidation_stabilization_alerts(db, quote), 0)
            notifications = db.execute(
                "SELECT stage, content FROM notifications ORDER BY user_id"
            ).fetchall()
            events = db.execute(
                "SELECT stage, rule_version, metrics_json FROM alert_signal_events ORDER BY user_id"
            ).fetchall()
        self.assertEqual([row["stage"] for row in notifications].count("CANDIDATE"), 2)
        self.assertEqual([row["stage"] for row in notifications].count("CONFIRMED"), 2)
        self.assertTrue(any("下跌异动观察" in row["content"] for row in notifications))
        self.assertTrue(any("下跌反转确认" in row["content"] for row in notifications))
        self.assertEqual(len(events), 4)
        self.assertTrue(all(row["rule_version"] == self.module.CONSOLIDATION_STABILIZATION_RULE_VERSION for row in events))
        self.assertTrue(all(json.loads(row["metrics_json"])["pattern_type"] == "ANOMALY" for row in events))

    def test_notification_api_is_user_scoped(self):
        now = "2026-08-14T15:00:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('notify-second', ?)", (now,)).lastrowid
            alert_type_id = db.execute("SELECT id FROM alert_types WHERE code = ?", (self.module.THREE_DAY_DIP_CODE,)).fetchone()["id"]
            notification_id = db.execute(
                """INSERT INTO notifications
                (user_id, alert_type_id, stock_code, stock_name, stage, title, content,
                 details_json, quote_time, created_at, dedupe_key)
                VALUES (?, ?, '600482', '中国动力', 'CANDIDATE', '测试提醒', '只属于第二用户', '{}', ?, ?, 'scope-test')""",
                (second_id, alert_type_id, now, now),
            ).lastrowid

        response = self.client.get("/api/notifications")
        self.assertEqual(response.get_json()["unread_count"], 0)
        self.assertEqual(self.client.post(f"/api/notifications/{notification_id}/read").status_code, 404)
        with self.client.session_transaction() as session:
            session["user_id"] = second_id
        response = self.client.get("/api/notifications")
        self.assertEqual(response.get_json()["unread_count"], 1)
        self.assertEqual(response.get_json()["notifications"][0]["title"], "测试提醒")
        self.assertEqual(self.client.post(f"/api/notifications/{notification_id}/read").status_code, 200)
        self.assertEqual(self.client.get("/api/notifications").get_json()["unread_count"], 0)

    def _insert_notification(self, db, user_id, index, read=False, alert_code=None):
        now = "2026-08-14T15:00:00"
        alert_code = alert_code or self.module.THREE_DAY_DIP_CODE
        alert_type_id = db.execute("SELECT id FROM alert_types WHERE code = ?", (alert_code,)).fetchone()["id"]
        return db.execute(
            """INSERT INTO notifications
            (user_id, alert_type_id, stock_code, stock_name, stage, title, content,
             details_json, quote_time, created_at, read_at, dedupe_key)
            VALUES (?, ?, '600482', '中国动力', ?, ?, ?, '{}', ?, ?, ?, ?)""",
            (
                user_id, alert_type_id, "CANDIDATE" if index % 2 else "CONFIRMED",
                f"通知标题 {index}", f"通知内容 {index}",
                now, now, now if read else None, f"page-key-{user_id}-{alert_code}-{index}",
            ),
        ).lastrowid

    def test_notifications_page_pagination_and_read(self):
        now = "2026-08-14T15:00:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            for index in range(25):
                self._insert_notification(db, user_id, index)

        page = self.client.get("/analysis/notifications").get_data(as_text=True)
        self.assertIn("共 25 条", page)
        self.assertIn("第 1 / 2 页", page)
        self.assertIn("下一页", page)
        self.assertIn("通知标题 24", page)
        self.assertIn("通知标题 5", page)
        second = self.client.get("/analysis/notifications?page=2").get_data(as_text=True)
        self.assertIn("通知标题 0", second)
        self.assertIn("通知标题 4", second)
        self.assertNotIn("通知标题 5", second)
        self.assertIn("第 2 / 2 页", second)

        with self.module.db_connect() as db:
            newest = db.execute("SELECT id FROM notifications ORDER BY created_at DESC, id DESC LIMIT 1").fetchone()["id"]
        self.assertEqual(self.client.post(f"/api/notifications/{newest}/read").status_code, 200)
        page = self.client.get("/analysis/notifications").get_data(as_text=True)
        self.assertIn("24 条未读", page)
        first_item = page.split('class="notification-page-item', 1)
        self.assertEqual(len(first_item), 2)
        self.assertNotIn("unread", first_item[1].split('"', 1)[0])

    def test_notifications_page_user_isolated_and_empty_state(self):
        now = "2026-08-14T15:00:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('page-second', ?)", (now,)).lastrowid
            self._insert_notification(db, second_id, 1)

        page = self.client.get("/analysis/notifications").get_data(as_text=True)
        self.assertIn("还没有提醒通知", page)
        self.assertNotIn("通知标题 1", page)
        with self.client.session_transaction() as session:
            session["user_id"] = second_id
        page = self.client.get("/analysis/notifications").get_data(as_text=True)
        self.assertIn("通知标题 1", page)

    def test_notifications_page_filters_supported_strategy_types(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            self._insert_notification(db, user_id, 1, alert_code=self.module.THREE_DAY_DIP_CODE)
            self._insert_notification(db, user_id, 2, alert_code=self.module.INTRADAY_REBOUND_CODE)
            self._insert_notification(db, user_id, 3, alert_code=self.module.WATCHLIST_LIMIT_UP_CODE)
            self._insert_notification(db, user_id, 4, alert_code=self.module.WATCHLIST_OPEN_GAIN_CODE)
            self._insert_notification(db, user_id, 5, alert_code=self.module.WATCHLIST_OPEN_LOSS_CODE)
            self._insert_notification(db, user_id, 6, alert_code=self.module.WATCHLIST_LIMIT_DOWN_CODE)
            today = datetime.now().date().isoformat()
            db.execute(
                "UPDATE notifications SET created_at = ?, quote_time = ? "
                "WHERE user_id = ? AND title IN (?, ?, ?, ?)",
                (f"{today}T15:00:00", f"{today}T15:00:00", user_id,
                 "通知标题 3", "通知标题 4", "通知标题 5", "通知标题 6"),
            )

        all_page = self.client.get("/analysis/notifications").get_data(as_text=True)
        self.assertIn("通知标题 1", all_page)
        self.assertIn("通知标题 2", all_page)
        self.assertIn("通知标题 3", all_page)
        self.assertIn("通知标题 4", all_page)
        self.assertIn("通知标题 5", all_page)
        self.assertIn("通知标题 6", all_page)
        self.assertIn("涨停提醒", all_page)
        self.assertIn("涨幅提醒", all_page)
        self.assertIn("跌幅提醒", all_page)
        self.assertIn("跌停提醒", all_page)

        dip_page = self.client.get(
            f"/analysis/notifications?type={self.module.THREE_DAY_DIP_CODE}"
        ).get_data(as_text=True)
        dip_list = dip_page.split('<div class="notification-page-list">', 1)[1].split("</div>", 1)[0]
        self.assertIn("通知标题 1", dip_list)
        self.assertNotIn("通知标题 2", dip_list)
        self.assertNotIn("通知标题 3", dip_list)
        self.assertNotIn("通知标题 4", dip_list)
        self.assertNotIn("通知标题 5", dip_list)
        self.assertNotIn("通知标题 6", dip_list)
        self.assertIn("当前 1 条", dip_page)

        rebound_page = self.client.get(
            f"/analysis/notifications?type={self.module.INTRADAY_REBOUND_CODE}"
        ).get_data(as_text=True)
        rebound_list = rebound_page.split('<div class="notification-page-list">', 1)[1].split("</div>", 1)[0]
        self.assertNotIn("通知标题 1", rebound_list)
        self.assertIn("通知标题 2", rebound_list)
        self.assertNotIn("通知标题 3", rebound_list)
        self.assertNotIn("通知标题 4", rebound_list)
        self.assertNotIn("通知标题 5", rebound_list)
        self.assertNotIn("通知标题 6", rebound_list)

        daily_page = self.client.get(
            f"/analysis/notifications?type={self.module.DAILY_ALERTS_FILTER_CODE}"
        ).get_data(as_text=True)
        daily_list = daily_page.split('<div class="notification-page-list">', 1)[1].split("</div>", 1)[0]
        self.assertNotIn("通知标题 1", daily_list)
        self.assertNotIn("通知标题 2", daily_list)
        self.assertIn("通知标题 3", daily_list)
        self.assertIn("通知标题 4", daily_list)
        self.assertIn("通知标题 5", daily_list)
        self.assertIn("通知标题 6", daily_list)
        self.assertIn("日常提醒", daily_page)
        self.assertIn("当前 4 条", daily_page)

        unknown_page = self.client.get("/analysis/notifications?type=UNKNOWN").get_data(as_text=True)
        self.assertIn("通知标题 1", unknown_page)
        self.assertIn("通知标题 2", unknown_page)
        self.assertIn("通知标题 3", unknown_page)
        self.assertIn("通知标题 4", unknown_page)
        self.assertIn("通知标题 5", unknown_page)
        self.assertIn("通知标题 6", unknown_page)

    def test_daily_notifications_only_show_today(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            self._insert_notification(db, user_id, 1, alert_code=self.module.WATCHLIST_LIMIT_UP_CODE)
            self._insert_notification(db, user_id, 2, alert_code=self.module.WATCHLIST_OPEN_GAIN_CODE)
            today = datetime.now().date().isoformat()
            db.execute(
                "UPDATE notifications SET created_at = ?, quote_time = ? WHERE title = ?",
                (f"{today}T00:00:00", f"{today}T00:00:00", "通知标题 2"),
            )

        page = self.client.get(
            f"/analysis/notifications?type={self.module.DAILY_ALERTS_FILTER_CODE}"
        ).get_data(as_text=True)
        daily_list = page.split('<div class="notification-page-list">', 1)[1].split("</div>", 1)[0]
        self.assertIn("通知标题 2", daily_list)
        self.assertNotIn("通知标题 1", daily_list)
        self.assertIn("当前 1 条", page)

    def test_notifications_filter_pagination_preserves_type(self):
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            for index in range(21):
                self._insert_notification(
                    db, user_id, index, alert_code=self.module.INTRADAY_REBOUND_CODE,
                )
            self._insert_notification(db, user_id, 30, alert_code=self.module.THREE_DAY_DIP_CODE)

        page = self.client.get(
            f"/analysis/notifications?type={self.module.INTRADAY_REBOUND_CODE}"
        ).get_data(as_text=True)
        self.assertIn("共 21 条", page)
        self.assertIn("type=INTRADAY_REBOUND", page)
        self.assertIn("page=2", page)

    def test_alert_type_admin_updates_parameters_without_builtin_sample(self):
        page = self.client.get("/admin/alert-types/three-day-dip").get_data(as_text=True)
        self.assertIn("衰竭止跌型", page)
        self.assertIn("强修复型", page)
        self.assertNotIn("中国动力", page)

        response = self.client.post(
            "/admin/alert-types/three-day-dip",
            data={
                "enabled": "on", "decline_days": "2", "require_bullish_baseline": "on",
                "require_bearish_candles": "on",
                "require_declining_closes": "on", "require_signal_day_new_low": "on",
                "min_decline_percent": "9", "max_volume_percent": "95",
                "exhaustion_min_repair_percent": "2", "exhaustion_max_body_percent": "10",
                "exhaustion_min_change_percent": "-1.5", "exhaustion_max_change_percent": "3",
                "reversal_min_decline_percent": "11", "reversal_min_repair_percent": "7",
                "reversal_min_recovery_percent": "85", "reversal_min_change_percent": "5",
                "reversal_max_volume_percent": "75",
                "intraday_candidate_enabled": "on", "close_confirmation_enabled": "on",
            },
            follow_redirects=True,
        )
        text = response.get_data(as_text=True)
        self.assertIn("已保存三日低吸提醒参数", text)
        self.assertIn("9.0", text)

    def test_daily_alert_settings_save_global_thresholds_and_switches(self):
        page = self.client.get("/admin/alert-types/daily-alerts").get_data(as_text=True)
        self.assertIn("日常提醒设置", page)
        self.assertIn("跌停提醒", page)
        response = self.client.post(
            "/admin/alert-types/daily-alerts",
            data={"gain_enabled": "on", "gain_threshold_percent": "5.5", "loss_threshold_percent": "3.5", "limit_down_enabled": "on"},
            follow_redirects=True,
        )
        self.assertIn("已保存日常提醒设置", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            rows = {
                row["code"]: row for row in db.execute(
                    "SELECT code, params_json, enabled FROM alert_types WHERE code IN (?, ?, ?, ?)",
                    (self.module.WATCHLIST_OPEN_GAIN_CODE, self.module.WATCHLIST_OPEN_LOSS_CODE,
                     self.module.WATCHLIST_LIMIT_UP_CODE, self.module.WATCHLIST_LIMIT_DOWN_CODE),
                ).fetchall()
            }
        self.assertTrue(rows[self.module.WATCHLIST_OPEN_GAIN_CODE]["enabled"])
        self.assertFalse(rows[self.module.WATCHLIST_OPEN_LOSS_CODE]["enabled"])
        self.assertFalse(rows[self.module.WATCHLIST_LIMIT_UP_CODE]["enabled"])
        self.assertTrue(rows[self.module.WATCHLIST_LIMIT_DOWN_CODE]["enabled"])
        self.assertAlmostEqual(json.loads(rows[self.module.WATCHLIST_OPEN_GAIN_CODE]["params_json"])["threshold_ratio"], 0.055)
        self.assertAlmostEqual(json.loads(rows[self.module.WATCHLIST_OPEN_LOSS_CODE]["params_json"])["threshold_ratio"], 0.035)

    def test_intraday_rebound_management_menu_and_parameters(self):
        page = self.client.get("/admin/alert-types/intraday-rebound").get_data(as_text=True)
        self.assertIn("日内反弹", page)
        self.assertIn("分钟级提醒", page)
        self.assertIn("急跌幅度至少", page)

        response = self.client.post(
            "/admin/alert-types/intraday-rebound",
            data={
                "enabled": "on", "lookback_minutes": "90", "min_drop_percent": "2.5",
                "candidate_min_rebound_percent": "1.2", "min_trough_age_minutes": "6",
                "candidate_min_volume_multiple": "1", "confirmation_min_volume_multiple": "2",
                "first_bounce_min_percent": "0.5", "max_entry_rebound_percent": "2",
            },
            follow_redirects=True,
        )
        self.assertIn("已保存日内反弹提醒参数", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            row = db.execute("SELECT params_json FROM alert_types WHERE code = ?", (self.module.INTRADAY_REBOUND_CODE,)).fetchone()
        params = json.loads(row["params_json"])
        self.assertEqual(params["lookback_minutes"], 90)
        self.assertEqual(params["min_drop_ratio"], 0.025)
        self.assertEqual(params["confirmation_min_volume_multiple"], 2.0)

    def test_consolidation_stabilization_admin_and_kanglong_backtest(self):
        now = "2026-06-18T15:01:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '300759', '康龙化成', 3, '', ?, ?)""",
                (user_id, now, now),
            )
            for row in self._consolidation_stabilization_sample():
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('300759', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )

        page = self.client.get(
            "/admin/alert-types/consolidation-stabilization?run_test=1&stock_code=300759&signal_date=2026-06-12"
        ).get_data(as_text=True)
        self.assertIn("异动提醒", page)
        self.assertIn("康龙化成", page)
        self.assertIn("下跌反转确认", page)
        self.assertIn("2.89 倍", page)
        self.assertIn("共同前提", page)
        self.assertIn("所有异动分支都必须先满足", page)
        self.assertIn("普通异动分支", page)
        self.assertIn("下跌异动分支", page)
        self.assertIn("共同最低涨幅", page)
        self.assertIn("这些条件不会单独触发异动", page)
        self.assertIn('name="signal_date"', page)
        self.assertNotIn('name="end_date"', page)

        legacy_page = self.client.get(
            "/admin/alert-types/consolidation-stabilization?run_test=1&stock_code=300759&start_date=2026-06-12&end_date=2026-06-12"
        ).get_data(as_text=True)
        self.assertIn("下跌反转确认", legacy_page)

        response = self.client.post(
            "/admin/alert-types/consolidation-stabilization",
            data={
                "enabled": "on", "volume_window": "15", "candidate_volume_ratio": "1.6",
                "confirmation_volume_ratio": "2.2", "min_change_percent": "2.5",
                "strong_change_percent": "5", "strong_volume_ratio": "3.5",
                "short_consolidation_days": "4", "short_close_range_percent": "4",
                "downtrend_lookback_days": "18", "downtrend_min_drop_percent": "9",
                "downtrend_volume_window": "8", "downtrend_candidate_volume_ratio": "1.6",
                "downtrend_confirmation_close_position_percent": "65",
            },
            follow_redirects=True,
        )
        self.assertIn("已保存异动提醒参数", response.get_data(as_text=True))
        with self.module.db_connect() as db:
            row = db.execute(
                "SELECT params_json FROM alert_types WHERE code = ?",
                (self.module.CONSOLIDATION_STABILIZATION_CODE,),
            ).fetchone()
        params = json.loads(row["params_json"])
        self.assertEqual(params["volume_window"], 15)
        self.assertEqual(params["short_consolidation_days"], 4)
        self.assertEqual(params["downtrend_lookback_days"], 18)
        self.assertEqual(params["downtrend_volume_window"], 8)
        self.assertAlmostEqual(params["min_change_ratio"], 0.025)
        self.assertAlmostEqual(params["confirmation_volume_ratio"], 2.2)
        self.assertAlmostEqual(params["downtrend_min_drop_ratio"], 0.09)
        self.assertAlmostEqual(params["downtrend_confirmation_close_position"], 0.65)

    def test_consolidation_stabilization_pool_manual_sample_scan_and_review(self):
        now = "2026-06-25T15:00:00"
        prices = self._consolidation_stabilization_sample() + [
            {"trade_date": "2026-06-19", "open": 22.55, "high": 23.20, "low": 22.40, "close": 23.00, "volume": 39000000},
        ]
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '300759', '康龙化成', 3, '', ?, ?)""",
                (user_id, now, now),
            )
            for row in prices:
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('300759', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )

        response = self.client.post(
            "/admin/alert-types/consolidation-stabilization/pool",
            data={"stock_code": "300759", "stock_name": "康龙化成", "signal_date": "2026-06-12", "note": "金标准"},
            follow_redirects=True,
        )
        text = response.get_data(as_text=True)
        self.assertIn("已加入异动样本池", text)
        self.assertIn("放量上涨异动", text)
        with self.module.db_connect() as db:
            sample = db.execute("SELECT * FROM alert_signal_samples").fetchone()
            outcome = db.execute("SELECT * FROM alert_signal_outcomes").fetchone()
        self.assertEqual(sample["pattern_type"], "ANOMALY")
        self.assertEqual(sample["rule_version"], self.module.CONSOLIDATION_STABILIZATION_RULE_VERSION)
        self.assertAlmostEqual(outcome["day_1_close_return"], 22.27 / 22.18 - 1)

        response = self.client.post(
            "/admin/alert-types/consolidation-stabilization/pool/scan",
            data={"start_date": "2026-06-12", "end_date": "2026-06-12"}, follow_redirects=True,
        )
        self.assertIn("历史扫描完成，处理 1 个命中样本", response.get_data(as_text=True))
        self.assertEqual(self.client.post(
            f"/admin/alert-types/consolidation-stabilization/pool/{sample['id']}/review",
            data={"review_status": "REJECTED", "note": "复核否决"},
        ).status_code, 302)
        with self.module.db_connect() as db:
            sample = db.execute("SELECT review_status, note FROM alert_signal_samples").fetchone()
        self.assertEqual(sample["review_status"], "REJECTED")
        self.assertEqual(sample["note"], "复核否决")

    def test_three_day_dip_backtest_finds_hits_without_writing_notifications(self):
        now = "2026-08-15T10:00:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            stock = db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '600482', '中国动力', 3, '', ?, ?) RETURNING stock_code, stock_name""",
                (user_id, now, now),
            ).fetchone()
            for row in self._exhaustion_sample():
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('600482', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )
            result = self.module.backtest_three_day_dip(
                db, stock, "2026-08-20", "2026-08-20", self.module.THREE_DAY_DIP_DEFAULT_PARAMS,
            )

        self.assertEqual(result["scanned_days"], 1)
        self.assertEqual([hit["signal_date"] for hit in result["hits"]], ["2026-08-20"])
        self.assertEqual(result["hits"][0]["pattern_type"], "EXHAUSTION")
        response = self.client.get(
            "/admin/alert-types/three-day-dip?run_test=1&stock_code=600482&start_date=2026-08-20&end_date=2026-08-20"
        )
        text = response.get_data(as_text=True)
        self.assertIn("重点观察历史命中测试", text)
        self.assertIn("衰竭止跌型", text)
        self.assertIn("衰竭止跌型", text)
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM alert_rule_states").fetchone()[0], 0)

    def test_three_day_dip_backtest_handles_insufficient_data_and_user_scope(self):
        now = "2026-08-15T10:00:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('backtest-second', ?)", (now,)).lastrowid
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '000001', '平安银行', 2, '', ?, ?)""",
                (first_id, now, now),
            )
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '600482', '中国动力', 2, '', ?, ?)""",
                (second_id, now, now),
            )
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, source, fetched_at)
                VALUES ('000001', '2026-08-14', 10, 11, 9, 10, 'test', ?)""",
                (now,),
            )

        insufficient = self.client.get(
            "/admin/alert-types/three-day-dip?run_test=1&stock_code=000001&start_date=2026-08-14&end_date=2026-08-14"
        ).get_data(as_text=True)
        self.assertIn("历史行情不足，无法完成测试", insufficient)
        scoped = self.client.get(
            "/admin/alert-types/three-day-dip?run_test=1&stock_code=600482&start_date=2026-08-14&end_date=2026-08-14"
        ).get_data(as_text=True)
        self.assertIn("请选择当前用户重点观察中的股票", scoped)
        self.assertNotIn("中国动力 · 600482", scoped)

    def test_three_day_dip_backtest_validates_date_range(self):
        now = "2026-08-15T10:00:00"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '000001', '平安银行', 2, '', ?, ?)""",
                (user_id, now, now),
            )
        text = self.client.get(
            "/admin/alert-types/three-day-dip?run_test=1&stock_code=000001&start_date=2026-08-15&end_date=2026-08-14"
        ).get_data(as_text=True)
        self.assertIn("开始日期不能晚于截止日期", text)

    def test_three_day_dip_pool_manual_sample_review_and_outcomes(self):
        now = "2026-08-23T15:00:00"
        prices = self._exhaustion_sample() + [
            {"trade_date": "2026-08-21", "open": 118.83, "high": 129.98, "low": 117.06, "close": 128.06, "volume": 34277800},
        ]
        with self.module.db_connect() as db:
            for row in prices:
                db.execute(
                    """INSERT INTO daily_prices
                    (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                    VALUES ('002851', ?, ?, ?, ?, ?, ?, 'test', ?)""",
                    (row["trade_date"], row["open"], row["high"], row["low"], row["close"], row["volume"], now),
                )
        response = self.client.post(
            "/admin/alert-types/three-day-dip/pool",
            data={"stock_code": "002851", "stock_name": "麦格米特", "signal_date": "2026-08-20", "note": "金标准"},
            follow_redirects=True,
        )
        text = response.get_data(as_text=True)
        self.assertIn("已加入三日低吸池", text)
        self.assertIn("麦格米特", text)
        self.assertIn("8.3%", text)
        with self.module.db_connect() as db:
            sample = db.execute("SELECT * FROM alert_signal_samples").fetchone()
            outcome = db.execute("SELECT * FROM alert_signal_outcomes").fetchone()
        self.assertEqual(sample["review_status"], "CONFIRMED")
        self.assertEqual(sample["source"], "MANUAL")
        self.assertAlmostEqual(outcome["day_1_close_return"], 128.06 / 118.28 - 1)

        self.assertEqual(self.client.post(
            f"/admin/alert-types/three-day-dip/pool/{sample['id']}/review",
            data={"review_status": "REJECTED", "note": "复核否决"},
        ).status_code, 302)
        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT review_status FROM alert_signal_samples").fetchone()[0], "REJECTED")

    def test_fetch_realtime_prices_parses_tencent_response(self):
        fields = [""] * 88
        fields[1] = "通富微电"
        fields[2] = "002156"
        fields[3] = "11.50"
        fields[4] = "11.00"
        fields[5] = "11.10"
        fields[30] = "20260821150000"
        fields[33] = "12.00"
        fields[34] = "10.80"
        fields[35] = "11.50/1000/11500"
        fields[36] = "1000"
        fields[47] = "12.10"
        fields[48] = "9.90"
        payload = ('v_sz002156="' + "~".join(fields) + '";').encode("gbk")

        class Response:
            def __enter__(self):
                return self

            def __exit__(self, *args):
                return None

            def read(self):
                return payload

        with patch.object(self.module, "urlopen", return_value=Response()):
            quotes = self.module.fetch_realtime_prices(["002156"])

        quote = quotes["002156"]
        self.assertEqual(quote["trade_date"], "2026-08-21")
        self.assertEqual(quote["open"], 11.1)
        self.assertEqual(quote["high"], 12.0)
        self.assertEqual(quote["low"], 10.8)
        self.assertEqual(quote["close"], 11.5)
        self.assertEqual(quote["volume"], 1000.0)
        self.assertEqual(quote["amount"], 11500.0)
        self.assertEqual(quote["previous_close"], 11.0)
        self.assertEqual(quote["stock_name"], "通富微电")
        self.assertEqual(quote["limit_up"], 12.1)
        self.assertEqual(quote["limit_down"], 9.9)

    def test_fetch_realtime_prices_keeps_quote_when_limit_price_is_invalid(self):
        fields = [""] * 88
        fields[1], fields[2], fields[3], fields[4], fields[5] = "通富微电", "002156", "11.50", "11.00", "11.10"
        fields[30], fields[33], fields[34] = "20260821150000", "12.00", "10.80"
        fields[35], fields[36], fields[47] = "11.50/1000/11500", "1000", "--"
        payload = ('v_sz002156="' + "~".join(fields) + '";').encode("gbk")

        class Response:
            def __enter__(self):
                return self

            def __exit__(self, *args):
                return None

            def read(self):
                return payload

        with patch.object(self.module, "urlopen", return_value=Response()):
            quote = self.module.fetch_realtime_prices(["002156"])["002156"]
        self.assertEqual(quote["close"], 11.5)
        self.assertIsNone(quote["limit_up"])
        self.assertIsNone(quote["limit_down"])

    def test_watchlist_limit_up_alert_is_daily_and_user_scoped(self):
        now = "2026-08-21T10:15:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('limit-second', ?)", (now,)).lastrowid
            for user_id in (first_id, second_id):
                db.execute(
                    """INSERT INTO watchlist_stocks
                    (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                    VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                    (user_id, now, now),
                )

        quote = {
            "002156": {
                "trade_date": "2026-08-21", "quote_time": now, "open": 11.2, "high": 12.1,
                "low": 11.1, "close": 12.1, "previous_close": 11.0, "limit_up": 12.1,
                "volume": 1000, "amount": 12100, "fetched_at": now,
            },
        }
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_limit_up_alerts(db, quote), 2)
            self.assertEqual(self.module.evaluate_watchlist_limit_up_alerts(db, quote), 0)
            notifications = db.execute(
                """SELECT user_id, title, details_json FROM notifications
                ORDER BY user_id"""
            ).fetchall()
        self.assertEqual([row["user_id"] for row in notifications], [first_id, second_id])
        self.assertTrue(all(row["title"] == "通富微电达到涨停" for row in notifications))
        self.assertEqual(json.loads(notifications[0]["details_json"])["limit_up"], 12.1)

        next_day = {
            "002156": dict(
                quote["002156"], trade_date="2026-08-24", quote_time="2026-08-24T10:15:00",
                fetched_at="2026-08-24T10:15:00",
            ),
        }
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_limit_up_alerts(db, next_day), 2)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 4)

    def test_watchlist_limit_up_alert_requires_explicit_limit_price_and_watch(self):
        now = "2026-08-21T10:15:00"
        quote = {
            "002156": {
                "trade_date": "2026-08-21", "quote_time": now, "close": 12.1,
                "previous_close": 11.0, "limit_up": None, "fetched_at": now,
            },
            "600000": {
                "trade_date": "2026-08-21", "quote_time": now, "close": 11.0,
                "previous_close": 10.0, "limit_up": 11.0, "fetched_at": now,
            },
        }
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                (user_id, now, now),
            )
            self.assertEqual(self.module.evaluate_watchlist_limit_up_alerts(db, quote), 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 0)

    def test_watchlist_open_gain_alert_triggers_at_four_percent_once_per_day_and_user(self):
        now = "2026-08-21T10:15:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('gain-second', ?)", (now,)).lastrowid
            for user_id in (first_id, second_id):
                db.execute(
                    """INSERT INTO watchlist_stocks
                    (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                    VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                    (user_id, now, now),
                )

        below = {
            "002156": {
                "trade_date": "2026-08-21", "quote_time": now, "open": 9.8,
                "previous_close": 10.0, "close": 10.399,
                "fetched_at": now,
            },
        }
        threshold = {"002156": dict(below["002156"], close=10.4)}
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, below), 0)
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, threshold), 2)
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, threshold), 0)
            notifications = db.execute(
                "SELECT user_id, title, content, details_json FROM notifications ORDER BY user_id"
            ).fetchall()
        self.assertEqual([row["user_id"] for row in notifications], [first_id, second_id])
        self.assertTrue(all(row["title"] == "通富微电涨幅达到 4%" for row in notifications))
        self.assertIn("当日上涨 4.00%", notifications[0]["content"])
        self.assertAlmostEqual(json.loads(notifications[0]["details_json"])["gain_ratio"], 0.04)
        self.assertEqual(json.loads(notifications[0]["details_json"])["previous_close"], 10.0)

        next_day = {
            "002156": dict(
                threshold["002156"], trade_date="2026-08-24", quote_time="2026-08-24T10:15:00",
                fetched_at="2026-08-24T10:15:00",
            ),
        }
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, next_day), 2)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 4)

    def test_watchlist_open_gain_alert_requires_valid_previous_close_and_watch(self):
        now = "2026-08-21T10:15:00"
        quotes = {
            "002156": {
                "trade_date": "2026-08-21", "quote_time": now, "previous_close": 0, "close": 10.4,
                "fetched_at": now,
            },
            "600000": {
                "trade_date": "2026-08-21", "quote_time": now, "previous_close": 10.0, "close": 10.4,
                "fetched_at": now,
            },
        }
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                (user_id, now, now),
            )
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, quotes), 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 0)

    def test_watchlist_open_gain_alert_uses_saved_threshold_and_enabled_switch(self):
        now = "2026-08-21T10:15:00"
        quote = {"002156": {
            "trade_date": "2026-08-21", "quote_time": now, "previous_close": 10.0, "close": 10.5,
            "fetched_at": now,
        }}
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                (user_id, now, now),
            )
            db.execute(
                "UPDATE alert_types SET params_json = ?, enabled = 1 WHERE code = ?",
                (json.dumps({"threshold_ratio": 0.06}), self.module.WATCHLIST_OPEN_GAIN_CODE),
            )
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, quote), 0)
            db.execute(
                "UPDATE alert_types SET params_json = ?, enabled = 0 WHERE code = ?",
                (json.dumps({"threshold_ratio": 0.04}), self.module.WATCHLIST_OPEN_GAIN_CODE),
            )
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, quote), 0)
            db.execute("UPDATE alert_types SET enabled = 1 WHERE code = ?", (self.module.WATCHLIST_OPEN_GAIN_CODE,))
            self.assertEqual(self.module.evaluate_watchlist_open_gain_alerts(db, quote), 1)

    def test_watchlist_limit_down_alert_is_daily_and_user_scoped(self):
        now = "2026-08-21T10:15:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('limit-down-second', ?)", (now,)).lastrowid
            for user_id in (first_id, second_id):
                db.execute(
                    """INSERT INTO watchlist_stocks
                    (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                    VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                    (user_id, now, now),
                )
        quote = {"002156": {
            "trade_date": "2026-08-21", "quote_time": now, "close": 9.9,
            "previous_close": 11.0, "limit_down": 9.9, "fetched_at": now,
        }}
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_limit_down_alerts(db, quote), 2)
            self.assertEqual(self.module.evaluate_watchlist_limit_down_alerts(db, quote), 0)
            notifications = db.execute(
                "SELECT user_id, title, content, details_json FROM notifications ORDER BY user_id"
            ).fetchall()
        self.assertEqual([row["user_id"] for row in notifications], [first_id, second_id])
        self.assertTrue(all(row["title"] == "通富微电达到跌停" for row in notifications))
        self.assertIn("今日跌停价 9.90 元", notifications[0]["content"])
        self.assertEqual(json.loads(notifications[0]["details_json"])["limit_down"], 9.9)

        next_day = {"002156": dict(
            quote["002156"], trade_date="2026-08-24", quote_time="2026-08-24T10:15:00", fetched_at="2026-08-24T10:15:00",
        )}
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_limit_down_alerts(db, next_day), 2)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 4)

    def test_watchlist_limit_down_alert_requires_explicit_limit_price_and_watch(self):
        now = "2026-08-21T10:15:00"
        quotes = {
            "002156": {"trade_date": "2026-08-21", "quote_time": now, "close": 9.9, "limit_down": None, "fetched_at": now},
            "600000": {"trade_date": "2026-08-21", "quote_time": now, "close": 9.0, "limit_down": 9.0, "fetched_at": now},
        }
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                (user_id, now, now),
            )
            self.assertEqual(self.module.evaluate_watchlist_limit_down_alerts(db, quotes), 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 0)

    def test_watchlist_open_loss_alert_triggers_at_four_percent_once_per_day_and_user(self):
        now = "2026-08-21T10:15:00"
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("INSERT INTO users (name, created_at) VALUES ('loss-second', ?)", (now,)).lastrowid
            for user_id in (first_id, second_id):
                db.execute(
                    """INSERT INTO watchlist_stocks
                    (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                    VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                    (user_id, now, now),
                )

        below_threshold = {
            "002156": {
                "trade_date": "2026-08-21", "quote_time": now, "open": 10.1,
                "previous_close": 10.0, "close": 9.601, "fetched_at": now,
            },
        }
        threshold = {"002156": dict(below_threshold["002156"], close=9.6)}
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_open_loss_alerts(db, below_threshold), 0)
            self.assertEqual(self.module.evaluate_watchlist_open_loss_alerts(db, threshold), 2)
            self.assertEqual(self.module.evaluate_watchlist_open_loss_alerts(db, threshold), 0)
            notifications = db.execute(
                "SELECT user_id, title, content, details_json FROM notifications ORDER BY user_id"
            ).fetchall()
        self.assertEqual([row["user_id"] for row in notifications], [first_id, second_id])
        self.assertTrue(all(row["title"] == "通富微电跌幅达到 4%" for row in notifications))
        self.assertIn("当日下跌 4.00%", notifications[0]["content"])
        self.assertAlmostEqual(json.loads(notifications[0]["details_json"])["loss_ratio"], 0.04)
        self.assertEqual(json.loads(notifications[0]["details_json"])["previous_close"], 10.0)

        next_day = {
            "002156": dict(
                threshold["002156"], trade_date="2026-08-24", quote_time="2026-08-24T10:15:00",
                fetched_at="2026-08-24T10:15:00",
            ),
        }
        with self.module.db_connect() as db:
            self.assertEqual(self.module.evaluate_watchlist_open_loss_alerts(db, next_day), 2)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 4)

    def test_watchlist_open_loss_alert_requires_valid_previous_close_and_watch(self):
        now = "2026-08-21T10:15:00"
        quotes = {
            "002156": {
                "trade_date": "2026-08-21", "quote_time": now, "previous_close": 0, "close": 9.6,
                "fetched_at": now,
            },
            "600000": {
                "trade_date": "2026-08-21", "quote_time": now, "previous_close": 10.0, "close": 9.6,
                "fetched_at": now,
            },
        }
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '002156', '通富微电', 2, '', ?, ?)""",
                (user_id, now, now),
            )
            self.assertEqual(self.module.evaluate_watchlist_open_loss_alerts(db, quotes), 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM notifications").fetchone()[0], 0)

    def test_kline_volume_converts_standardized_realtime_shares_to_lots(self):
        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('000657', '2026-08-24', 69, 70, 64, 66.76, 89639600, 'tencent-realtime', '2026-08-24T15:03:06')"""
            )
            row = db.execute("SELECT volume, source FROM daily_prices WHERE stock_code = '000657'").fetchone()
        self.assertEqual(self.module.kline_volume_lots(row), 896396)

    def test_kline_volume_converts_historical_shares_to_lots(self):
        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('000657', '2026-08-24', 69, 70, 64, 66.76, 896396, 'akshare-tencent', '2026-08-24T15:03:06')"""
            )
            row = db.execute("SELECT volume, source FROM daily_prices WHERE stock_code = '000657'").fetchone()
        self.assertEqual(self.module.kline_volume_lots(row), 8963.96)

    def test_kline_api_declares_lot_volume_unit(self):
        now = "2026-08-24T15:03:06"
        with self.module.db_connect() as db:
            user_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            db.execute(
                """INSERT INTO watchlist_stocks
                (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                VALUES (?, '000657', '中钨高新', 2, '', ?, ?)""",
                (user_id, now, now),
            )
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('000657', '2026-08-24', 69, 70, 64, 66.76, 89639600,
                'tencent-realtime', ?)""",
                (now,),
            )
        payload = self.client.get("/analysis/watchlist/kline/000657").get_json()
        self.assertEqual(payload["volume_unit"], "lot")
        self.assertEqual(payload["data"][0]["volume"], 896396)

    def test_eastmoney_history_volume_is_normalized_from_lots_to_shares(self):
        frame = pd.DataFrame([{
            "日期": "2026-08-24", "开盘": 69, "最高": 70, "最低": 64, "收盘": 66.76,
            "成交量": 896396, "成交额": 600000000,
        }])
        with self.module.db_connect() as db:
            self.module.save_daily_prices(
                db, "000657", frame, "akshare-eastmoney", "2026-08-24", "2026-08-24",
            )
            row = db.execute("SELECT volume, source FROM daily_prices WHERE stock_code = '000657'").fetchone()
        self.assertEqual(row["volume"], 89639600)
        self.assertEqual(self.module.kline_volume_lots(row), 896396)

    def test_tencent_history_volume_is_normalized_only_when_returned_in_lots(self):
        lots_frame = pd.DataFrame([{
            "日期": "2026-08-24", "开盘": 69, "最高": 70, "最低": 64, "收盘": 66.76,
            "成交量": 896396, "成交额": 5992712800,
        }])
        shares_frame = pd.DataFrame([{
            "日期": "2026-08-25", "开盘": 69, "最高": 70, "最低": 64, "收盘": 66.76,
            "成交量": 89639600, "成交额": 5992712800,
        }])
        with self.module.db_connect() as db:
            self.module.save_daily_prices(
                db, "000657", lots_frame, "akshare-tencent", "2026-08-24", "2026-08-24",
            )
            self.module.save_daily_prices(
                db, "002156", shares_frame, "akshare-tencent", "2026-08-25", "2026-08-25",
            )
            rows = db.execute(
                "SELECT stock_code, volume FROM daily_prices ORDER BY stock_code"
            ).fetchall()
        self.assertEqual(dict(rows), {"000657": 89639600, "002156": 89639600})

    def test_daily_volume_migration_repairs_only_confirmed_tencent_lots_once(self):
        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
                VALUES ('000703', '2026-08-24', 19, 20, 18, 19, 1542976, 3006417500,
                'akshare-tencent', '2026-08-25T10:12:23')"""
            )
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
                VALUES ('002156', '2026-08-24', 19, 20, 18, 19, 154297600, 3006417500,
                'akshare-tencent', '2026-08-25T10:12:23')"""
            )
            db.execute("PRAGMA user_version = 17")

        self.module.init_db()
        self.module.init_db()

        with self.module.db_connect() as db:
            volumes = dict(db.execute(
                "SELECT stock_code, volume FROM daily_prices ORDER BY stock_code"
            ).fetchall())
        self.assertEqual(volumes, {"000703": 154297600, "002156": 154297600})

    def test_realtime_daily_volume_migration_reverses_only_confirmed_over_scaling_once(self):
        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
                VALUES ('688981', '2026-08-25', 118, 122, 117, 120, 3106434200, 3718701204,
                'tencent-realtime', '2026-08-25T15:00:30')"""
            )
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
                VALUES ('000703', '2026-08-25', 18, 19, 18, 18.34, 119116800, 2207488955,
                'tencent-realtime', '2026-08-25T15:00:30')"""
            )
            db.execute("PRAGMA user_version = 18")

        self.module.init_db()
        self.module.init_db()

        with self.module.db_connect() as db:
            volumes = dict(db.execute(
                "SELECT stock_code, volume FROM daily_prices ORDER BY stock_code"
            ).fetchall())
        self.assertEqual(volumes, {"000703": 119116800, "688981": 31064342})

    def test_realtime_daily_volume_trigger_corrects_legacy_writer_scaling(self):
        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
                VALUES ('688981', '2026-08-26', 120, 125, 120, 124, 3386636900, 4155943765,
                'tencent-realtime', '2026-08-26T14:22:15')"""
            )
            volume = db.execute(
                "SELECT volume FROM daily_prices WHERE stock_code = '688981'"
            ).fetchone()["volume"]
        self.assertEqual(volume, 33866369)

    def test_three_day_dip_volume_ratio_uses_standardized_daily_shares(self):
        prices = self._exhaustion_sample()
        for row in prices:
            row["source"] = "akshare-tencent"
        result = self.module.evaluate_three_day_dip(prices, enforce_volume=True)
        self.assertAlmostEqual(result["volume_ratio"], 29231900 / 31212500)

        quote = {"000657": {
            "trade_date": "2026-08-24", "open": 69, "high": 70, "low": 64, "close": 66.76,
            "volume": 896396, "amount": 5992712800, "fetched_at": "2026-08-24T15:03:06",
        }}
        with self.module.db_connect() as db:
            self.module.save_realtime_prices(db, quote)
            row = db.execute("SELECT volume, source FROM daily_prices WHERE stock_code = '000657'").fetchone()
        self.assertEqual(row["volume"], 89639600)

    def test_closing_quote_uses_market_quote_time(self):
        delayed = {
            "trade_date": "2026-08-24", "quote_time": "2026-08-24T14:59:30",
            "fetched_at": "2026-08-24T15:03:06",
        }
        self.assertFalse(self.module.is_closing_quote(delayed))
        self.assertTrue(self.module.is_closing_quote(dict(delayed, quote_time="2026-08-24T15:00:00")))
        self.assertFalse(self.module.is_closing_quote(dict(delayed, quote_time="2026-08-23T15:00:00")))

    def test_fetch_market_indexes_parses_index_quotes(self):
        fields = [""] * 88
        fields[2] = "sh000001"
        fields[3] = "3500.00"
        fields[4] = "3490.00"
        fields[5] = "3480.00"
        fields[30] = "20260821150000"
        fields[33] = "3510.00"
        fields[34] = "3470.00"
        fields[35] = "100000/1000/100000000"
        fields[36] = "1000000"
        payload = ('v_sh000001="' + "~".join(fields) + '";').encode("gbk")

        class Response:
            def __enter__(self):
                return self

            def __exit__(self, *args):
                return None

            def read(self):
                return payload

        with patch.object(self.module, "urlopen", return_value=Response()):
            indexes = self.module.fetch_market_indexes()

        self.assertEqual(len(indexes), 1)
        self.assertEqual(indexes[0]["name"], "上证指数")
        self.assertEqual(indexes[0]["price"], 3500)
        self.assertEqual(indexes[0]["change"], 10)
        self.assertAlmostEqual(indexes[0]["change_rate"], 10 / 3490)

    def test_fetch_market_indexes_parses_multiple_records_with_newlines(self):
        def record(code, close, previous):
            fields = [""] * 88
            fields[2] = code[2:]
            fields[3] = str(close)
            fields[4] = str(previous)
            fields[5] = str(previous)
            fields[30] = "20260821150000"
            fields[33] = str(close)
            fields[34] = str(previous)
            fields[35] = "100/1000/100000"
            fields[36] = "1000"
            return "v_" + code + '=\"' + "~".join(fields) + '\"'

        payload = (record("sh000001", 3500, 3490) + ";\n" + record("sz399001", 14000, 13900) + ";\n" + record("sz399006", 3500, 3400) + ";\n" + record("sh000688", 1600, 1590) + ";").encode("gbk")

        class Response:
            def __enter__(self):
                return self

            def __exit__(self, *args):
                return None

            def read(self):
                return payload

        with patch.object(self.module, "urlopen", return_value=Response()):
            indexes = self.module.fetch_market_indexes()

        self.assertEqual([item["code"] for item in indexes], list(self.module.MARKET_INDEXES)[:4])

    def test_portfolio_indexes_api_starts_first_refresh_in_background(self):
        with self.module.MARKET_INDEX_STATE_LOCK:
            self.module.MARKET_INDEX_STATE.update({"data": [], "updated_at": None, "error": None})
        with patch.object(self.module, "refresh_market_indexes_in_background") as refresh, \
             patch.object(self.module, "fetch_market_indexes") as fetch:
            response = self.client.get("/analysis/portfolio/indexes")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"], [])
        refresh.assert_called_once_with()
        fetch.assert_not_called()

    def test_portfolio_indexes_api_returns_cached_indexes_without_waiting_for_refresh(self):
        with self.module.MARKET_INDEX_STATE_LOCK:
            self.module.MARKET_INDEX_STATE.update({
                "data": [{"code": "sh000001", "name": "上证指数", "price": 3500, "change": 10,
                          "change_rate": 0.01, "trade_date": "2026-08-21", "fetched_at": "2026-08-21T15:00:00"}],
                "updated_at": datetime.now().isoformat(timespec="seconds"), "error": None,
            })
        with patch.object(self.module, "fetch_market_indexes") as fetch:
            response = self.client.get("/analysis/portfolio/indexes")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"][0]["name"], "上证指数")
        fetch.assert_not_called()

    def test_portfolio_indexes_api_returns_stale_cache_and_refreshes_in_background(self):
        with self.module.MARKET_INDEX_STATE_LOCK:
            self.module.MARKET_INDEX_STATE.update({
                "data": [{"code": "sh000001", "name": "上证指数", "price": 3500, "change": 10,
                          "change_rate": 0.01, "trade_date": "2026-08-21", "fetched_at": "2026-08-21T15:00:00"}],
                "updated_at": (datetime.now() - timedelta(seconds=31)).isoformat(timespec="seconds"),
                "error": None,
            })
        with patch.object(self.module, "refresh_market_indexes_in_background") as refresh, \
             patch.object(self.module, "fetch_market_indexes") as fetch:
            response = self.client.get("/analysis/portfolio/indexes")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"][0]["name"], "上证指数")
        refresh.assert_called_once_with()
        fetch.assert_not_called()

    def test_auto_sync_trading_windows(self):
        self.assertEqual(self.module.AUTO_REALTIME_SYNC_INTERVAL_SECONDS, 5)
        self.assertFalse(self.module.is_auto_sync_time(datetime(2026, 8, 24, 9, 14)))
        self.assertTrue(self.module.is_auto_sync_time(datetime(2026, 8, 24, 9, 15)))
        self.assertTrue(self.module.is_auto_sync_time(datetime(2026, 8, 24, 11, 30, 59)))
        self.assertFalse(self.module.is_auto_sync_time(datetime(2026, 8, 24, 11, 31)))
        self.assertFalse(self.module.is_auto_sync_time(datetime(2026, 8, 24, 12, 59)))
        self.assertTrue(self.module.is_auto_sync_time(datetime(2026, 8, 24, 13, 0)))
        self.assertTrue(self.module.is_auto_sync_time(datetime(2026, 8, 24, 15, 0, 59)))
        self.assertFalse(self.module.is_auto_sync_time(datetime(2026, 8, 24, 15, 1)))
        self.assertFalse(self.module.is_auto_sync_time(datetime(2026, 8, 23, 10, 0)))

    def test_auto_sync_deduplicates_all_users_stocks(self):
        now = "2026-08-24T09:00:00"
        with self.module.db_connect() as db:
            first_user = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_user = db.execute("INSERT INTO users (name, created_at) VALUES ('auto-second', ?)", (now,)).lastrowid
            for user_id, stock_code in ((first_user, "002156"), (second_user, "002156"), (second_user, "300308")):
                db.execute(
                    """INSERT INTO watchlist_stocks
                    (user_id, stock_code, stock_name, priority, note, created_at, updated_at)
                    VALUES (?, ?, ?, 2, '', ?, ?)""",
                    (user_id, stock_code, stock_code, now, now),
                )

        with patch.object(self.module, "sync_realtime_codes", return_value=(2, [])) as sync, \
             patch.object(self.module, "refresh_market_indexes_in_background") as refresh:
            result = self.module.run_auto_realtime_sync(datetime(2026, 8, 24, 10, 0))

        self.assertTrue(result)
        sync.assert_called_once_with(["002156", "300308"])
        refresh.assert_called_once_with()
        status = self.module.auto_sync_status()
        self.assertEqual(status["last_success_at"], "2026-08-24T10:00:00")
        self.assertEqual(status["last_synced"], 2)
        self.assertIsNone(status["last_error"])

    def test_retired_realtime_stream_stops_existing_event_source_clients(self):
        self.assertEqual(self.client.get("/api/realtime-stream?view=unknown").status_code, 400)
        for view in ("portfolio", "watchlist", "notifications"):
            response = self.client.get(f"/api/realtime-stream?view={view}")
            self.assertEqual(response.status_code, 204)
            self.assertEqual(response.get_data(), b"")

    def test_realtime_sync_publishes_update(self):
        quote = {
            "002156": {
                "trade_date": "2026-08-24", "quote_time": "2026-08-24T10:00:00",
                "open": 10.0, "high": 10.5, "low": 9.9, "close": 10.4,
                "previous_close": 10.0, "limit_up": 11.0, "volume": 1000,
                "amount": 10400, "fetched_at": "2026-08-24T10:00:00",
            },
        }
        with patch.object(self.module, "fetch_realtime_prices", return_value=quote), \
             patch.object(self.module, "publish_realtime_stream_update", wraps=self.module.publish_realtime_stream_update) as publish:
            self.module.sync_realtime_codes(["002156"])
        publish.assert_called_once_with()

    def test_auto_sync_skips_outside_window_and_empty_watchlist(self):
        with patch.object(self.module, "sync_realtime_codes") as sync:
            self.assertFalse(self.module.run_auto_realtime_sync(datetime(2026, 8, 24, 12, 0)))
            self.assertFalse(self.module.run_auto_realtime_sync(datetime(2026, 8, 23, 10, 0)))
            self.assertFalse(self.module.run_auto_realtime_sync(datetime(2026, 8, 24, 10, 0)))
        sync.assert_not_called()

    def test_watchlist_kline_only_reads_saved_history(self):
        with patch.object(self.module, "sync_history_codes", return_value=(1, [])):
            self.client.post(
                "/analysis/watchlist",
                data={"stock_code": "300308", "stock_name": "中际旭创", "priority": "2"},
            )
        with self.module.db_connect() as db:
            db.execute(
                """INSERT INTO daily_prices
                (stock_code, trade_date, open, high, low, close, volume, source, fetched_at)
                VALUES ('300308', '2026-08-20', 10, 12, 9, 11, 1000, 'test', '2026-08-21T15:00:00')"""
            )
        with patch.object(self.module, "fetch_daily_prices") as fetch:
            response = self.client.get("/analysis/watchlist/kline/300308")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"][0]["date"], "2026-08-20")
        fetch.assert_not_called()

    def test_strategy_filters_and_backtest(self):
        self.client.post("/admin/preview", data={"statement": (self.make_statement(), "statement.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
        self.client.post(f"/admin/import/{job['id']}")

        strategy = self.client.get("/analysis/strategy?min_days=2&max_days=2")
        self.assertEqual(strategy.status_code, 200)
        text = strategy.get_data(as_text=True)
        self.assertIn("持仓 T+2 至 T+2", text)
        self.assertIn("命中样本", text)
        self.assertIn("1", text)
        self.assertIn("002156", text)

        invalid = self.client.get("/analysis/strategy?buy_session=invalid&min_days=9&max_days=2&period=17")
        self.assertEqual(invalid.status_code, 200)
        invalid_text = invalid.get_data(as_text=True)
        self.assertIn("持仓 T+2 至 T+9", invalid_text)
        self.assertIn('<option value="0" selected>全部历史</option>', invalid_text)
        self.assertIn('<option value="all">不限</option>', invalid_text)

    def test_local_users_isolate_imports_and_fifo(self):
        first_user = self.client.get("/admin/statements")
        self.assertIn("yutaoGS", first_user.get_data(as_text=True))
        self.client.post("/users/create", data={"name": "second", "next": "/admin/statements"})
        with self.module.db_connect() as db:
            first_id = db.execute("SELECT id FROM users WHERE name = 'yutaoGS'").fetchone()["id"]
            second_id = db.execute("SELECT id FROM users WHERE name = 'second'").fetchone()["id"]

        self.client.post("/users/switch", data={"user_id": first_id, "next": "/admin/statements"})
        self.client.post("/admin/preview", data={"statement": (self.make_statement(), "first.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            first_job = db.execute("SELECT id FROM import_jobs WHERE user_id = ?", (first_id,)).fetchone()["id"]
        self.client.post(f"/admin/import/{first_job}")

        self.client.post("/users/switch", data={"user_id": second_id, "next": "/admin/statements"})
        second_page = self.client.get("/admin/statements").get_data(as_text=True)
        self.assertIn("共 0 条成交记录", second_page)
        self.assertEqual(self.client.post(f"/admin/import/{first_job}").status_code, 404)
        preview = self.client.post("/admin/preview", data={"statement": (self.make_statement(), "second.xlsx")}, content_type="multipart/form-data")
        preview_text = preview.get_data(as_text=True)
        self.assertIn("数据库重复", preview_text)
        self.assertIn("预计新增", preview_text)
        with self.module.db_connect() as db:
            second_job = db.execute("SELECT id FROM import_jobs WHERE user_id = ?", (second_id,)).fetchone()["id"]
        self.client.post(f"/admin/import/{second_job}")

        with self.module.db_connect() as db:
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions WHERE user_id = ?", (first_id,)).fetchone()[0], 3)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM executions WHERE user_id = ?", (second_id,)).fetchone()[0], 3)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM fifo_matches WHERE user_id = ?", (first_id,)).fetchone()[0], 2)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM fifo_matches WHERE user_id = ?", (second_id,)).fetchone()[0], 2)
            self.assertEqual(db.execute("PRAGMA foreign_key_check").fetchall(), [])

    def test_user_management_validation(self):
        response = self.client.post("/users/create", data={"name": "yutaogs", "next": "https://example.com"})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/analysis/stocks")
        self.assertIn("用户名称已存在", self.client.get("/analysis/stocks").get_data(as_text=True))

        response = self.client.post("/users/create", data={"name": "second", "next": "/analysis/summary"})
        self.assertEqual(response.headers["Location"], "/analysis/summary")
        response = self.client.post("/users/999999/rename", data={"name": "missing", "next": "/analysis/summary"}, follow_redirects=True)
        self.assertIn("用户不存在", response.get_data(as_text=True))

    def test_import_errors_are_visible(self):
        response = self.client.post("/admin/preview", data={}, follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("请选择交割单文件", response.get_data(as_text=True))

        response = self.client.post(
            "/admin/preview",
            data={"statement": (io.BytesIO(b"not a statement"), "statement.pdf")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("仅支持 .xlsx、.xls、.csv、.txt 文件", response.get_data(as_text=True))

    def test_invalid_trade_error_shows_values_read_from_file(self):
        rows = [
            ["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交价格"],
            [20260820, "13:19:01", 300209, "行云科技", "证券买入", 300, ""],
        ]
        parsed, errors, ignored, total_rows = self.module.parse_rows(rows)
        self.assertEqual(total_rows, 1)
        self.assertEqual(parsed, [])
        self.assertEqual(ignored, [])
        self.assertIn("数量=300", errors[0])
        self.assertIn("价格=", errors[0])
        self.assertIn("操作=证券买入", errors[0])

    def test_non_trade_rows_are_ignored_instead_of_failed(self):
        rows = [
            ["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交价格"],
            [20260819, "19:42:00", 301697, "贝特利", "申购配号", 2, 0],
            [20260818, "09:00:00", 600584, "长电科技", "股息入账", 0, 42.82],
        ]
        parsed, errors, ignored, total_rows = self.module.parse_rows(rows)
        self.assertEqual(total_rows, 2)
        self.assertEqual(parsed, [])
        self.assertEqual(errors, [])
        self.assertEqual(len(ignored), 2)
        self.assertIn("申购配号", ignored[0])

    def test_stock_detail(self):
        self.client.post("/admin/preview", data={"statement": (self.make_statement(), "statement.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW' ORDER BY created_at DESC, rowid DESC LIMIT 1").fetchone()
        self.client.post(f"/admin/import/{job['id']}")
        self.client.post(
            "/admin/manual",
            data={
                "trade_date": "2026-08-04", "trade_time": "10:00:00",
                "stock_code": "002156", "stock_name": "通富微电", "action": "SELL",
                "quantity": "50", "deal_price": "14", "commission": "0",
                "stamp_tax": "0", "transfer_fee": "0",
            },
        )

        stocks = self.client.get("/analysis/stocks")
        self.assertEqual(stocks.status_code, 200)
        text = stocks.get_data(as_text=True)
        self.assertIn("002156", text)
        self.assertIn("通富微电", text)
        self.assertIn("2026-08-04", text)
        self.assertIn("595.00", text)

        detail = self.client.get("/analysis/stocks?code=002156")
        self.assertEqual(detail.status_code, 200)
        detail_text = detail.get_data(as_text=True)
        self.assertIn("08-01 09:31", detail_text)
        self.assertIn("08-04 10:00", detail_text)
        self.assertIn("总交易笔数", detail_text)
        self.assertIn("交易明细", detail_text)
        self.assertNotIn("<svg", detail_text)
        self.assertNotIn("查看复盘", detail_text)
        self.assertNotIn("FIFO 成本匹配明细", detail_text)

    def test_stock_list_detailed_cycle_metrics(self):
        self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260803, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S1", 13],
        ])
        self._import_workbook_rows([
            [20260805, "09:30:00", 2156, "通富微电", "证券买入", 100, "B2", 10],
            [20260806, "10:00:00", 2156, "通富微电", "证券卖出", -100, "S2", 8],
        ])

        with self.module.db_connect() as db:
            stock = self.module.analysis_data(db)["by_stock"][0]
        self.assertEqual(stock["trades"], 2)
        self.assertAlmostEqual(stock["profit"], 100.0)
        self.assertAlmostEqual(stock["profit_rate"], 0.05)
        self.assertAlmostEqual(stock["win_rate"], 0.5)
        self.assertAlmostEqual(stock["max_profit"], 300.0)
        self.assertAlmostEqual(stock["max_loss"], -200.0)
        self.assertAlmostEqual(stock["avg_holding_days"], 1.5)

        page = self.client.get("/analysis/stocks").get_data(as_text=True)
        for label in ("累计盈亏", "累计收益率", "交易总次数", "胜率", "最大盈利", "最大亏损", "平均持仓天数"):
            self.assertIn(label, page)
        self.assertIn("+100.00 元", page)
        self.assertIn("+5.0%", page)
        self.assertIn("2 笔", page)
        self.assertIn("50%", page)
        self.assertIn("+300.00", page)
        self.assertIn("-200.00", page)
        self.assertIn("1.5 天", page)

        self.assertIn("总交易笔数", page)
        self.assertIn("盈亏比", page)
        self.assertIn("1.50", page)
        self.assertIn("+100.00", page)
        self.assertIn("+30.00%", page)
        self.assertIn("-20.00%", page)
        self.assertLess(page.index("08-06 10:00"), page.index("08-03 10:00"))

    def test_stock_detail_same_day_and_no_loss_profit_factor(self):
        self._import_workbook_rows([
            [20260801, "09:30:00", 2156, "通富微电", "证券买入", 100, "B1", 10],
            [20260801, "14:30:00", 2156, "通富微电", "证券卖出", -100, "S1", 11],
        ])
        page = self.client.get("/analysis/stocks?code=002156").get_data(as_text=True)
        self.assertIn("08-01 09:30", page)
        self.assertIn("08-01 14:30", page)
        self.assertIn("&lt;1天", page)
        self.assertIn("--", page)

    def test_stock_list_sorting(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["成交日期", "成交时间", "证券代码", "证券名称", "操作", "成交数量", "成交编号", "成交价格", "成交金额"])
        sheet.append([20260801, "09:30:00", 100001, "较早盈利", "证券买入", 100, "A1", 10, 1000])
        sheet.append([20260802, "10:00:00", 100001, "较早盈利", "证券卖出", -100, "A2", 20, 2000])
        sheet.append([20260810, "09:30:00", 100002, "最近亏损", "证券买入", 100, "B1", 20, 2000])
        sheet.append([20260811, "10:00:00", 100002, "最近亏损", "证券卖出", -100, "B2", 15, 1500])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        self.client.post("/admin/preview", data={"statement": (output, "sorting.xlsx")}, content_type="multipart/form-data")
        with self.module.db_connect() as db:
            job = db.execute("SELECT id FROM import_jobs WHERE status = 'PREVIEW'").fetchone()
        self.client.post(f"/admin/import/{job['id']}")

        recent = self.client.get("/analysis/stocks").get_data(as_text=True)
        self.assertLess(recent.index("最近亏损"), recent.index("较早盈利"))
        self.assertIn('<option value="recent" selected>最近交易</option>', recent)
        self.assertIn("2026-08-11", recent)

        profit = self.client.get("/analysis/stocks?sort=profit").get_data(as_text=True)
        self.assertLess(profit.index("较早盈利"), profit.index("最近亏损"))
        self.assertIn('<option value="profit" selected>盈利最高</option>', profit)
        self.assertIn("sort=profit", profit)

        loss = self.client.get("/analysis/stocks?sort=loss").get_data(as_text=True)
        self.assertLess(loss.index("最近亏损"), loss.index("较早盈利"))

        invalid = self.client.get("/analysis/stocks?sort=invalid").get_data(as_text=True)
        self.assertLess(invalid.index("最近亏损"), invalid.index("较早盈利"))
        self.assertIn('<option value="recent" selected>最近交易</option>', invalid)

        by_name = self.client.get("/analysis/stocks?q=较早&sort=profit").get_data(as_text=True)
        self.assertIn("较早盈利", by_name)
        self.assertNotIn("最近亏损", by_name)
        self.assertIn("找到 1 只股票", by_name)
        self.assertIn("q=%E8%BE%83%E6%97%A9", by_name)
        self.assertIn("sort=profit", by_name)

        by_code = self.client.get("/analysis/stocks?q=100002").get_data(as_text=True)
        self.assertIn("最近亏损", by_code)
        self.assertNotIn("较早盈利", by_code)

        partial_code = self.client.get("/analysis/stocks?q=0002").get_data(as_text=True)
        self.assertIn("最近亏损", partial_code)

        empty = self.client.get("/analysis/stocks?q=不存在").get_data(as_text=True)
        self.assertIn("没有匹配名称或代码的股票", empty)



if __name__ == "__main__":
    unittest.main()
