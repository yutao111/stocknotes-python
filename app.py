from __future__ import annotations

import csv
import calendar
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import io
import json
import math
import os
import sqlite3
import ssl
import tempfile
import threading
import time
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.request import Request, urlopen

import certifi

from flask import Flask, abort, flash, g, has_request_context, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
import xlrd


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.environ.get("STOCKNOTES_DB", BASE_DIR / "stocknotes.db"))
ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv", "txt"}
IGNORED_ACTIONS = {"申购配号", "股息入账", "股息红利税补", "指定交易"}
STATEMENTS_PER_PAGE = 20
SCHEMA_VERSION = 11
SSL_CONTEXT = ssl.create_default_context(cafile=certifi.where())
DEFAULT_USER_NAME = "yutaoGS"
MAX_USER_NAME_LENGTH = 50
EXECUTION_SOURCES = ("STATEMENT", "BROKER_TODAY", "MANUAL")
REVIEW_STATUSES = ("PENDING", "COMPLETED")
REVIEW_REASON_TYPES = (
    "TREND_BREAKOUT", "PULLBACK", "EVENT_DRIVEN", "EARNINGS_DRIVEN",
    "VALUATION", "SECTOR_CONFIRMATION", "TECHNICAL_PATTERN", "SENTIMENT", "OTHER",
)
REVIEW_JUDGEMENT_RESULTS = ("CORRECT", "PARTIAL", "WRONG")
PRICE_SYNC_INITIAL_DAYS = 250
PRICE_SYNC_OVERLAP_DAYS = 7
PRICE_SYNC_WORKERS = 4
AUTO_REALTIME_SYNC_INTERVAL_SECONDS = 30
REALTIME_SYNC_LOCK = threading.Lock()
AUTO_SYNC_STOP = threading.Event()
AUTO_SYNC_STATE_LOCK = threading.Lock()
AUTO_SYNC_START_LOCK = threading.Lock()
AUTO_SYNC_THREAD = None
AUTO_SYNC_STATE = {
    "running": False,
    "last_attempt_at": None,
    "last_success_at": None,
    "last_synced": 0,
    "last_error": None,
    "indexes_last_success_at": None,
    "indexes_last_error": None,
}
MARKET_INDEX_STATE_LOCK = threading.Lock()
MARKET_INDEX_STATE = {"data": [], "updated_at": None, "error": None}
HOT_SECTORS_CACHE_LOCK = threading.Lock()
HOT_SECTORS_CACHE = {"modules": {}, "updated_at": None}
HOT_SECTORS_CACHE_TTL_SECONDS = 30
THREE_DAY_DIP_CODE = "THREE_DAY_DIP"
THREE_DAY_DIP_DEFAULT_PARAMS = {
    "decline_days": 2,
    "require_bearish_candles": True,
    "require_declining_closes": True,
    "max_signal_low_above_prior_ratio": 0.04,
    "min_decline_ratio": 0.08,
    "max_volume_ratio": 1.0,
    "exhaustion_min_repair_ratio": 0.02,
    "exhaustion_max_body_range_ratio": 0.10,
    "exhaustion_min_change_ratio": -0.015,
    "exhaustion_max_change_ratio": 0.03,
    "shadow_min_repair_ratio": 0.015,
    "shadow_max_body_range_ratio": 0.20,
    "shadow_min_lower_shadow_body_ratio": 1.0,
    "shadow_min_change_ratio": -0.015,
    "shadow_max_change_ratio": 0.03,
    "reversal_min_decline_ratio": 0.10,
    "reversal_min_repair_ratio": 0.06,
    "reversal_min_recovery_range_ratio": 0.60,
    "reversal_min_change_ratio": 0.04,
    "reversal_max_volume_ratio": 1.0,
    "intraday_candidate_enabled": True,
    "close_confirmation_enabled": True,
}
THREE_DAY_DIP_RULE_VERSION = "three-day-dip-v6"
INTRADAY_REBOUND_CODE = "INTRADAY_REBOUND"
INTRADAY_REBOUND_DEFAULT_PARAMS = {
    "lookback_minutes": 120,
    "min_drop_ratio": 0.02,
    "min_rebound_ratio": 0.008,
    "min_trough_age_minutes": 5,
    "min_volume_multiple": 1.5,
    "enabled": True,
}
REVIEW_MAIN_PROBLEMS = (
    "STOCK_SELECTION", "ENTRY_TIMING", "EXIT_TIMING", "POSITION_SIZE",
    "STOP_LOSS", "EARLY_PROFIT_TAKING", "CHASE", "BOTTOM_FISHING",
    "EMOTIONAL", "PLAN_VIOLATION", "PLAN_ERROR", "EXTERNAL_EVENT", "NO_OBVIOUS_PROBLEM",
)
REVIEW_REASON_TYPE_LABELS = {
    "TREND_BREAKOUT": "趋势突破",
    "PULLBACK": "回调买入",
    "EVENT_DRIVEN": "事件驱动",
    "EARNINGS_DRIVEN": "业绩驱动",
    "VALUATION": "估值",
    "SECTOR_CONFIRMATION": "板块共振",
    "TECHNICAL_PATTERN": "技术形态",
    "SENTIMENT": "情绪博弈",
    "OTHER": "其他",
}
REVIEW_JUDGEMENT_LABELS = {"CORRECT": "完全正确", "PARTIAL": "部分正确", "WRONG": "错误"}
REVIEW_MAIN_PROBLEM_LABELS = {
    "STOCK_SELECTION": "选股错误",
    "ENTRY_TIMING": "买点错误",
    "EXIT_TIMING": "卖点错误",
    "POSITION_SIZE": "仓位错误",
    "STOP_LOSS": "止损错误",
    "EARLY_PROFIT_TAKING": "止盈过早",
    "CHASE": "追涨",
    "BOTTOM_FISHING": "抄底",
    "EMOTIONAL": "情绪化交易",
    "PLAN_VIOLATION": "没有按照计划执行",
    "PLAN_ERROR": "交易计划本身错误",
    "EXTERNAL_EVENT": "外部事件",
    "NO_OBVIOUS_PROBLEM": "无明显问题",
}
STOCK_SORTS = {
    "recent": "last_trade_at DESC, profit DESC, episodes.stock_code",
    "profit": "profit DESC, last_trade_at DESC, episodes.stock_code",
    "loss": "profit ASC, last_trade_at DESC, episodes.stock_code",
    "trades": "trades DESC, last_trade_at DESC, episodes.stock_code",
    "win_rate": "win_rate DESC, trades DESC, last_trade_at DESC, episodes.stock_code",
}

FIELDS = [
    "trade_date", "trade_time", "stock_code", "stock_name", "action",
    "raw_quantity", "deal_id", "deal_price", "deal_amount", "balance",
    "stock_balance", "cash_change", "commission", "stamp_tax", "other_fee",
    "cash_balance", "current_amount", "contract_id", "extra_fee",
    "transfer_fee", "market",
]

HEADER_ALIASES = {
    "成交日期": "trade_date", "日期": "trade_date",
    "成交时间": "trade_time", "时间": "trade_time",
    "证券代码": "stock_code", "股票代码": "stock_code",
    "证券名称": "stock_name", "股票名称": "stock_name",
    "操作": "action", "业务名称": "action", "买卖标志": "action",
    "成交数量": "raw_quantity", "数量": "raw_quantity",
    "成交编号": "deal_id", "成交序号": "deal_id",
    "成交价格": "deal_price", "成交价": "deal_price", "成交均价": "deal_price",
    "成交金额": "deal_amount", "发生金额": "cash_change",
    "余额": "balance", "股票余额": "stock_balance", "证券余额": "stock_balance",
    "佣金": "commission", "印花税": "stamp_tax", "其他杂费": "other_fee",
    "资金余额": "cash_balance", "本次金额": "current_amount",
    "合同编号": "contract_id", "合同号": "contract_id", "其他费": "extra_fee",
    "过户费": "transfer_fee", "交易市场": "market", "市场": "market",
}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "stocknotes-local-only")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024


def db_connect() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        db.execute("PRAGMA foreign_keys = OFF")
        db.execute("BEGIN IMMEDIATE")
        db.execute(
            """CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL COLLATE NOCASE UNIQUE CHECK(name = TRIM(name) AND LENGTH(name) BETWEEN 1 AND 50),
                created_at TEXT NOT NULL
            )"""
        )
        default_user = db.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
        if default_user is None:
            cursor = db.execute(
                "INSERT INTO users (name, created_at) VALUES (?, ?)",
                (DEFAULT_USER_NAME, datetime.now().isoformat(timespec="seconds")),
            )
            default_user_id = cursor.lastrowid
        else:
            default_user_id = default_user["id"]

        business_tables = ("import_jobs", "executions", "fifo_matches", "positions", "unmatched_sells")
        existing_tables = {
            row["name"] for row in db.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name IN (?, ?, ?, ?, ?)",
                business_tables,
            )
        }
        needs_migration = any(
            "user_id" not in {column["name"] for column in db.execute(f"PRAGMA table_info({table})")}
            for table in existing_tables
        )
        legacy_columns = {}
        if needs_migration:
            db.execute("DROP INDEX IF EXISTS idx_executions_time")
            for table in reversed(business_tables):
                if table in existing_tables:
                    legacy_columns[table] = {column["name"] for column in db.execute(f"PRAGMA table_info({table})")}
                    db.execute(f"ALTER TABLE {table} RENAME TO __legacy_{table}")

        signal_pool_sql = db.execute(
            "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'alert_signal_samples'"
        ).fetchone()
        migrate_signal_pool = signal_pool_sql is not None and "SHADOW_STOP" not in signal_pool_sql["sql"]
        if migrate_signal_pool:
            db.execute("DROP INDEX IF EXISTS idx_alert_signal_samples_user_date")
            db.execute("ALTER TABLE alert_signal_outcomes RENAME TO __legacy_alert_signal_outcomes")
            db.execute("ALTER TABLE alert_signal_samples RENAME TO __legacy_alert_signal_samples")

        schema_statements = (
            """CREATE TABLE IF NOT EXISTS import_jobs (
                id TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                status TEXT NOT NULL,
                total_rows INTEGER NOT NULL,
                valid_rows INTEGER NOT NULL,
                invalid_rows INTEGER NOT NULL,
                duplicate_rows INTEGER NOT NULL DEFAULT 0,
                payload TEXT,
                errors TEXT,
                created_at TEXT NOT NULL,
                imported_at TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS executions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                import_job_id TEXT NOT NULL REFERENCES import_jobs(id) ON DELETE CASCADE,
                fingerprint TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                trade_time TEXT NOT NULL,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                action TEXT NOT NULL CHECK(action IN ('BUY', 'SELL')),
                raw_quantity REAL NOT NULL,
                quantity REAL NOT NULL,
                deal_id TEXT,
                deal_price REAL NOT NULL,
                deal_amount REAL NOT NULL,
                commission REAL NOT NULL DEFAULT 0,
                stamp_tax REAL NOT NULL DEFAULT 0,
                other_fee REAL NOT NULL DEFAULT 0,
                extra_fee REAL NOT NULL DEFAULT 0,
                transfer_fee REAL NOT NULL DEFAULT 0,
                market TEXT,
                raw_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, fingerprint)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_executions_time
            ON executions(user_id, stock_code, trade_date, trade_time, id)""",
            """CREATE TABLE IF NOT EXISTS fifo_matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                buy_execution_id INTEGER NOT NULL REFERENCES executions(id),
                sell_execution_id INTEGER NOT NULL REFERENCES executions(id),
                buy_date TEXT NOT NULL,
                sell_date TEXT NOT NULL,
                quantity REAL NOT NULL,
                buy_cost REAL NOT NULL,
                sell_income REAL NOT NULL,
                profit REAL NOT NULL,
                profit_rate REAL NOT NULL,
                holding_days INTEGER NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS positions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                buy_execution_id INTEGER NOT NULL REFERENCES executions(id),
                buy_date TEXT NOT NULL,
                quantity REAL NOT NULL,
                unit_cost REAL NOT NULL,
                total_cost REAL NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS unmatched_sells (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                sell_execution_id INTEGER NOT NULL REFERENCES executions(id),
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                sell_date TEXT NOT NULL,
                quantity REAL NOT NULL,
                sell_income REAL NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS trade_episodes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                opening_execution_id INTEGER NOT NULL REFERENCES executions(id) ON DELETE CASCADE,
                closing_execution_id INTEGER REFERENCES executions(id) ON DELETE SET NULL,
                opened_at TEXT NOT NULL,
                closed_at TEXT,
                status TEXT NOT NULL CHECK(status IN ('OPEN', 'CLOSED')),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(user_id, opening_execution_id)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_trade_episodes_stock
            ON trade_episodes(user_id, stock_code, status, opened_at)""",
            """CREATE TABLE IF NOT EXISTS trade_episode_executions (
                trade_episode_id INTEGER NOT NULL REFERENCES trade_episodes(id) ON DELETE CASCADE,
                execution_id INTEGER NOT NULL REFERENCES executions(id) ON DELETE CASCADE,
                role TEXT NOT NULL CHECK(role IN ('ENTRY', 'ADD', 'REDUCE', 'EXIT')),
                attributed_quantity REAL NOT NULL CHECK(attributed_quantity > 0),
                PRIMARY KEY(trade_episode_id, execution_id)
            )""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_episode_execution
            ON trade_episode_executions(execution_id)""",
            """CREATE TABLE IF NOT EXISTS trade_reviews (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                trade_episode_id INTEGER REFERENCES trade_episodes(id) ON DELETE SET NULL,
                review_status TEXT NOT NULL DEFAULT 'PENDING' CHECK(review_status IN ('PENDING', 'COMPLETED')),
                trade_reason TEXT NOT NULL DEFAULT '',
                expected_profit_percent REAL,
                expected_target_price REAL,
                stop_loss_price REAL,
                expected_holding_days INTEGER,
                confidence_level INTEGER CHECK(confidence_level IS NULL OR confidence_level BETWEEN 1 AND 5),
                sell_reason TEXT NOT NULL DEFAULT '',
                judgement_result TEXT CHECK(judgement_result IS NULL OR judgement_result IN ('CORRECT', 'PARTIAL', 'WRONG')),
                main_problem TEXT CHECK(main_problem IS NULL OR main_problem IN (
                    'STOCK_SELECTION', 'ENTRY_TIMING', 'EXIT_TIMING', 'POSITION_SIZE',
                    'STOP_LOSS', 'EARLY_PROFIT_TAKING', 'CHASE', 'BOTTOM_FISHING',
                    'EMOTIONAL', 'PLAN_VIOLATION', 'PLAN_ERROR', 'EXTERNAL_EVENT', 'NO_OBVIOUS_PROBLEM'
                )),
                review_note TEXT NOT NULL DEFAULT '',
                next_action TEXT NOT NULL DEFAULT '',
                original_opening_execution_id INTEGER,
                stock_code_snapshot TEXT NOT NULL DEFAULT '',
                opened_at_snapshot TEXT,
                closed_at_snapshot TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                completed_at TEXT,
                CHECK(expected_profit_percent IS NULL OR expected_profit_percent >= 0),
                CHECK(expected_target_price IS NULL OR expected_target_price > 0),
                CHECK(stop_loss_price IS NULL OR stop_loss_price > 0),
                CHECK(expected_holding_days IS NULL OR expected_holding_days >= 0)
            )""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_trade_reviews_episode
            ON trade_reviews(user_id, trade_episode_id) WHERE trade_episode_id IS NOT NULL""",
            """CREATE INDEX IF NOT EXISTS idx_trade_reviews_user_status
            ON trade_reviews(user_id, review_status, updated_at DESC)""",
            """CREATE TABLE IF NOT EXISTS trade_review_reason_types (
                trade_review_id INTEGER NOT NULL REFERENCES trade_reviews(id) ON DELETE CASCADE,
                reason_type TEXT NOT NULL CHECK(reason_type IN (
                    'TREND_BREAKOUT', 'PULLBACK', 'EVENT_DRIVEN', 'EARNINGS_DRIVEN',
                    'VALUATION', 'SECTOR_CONFIRMATION', 'TECHNICAL_PATTERN', 'SENTIMENT', 'OTHER'
                )),
                PRIMARY KEY(trade_review_id, reason_type)
            )""",
            """CREATE TABLE IF NOT EXISTS daily_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stock_code TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                open REAL,
                high REAL,
                low REAL,
                close REAL,
                volume REAL,
                amount REAL,
                source TEXT NOT NULL,
                fetched_at TEXT NOT NULL,
                UNIQUE(stock_code, trade_date)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_daily_prices_code_date
            ON daily_prices(stock_code, trade_date)""",
            """CREATE TABLE IF NOT EXISTS intraday_quotes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stock_code TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                quote_minute TEXT NOT NULL,
                price REAL NOT NULL CHECK(price > 0),
                day_high REAL,
                day_low REAL,
                previous_close REAL,
                volume REAL,
                amount REAL,
                source TEXT NOT NULL,
                fetched_at TEXT NOT NULL,
                UNIQUE(stock_code, quote_minute)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_intraday_quotes_code_date_minute
            ON intraday_quotes(stock_code, trade_date, quote_minute)""",
            """CREATE TABLE IF NOT EXISTS trade_excursion_metrics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trade_episode_id INTEGER NOT NULL REFERENCES trade_episodes(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                mfe REAL,
                mae REAL,
                highest_price REAL,
                lowest_price REAL,
                max_gain_amount REAL,
                capture_rate REAL,
                buy_price REAL,
                sell_price REAL,
                quantity REAL,
                data_start_date TEXT,
                data_end_date TEXT,
                price_source TEXT,
                calculated_at TEXT NOT NULL,
                UNIQUE(trade_episode_id)
            )""",
            """CREATE TABLE IF NOT EXISTS account_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                snapshot_date TEXT NOT NULL,
                total_assets REAL NOT NULL CHECK(total_assets > 0),
                available_cash REAL CHECK(available_cash IS NULL OR available_cash >= 0),
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(user_id, snapshot_date)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_account_snapshots_user_date
            ON account_snapshots(user_id, snapshot_date DESC, id DESC)""",
            """CREATE TABLE IF NOT EXISTS watchlist_stocks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                priority INTEGER NOT NULL DEFAULT 2 CHECK(priority BETWEEN 1 AND 3),
                note TEXT NOT NULL DEFAULT '' CHECK(LENGTH(note) <= 500),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(user_id, stock_code)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_watchlist_stocks_user_priority
            ON watchlist_stocks(user_id, priority DESC, updated_at DESC, id DESC)""",
            """CREATE TABLE IF NOT EXISTS alert_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                signal_side TEXT NOT NULL CHECK(signal_side IN ('BUY', 'SELL', 'NEUTRAL')),
                description TEXT NOT NULL DEFAULT '',
                params_json TEXT NOT NULL,
                sample_json TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1 CHECK(enabled IN (0, 1)),
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )""",
            """CREATE TABLE IF NOT EXISTS alert_rule_states (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                alert_type_id INTEGER NOT NULL REFERENCES alert_types(id) ON DELETE CASCADE,
                stock_code TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                candidate_triggered_at TEXT,
                confirmation_triggered_at TEXT,
                last_evaluated_at TEXT NOT NULL,
                last_matched INTEGER NOT NULL DEFAULT 0 CHECK(last_matched IN (0, 1)),
                UNIQUE(user_id, alert_type_id, stock_code, trade_date)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_alert_rule_states_lookup
            ON alert_rule_states(alert_type_id, stock_code, trade_date)""",
            """CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                alert_type_id INTEGER REFERENCES alert_types(id) ON DELETE SET NULL,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                stage TEXT NOT NULL CHECK(stage IN ('CANDIDATE', 'CONFIRMED')),
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                details_json TEXT NOT NULL DEFAULT '{}',
                quote_time TEXT NOT NULL,
                created_at TEXT NOT NULL,
                read_at TEXT,
                dedupe_key TEXT NOT NULL UNIQUE
            )""",
            """CREATE INDEX IF NOT EXISTS idx_notifications_user_unread
            ON notifications(user_id, read_at, created_at DESC, id DESC)""",
            """CREATE TABLE IF NOT EXISTS alert_signal_samples (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                alert_type_id INTEGER NOT NULL REFERENCES alert_types(id) ON DELETE CASCADE,
                stock_code TEXT NOT NULL,
                stock_name TEXT NOT NULL,
                signal_date TEXT NOT NULL,
                pattern_type TEXT NOT NULL CHECK(pattern_type IN ('EXHAUSTION', 'SHADOW_STOP', 'STRONG_REVERSAL')),
                source TEXT NOT NULL CHECK(source IN ('AUTO', 'MANUAL')),
                review_status TEXT NOT NULL DEFAULT 'PENDING' CHECK(review_status IN ('PENDING', 'CONFIRMED', 'REJECTED')),
                rule_version TEXT NOT NULL,
                metrics_json TEXT NOT NULL,
                candles_json TEXT NOT NULL,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                reviewed_at TEXT,
                UNIQUE(user_id, alert_type_id, stock_code, signal_date)
            )""",
            """CREATE INDEX IF NOT EXISTS idx_alert_signal_samples_user_date
            ON alert_signal_samples(user_id, signal_date DESC, id DESC)""",
            """CREATE TABLE IF NOT EXISTS alert_signal_outcomes (
                sample_id INTEGER PRIMARY KEY REFERENCES alert_signal_samples(id) ON DELETE CASCADE,
                day_1_close_return REAL,
                day_3_close_return REAL,
                day_5_close_return REAL,
                day_10_close_return REAL,
                day_1_max_return REAL,
                day_3_max_return REAL,
                day_5_max_return REAL,
                day_10_max_return REAL,
                day_1_max_drawdown REAL,
                day_3_max_drawdown REAL,
                day_5_max_drawdown REAL,
                day_10_max_drawdown REAL,
                evaluated_through_date TEXT,
                updated_at TEXT NOT NULL
            )""",
        )
        for statement in schema_statements:
            db.execute(statement)
        if migrate_signal_pool:
            sample_columns = (
                "id, user_id, alert_type_id, stock_code, stock_name, signal_date, pattern_type, source, "
                "review_status, rule_version, metrics_json, candles_json, note, created_at, updated_at, reviewed_at"
            )
            db.execute(
                f"INSERT INTO alert_signal_samples ({sample_columns}) SELECT {sample_columns} FROM __legacy_alert_signal_samples"
            )
            outcome_columns = (
                "sample_id, day_1_close_return, day_3_close_return, day_5_close_return, day_10_close_return, "
                "day_1_max_return, day_3_max_return, day_5_max_return, day_10_max_return, "
                "day_1_max_drawdown, day_3_max_drawdown, day_5_max_drawdown, day_10_max_drawdown, "
                "evaluated_through_date, updated_at"
            )
            db.execute(
                f"INSERT INTO alert_signal_outcomes ({outcome_columns}) SELECT {outcome_columns} FROM __legacy_alert_signal_outcomes"
            )
            db.execute("DROP TABLE __legacy_alert_signal_outcomes")
            db.execute("DROP TABLE __legacy_alert_signal_samples")
        now = datetime.now().isoformat(timespec="seconds")
        db.execute(
            """INSERT INTO alert_types
            (code, name, signal_side, description, params_json, sample_json, enabled, created_at, updated_at)
            VALUES (?, '三日低吸', 'BUY', '前两日连续收阴走低，第三日创新低并出现明显下影线。', ?, ?, 1, ?, ?)
            ON CONFLICT(code) DO NOTHING""",
            (
                THREE_DAY_DIP_CODE,
                json.dumps(THREE_DAY_DIP_DEFAULT_PARAMS, ensure_ascii=False),
                "{}",
                now,
                now,
            ),
        )
        db.execute(
            """INSERT INTO alert_types
            (code, name, signal_side, description, params_json, sample_json, enabled, created_at, updated_at)
            VALUES (?, '日内反弹', 'BUY', '急跌后形成更高低点，并放量突破第一波反弹高点。', ?, ?, 1, ?, ?)
            ON CONFLICT(code) DO NOTHING""",
            (
                INTRADAY_REBOUND_CODE,
                json.dumps(INTRADAY_REBOUND_DEFAULT_PARAMS, ensure_ascii=False),
                "{}",
                now,
                now,
            ),
        )
        execution_columns = {column["name"] for column in db.execute("PRAGMA table_info(executions)")}
        if "source" not in execution_columns:
            db.execute("ALTER TABLE executions ADD COLUMN source TEXT NOT NULL DEFAULT 'STATEMENT'")
        daily_price_columns = {column["name"] for column in db.execute("PRAGMA table_info(daily_prices)")}
        if "amount" not in daily_price_columns:
            db.execute("ALTER TABLE daily_prices ADD COLUMN amount REAL")
        if needs_migration:
            for table in business_tables:
                if table not in legacy_columns:
                    continue
                columns = [column["name"] for column in db.execute(f"PRAGMA table_info({table})")]
                copied = [column for column in columns if column != "user_id" and column in legacy_columns[table]]
                target_columns = ["user_id", *copied]
                user_expression = "user_id" if "user_id" in legacy_columns[table] else "?"
                db.execute(
                    f"INSERT INTO {table} ({', '.join(target_columns)}) "
                    f"SELECT {user_expression}, {', '.join(copied)} FROM __legacy_{table}",
                    () if user_expression == "user_id" else (default_user_id,),
                )
            for table in business_tables:
                if table in legacy_columns:
                    db.execute(f"DROP TABLE __legacy_{table}")
        for user in db.execute("SELECT id FROM users"):
            rebuild_trade_episodes(db, user["id"])
        pasted_manual = db.execute(
            """SELECT id, user_id, trade_date, deal_id, stock_code, action, quantity, deal_price
            FROM executions WHERE source = 'MANUAL' AND import_job_id IN
            (SELECT id FROM import_jobs WHERE filename = '当天成交文本补录')"""
        ).fetchall()
        for row in pasted_manual:
            identity = (
                f"BROKER_TODAY|{row['user_id']}|{row['trade_date']}|{row['deal_id']}|"
                f"{row['stock_code']}|{row['action']}|{row['quantity']}|{row['deal_price']}"
            )
            new_fingerprint = hashlib.sha256(identity.encode("utf-8")).hexdigest()
            db.execute(
                "UPDATE executions SET source = 'BROKER_TODAY', fingerprint = ? WHERE id = ?",
                (new_fingerprint, row["id"]),
            )
        foreign_key_errors = db.execute("PRAGMA foreign_key_check").fetchall()
        if foreign_key_errors:
            raise sqlite3.IntegrityError(f"数据库迁移外键检查失败：{foreign_key_errors}")
        db.execute(f"PRAGMA user_version = {SCHEMA_VERSION}")
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def current_user_id() -> int:
    if has_request_context() and hasattr(g, "current_user"):
        return int(g.current_user["id"])
    with db_connect() as db:
        return int(db.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()["id"])


@app.before_request
def load_current_user() -> None:
    with db_connect() as db:
        user = None
        try:
            requested_user_id = int(session.get("user_id", 0))
        except (TypeError, ValueError):
            requested_user_id = 0
        if requested_user_id:
            user = db.execute("SELECT * FROM users WHERE id = ?", (requested_user_id,)).fetchone()
        if user is None:
            user = db.execute("SELECT * FROM users ORDER BY id LIMIT 1").fetchone()
            session["user_id"] = user["id"]
        g.current_user = user


@app.context_processor
def inject_users() -> dict:
    with db_connect() as db:
        users = db.execute("SELECT * FROM users ORDER BY name COLLATE NOCASE, id").fetchall()
        unread_notifications, recent_notifications = notification_rows(db, current_user_id(), 8)
    return {
        "users": users,
        "current_user": g.current_user,
        "unread_notifications": unread_notifications,
        "recent_notifications": recent_notifications,
    }


def clean_header(value: object) -> str:
    return str(value or "").strip().replace(" ", "").replace("\n", "")


def to_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    text = str(value).strip().replace(",", "").replace("¥", "")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    return float(text or 0)


def normalize_date(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip().split(" ")[0].replace("/", "-").replace(".", "-")
    if text.isdigit() and len(text) == 8:
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    parsed = datetime.strptime(text, "%Y-%m-%d")
    return parsed.strftime("%Y-%m-%d")


def normalize_time(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M:%S")
    text = str(value or "00:00:00").strip()
    if text.isdigit() and len(text) == 6:
        return f"{text[:2]}:{text[2:4]}:{text[4:]}"
    parts = text.split(":")
    if len(parts) == 2:
        text += ":00"
    return datetime.strptime(text, "%H:%M:%S").strftime("%H:%M:%S")


def normalize_code(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value or "").strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    text = text.replace("SH", "").replace("SZ", "")
    digits = "".join(character for character in text if character.isdigit())
    if not digits:
        raise ValueError("证券代码为空")
    return digits.zfill(6)[-6:]


def normalize_action(value: object, quantity: float) -> str:
    text = str(value or "").strip().upper()
    if "卖" in text or text in {"SELL", "S"}:
        return "SELL"
    if "买" in text or text in {"BUY", "B"}:
        return "BUY"
    if quantity < 0:
        return "SELL"
    raise ValueError(f"无法识别操作：{value}")


def manual_number(name: str, label: str, *, positive: bool = False) -> float:
    try:
        value = to_number(request.form.get(name, ""))
    except ValueError as error:
        raise ValueError(f"{label}格式不正确") from error
    if positive and value <= 0:
        raise ValueError(f"{label}必须大于 0")
    if value < 0:
        raise ValueError(f"{label}不能小于 0")
    return value


def read_upload(file) -> list[list[object]]:
    extension = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx、.xls、.csv、.txt 文件")
    if extension == "xlsx":
        workbook = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
        try:
            return [list(row) for row in workbook.active.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if extension == "xls":
        workbook = xlrd.open_workbook(file_contents=file.read(), on_demand=True)
        try:
            sheet = workbook.sheet_by_index(0)
            return [sheet.row_values(index) for index in range(sheet.nrows)]
        finally:
            workbook.release_resources()

    content = file.read()
    text = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("文件编码无法识别，请另存为 UTF-8 CSV 后重试")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return [row for row in csv.reader(io.StringIO(text), dialect)]
    except csv.Error:
        return [line.split("\t") for line in text.splitlines()]


def parse_rows(rows: list[list[object]]) -> tuple[list[dict], list[str], list[str], int]:
    header_index = -1
    column_map: dict[int, str] = {}
    for index, row in enumerate(rows[:20]):
        candidate = {position: HEADER_ALIASES[clean_header(value)] for position, value in enumerate(row) if clean_header(value) in HEADER_ALIASES}
        if {"trade_date", "stock_code", "action", "raw_quantity", "deal_price"}.issubset(candidate.values()):
            header_index = index
            column_map = candidate
            break
    if header_index < 0:
        raise ValueError("未找到交割单表头，请确认包含成交日期、证券代码、操作、成交数量和成交价格")

    parsed: list[dict] = []
    errors: list[str] = []
    ignored: list[str] = []
    source_rows = rows[header_index + 1 :]
    for source_number, row in enumerate(source_rows, start=header_index + 2):
        if not any(value not in (None, "") for value in row):
            continue
        raw = {field: "" for field in FIELDS}
        for position, field in column_map.items():
            if position < len(row):
                raw[field] = row[position]
        action_text = str(raw["action"] or "").strip()
        if action_text in IGNORED_ACTIONS:
            ignored.append(f"第 {source_number} 行：{action_text}（非买卖成交）")
            continue
        try:
            raw_quantity = to_number(raw["raw_quantity"])
            quantity = abs(raw_quantity)
            price = to_number(raw["deal_price"])
            if quantity <= 0 or price <= 0:
                raise ValueError(
                    f"成交数量和价格必须大于 0（实际读取：数量={raw['raw_quantity']!s}，"
                    f"价格={raw['deal_price']!s}，操作={raw['action']!s}）"
                )
            action = normalize_action(raw["action"], raw_quantity)
            execution = {
                "trade_date": normalize_date(raw["trade_date"]),
                "trade_time": normalize_time(raw["trade_time"]),
                "stock_code": normalize_code(raw["stock_code"]),
                "stock_name": str(raw["stock_name"] or "未知证券").strip(),
                "action": action,
                "raw_quantity": raw_quantity,
                "quantity": quantity,
                "deal_id": str(raw["deal_id"] or "").strip(),
                "deal_price": price,
                "deal_amount": abs(to_number(raw["deal_amount"])) or price * quantity,
                "commission": abs(to_number(raw["commission"])),
                "stamp_tax": abs(to_number(raw["stamp_tax"])),
                "other_fee": abs(to_number(raw["other_fee"])),
                "extra_fee": abs(to_number(raw["extra_fee"])),
                "transfer_fee": abs(to_number(raw["transfer_fee"])),
                "market": str(raw["market"] or "").strip(),
                "raw": {key: str(value or "") for key, value in raw.items()},
                "source_row": source_number,
            }
            identity = "|".join(str(execution[key]) for key in ("trade_date", "trade_time", "stock_code", "action", "quantity", "deal_price", "deal_id"))
            execution["fingerprint"] = hashlib.sha256(identity.encode("utf-8")).hexdigest()
            parsed.append(execution)
        except (ValueError, TypeError) as error:
            errors.append(f"第 {source_number} 行：{error}")
    return parsed, errors, ignored, len(source_rows)


def total_fees(row: sqlite3.Row | dict) -> float:
    return sum(float(row[key] or 0) for key in ("commission", "stamp_tax", "other_fee", "extra_fee", "transfer_fee"))


def rebuild_trade_episodes(db: sqlite3.Connection, user_id: int) -> None:
    db.execute(
        "DELETE FROM trade_episode_executions WHERE trade_episode_id IN (SELECT id FROM trade_episodes WHERE user_id = ?)",
        (user_id,),
    )
    executions = db.execute(
        "SELECT * FROM executions WHERE user_id = ? ORDER BY stock_code, trade_date, trade_time, id",
        (user_id,),
    ).fetchall()
    states: dict[str, dict] = {}
    active_opening_ids: set[int] = set()
    now = datetime.now().isoformat(timespec="seconds")

    for row in executions:
        state = states.setdefault(row["stock_code"], {"quantity": 0.0, "episode_id": None})
        if row["action"] == "BUY":
            role = "ADD"
            if state["quantity"] <= 0.000001:
                role = "ENTRY"
                opened_at = f"{row['trade_date']} {row['trade_time']}"
                db.execute(
                    """INSERT INTO trade_episodes
                    (user_id, stock_code, stock_name, opening_execution_id, opened_at, status, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, 'OPEN', ?, ?)
                    ON CONFLICT(user_id, opening_execution_id) DO UPDATE SET
                    stock_code = excluded.stock_code, stock_name = excluded.stock_name,
                    opened_at = excluded.opened_at, updated_at = excluded.updated_at""",
                    (user_id, row["stock_code"], row["stock_name"], row["id"], opened_at, now, now),
                )
                episode = db.execute(
                    "SELECT id FROM trade_episodes WHERE user_id = ? AND opening_execution_id = ?",
                    (user_id, row["id"]),
                ).fetchone()
                state["episode_id"] = episode["id"]
                active_opening_ids.add(row["id"])
            db.execute(
                """INSERT INTO trade_episode_executions
                (trade_episode_id, execution_id, role, attributed_quantity) VALUES (?, ?, ?, ?)
                ON CONFLICT(trade_episode_id, execution_id) DO UPDATE SET
                role = excluded.role, attributed_quantity = excluded.attributed_quantity""",
                (state["episode_id"], row["id"], role, row["quantity"]),
            )
            state["quantity"] += row["quantity"]
            continue

        attributed = min(state["quantity"], row["quantity"])
        if attributed <= 0.000001 or state["episode_id"] is None:
            continue
        remaining = state["quantity"] - attributed
        role = "EXIT" if remaining <= 0.000001 else "REDUCE"
        db.execute(
            """INSERT INTO trade_episode_executions
            (trade_episode_id, execution_id, role, attributed_quantity) VALUES (?, ?, ?, ?)
            ON CONFLICT(trade_episode_id, execution_id) DO UPDATE SET
            role = excluded.role, attributed_quantity = excluded.attributed_quantity""",
            (state["episode_id"], row["id"], role, attributed),
        )
        state["quantity"] = max(0.0, remaining)
        if role == "EXIT":
            db.execute(
                """UPDATE trade_episodes SET stock_name = ?, closing_execution_id = ?, closed_at = ?,
                status = 'CLOSED', updated_at = ? WHERE id = ? AND user_id = ?""",
                (row["stock_name"], row["id"], f"{row['trade_date']} {row['trade_time']}", now,
                 state["episode_id"], user_id),
            )
            state["episode_id"] = None

    for state in states.values():
        if state["episode_id"] is not None:
            db.execute(
                """UPDATE trade_episodes SET closing_execution_id = NULL, closed_at = NULL,
                status = 'OPEN', updated_at = ? WHERE id = ? AND user_id = ?""",
                (now, state["episode_id"], user_id),
            )
    if active_opening_ids:
        placeholders = ",".join("?" for _ in active_opening_ids)
        db.execute(
            f"DELETE FROM trade_episodes WHERE user_id = ? AND opening_execution_id NOT IN ({placeholders})",
            (user_id, *active_opening_ids),
        )
    else:
        db.execute("DELETE FROM trade_episodes WHERE user_id = ?", (user_id,))


def trade_episode_rows(db: sqlite3.Connection, user_id: int, stock_code: str | None = None) -> list[sqlite3.Row]:
    stock_filter = "AND ep.stock_code = ?" if stock_code else ""
    parameters = (user_id, stock_code) if stock_code else (user_id,)
    return db.execute(
        f"""SELECT ep.*,
        SUBSTR(ep.opened_at, 1, 10) buy_date, SUBSTR(ep.closed_at, 1, 10) sell_date,
        COALESCE(SUM(CASE WHEN e.action = 'BUY' THEN link.attributed_quantity ELSE 0 END), 0) quantity,
        COALESCE(SUM(CASE WHEN e.action = 'BUY' THEN e.deal_price * link.attributed_quantity ELSE 0 END) /
            NULLIF(SUM(CASE WHEN e.action = 'BUY' THEN link.attributed_quantity ELSE 0 END), 0), 0) buy_price,
        COALESCE(SUM(CASE WHEN e.action = 'SELL' THEN e.deal_price * link.attributed_quantity ELSE 0 END) /
            NULLIF(SUM(CASE WHEN e.action = 'SELL' THEN link.attributed_quantity ELSE 0 END), 0), 0) sell_price,
        COALESCE(SUM(CASE WHEN e.action = 'BUY' THEN
            (e.deal_amount + e.commission + e.stamp_tax + e.other_fee + e.extra_fee + e.transfer_fee)
            * link.attributed_quantity / e.quantity ELSE 0 END), 0) buy_cost,
        COALESCE(SUM(CASE WHEN e.action = 'SELL' THEN
            (e.deal_amount - e.commission - e.stamp_tax - e.other_fee - e.extra_fee - e.transfer_fee)
            * link.attributed_quantity / e.quantity ELSE 0 END), 0) sell_income,
        COALESCE(SUM((e.commission + e.stamp_tax + e.other_fee + e.extra_fee + e.transfer_fee)
            * link.attributed_quantity / e.quantity), 0) fees,
        COALESCE((SELECT SUM(f.profit) FROM fifo_matches f
            JOIN trade_episode_executions buy_link ON buy_link.execution_id = f.buy_execution_id
            WHERE buy_link.trade_episode_id = ep.id), 0) profit,
        COALESCE((SELECT SUM(f.profit) FROM fifo_matches f
            JOIN trade_episode_executions buy_link ON buy_link.execution_id = f.buy_execution_id
            WHERE buy_link.trade_episode_id = ep.id) /
            NULLIF((SELECT SUM(f.buy_cost) FROM fifo_matches f
                JOIN trade_episode_executions buy_link ON buy_link.execution_id = f.buy_execution_id
                WHERE buy_link.trade_episode_id = ep.id), 0), 0) profit_rate,
        CASE WHEN ep.closed_at IS NOT NULL THEN
            CAST(julianday(SUBSTR(ep.closed_at, 1, 10)) - julianday(SUBSTR(ep.opened_at, 1, 10)) AS INTEGER)
            ELSE CAST(julianday('now') - julianday(SUBSTR(ep.opened_at, 1, 10)) AS INTEGER) END holding_days,
        CASE WHEN ep.closed_at IS NOT NULL THEN julianday(ep.closed_at) - julianday(ep.opened_at)
            ELSE julianday('now') - julianday(ep.opened_at) END holding_days_exact,
        COALESCE((SELECT CASE WHEN r.review_status = 'COMPLETED' THEN 'COMPLETED'
                              WHEN r.id IS NOT NULL THEN 'DRAFT' ELSE 'PENDING' END
                  FROM trade_reviews r
                  WHERE r.trade_episode_id = ep.id AND r.user_id = ep.user_id), 'PENDING') review_status
        FROM trade_episodes ep
        JOIN trade_episode_executions link ON link.trade_episode_id = ep.id
        JOIN executions e ON e.id = link.execution_id AND e.user_id = ep.user_id
        WHERE ep.user_id = ? {stock_filter}
        GROUP BY ep.id ORDER BY ep.opened_at DESC, ep.id DESC""",
        parameters,
    ).fetchall()


def trade_episode_detail(db: sqlite3.Connection, user_id: int, episode_id: int) -> dict | None:
    episodes = [row for row in trade_episode_rows(db, user_id) if row["id"] == episode_id]
    if not episodes:
        return None
    episode = episodes[0]
    executions = db.execute(
        """SELECT e.*, link.role, link.attributed_quantity FROM trade_episode_executions link
        JOIN executions e ON e.id = link.execution_id
        WHERE link.trade_episode_id = ? AND e.user_id = ? ORDER BY e.trade_date, e.trade_time, e.id""",
        (episode_id, user_id),
    ).fetchall()
    matches = db.execute(
        """SELECT f.* FROM fifo_matches f
        JOIN trade_episode_executions link ON link.execution_id = f.buy_execution_id
        WHERE link.trade_episode_id = ? AND f.user_id = ? ORDER BY f.sell_date, f.id""",
        (episode_id, user_id),
    ).fetchall()
    review = db.execute(
        "SELECT * FROM trade_reviews WHERE trade_episode_id = ? AND user_id = ?",
        (episode_id, user_id),
    ).fetchone()
    review_reason_types: list[str] = []
    if review is not None:
        review_reason_types = [
            row["reason_type"]
            for row in db.execute(
                "SELECT reason_type FROM trade_review_reason_types WHERE trade_review_id = ? ORDER BY reason_type",
                (review["id"],),
            )
        ]
    return {
        "episode": episode,
        "executions": executions,
        "buys": [row for row in executions if row["action"] == "BUY"],
        "sells": [row for row in executions if row["action"] == "SELL"],
        "matches": matches,
        "review": review,
        "review_reason_types": review_reason_types,
        "metrics": excursion_metrics_for(db, user_id, episode_id),
        "diagnosis": trade_diagnosis(db, review, excursion_metrics_for(db, user_id, episode_id), episode),
        "chart": build_chart(executions, matches),
    }


def rebuild_fifo(db: sqlite3.Connection, user_id: int | None = None) -> None:
    user_id = user_id or current_user_id()
    db.execute("DELETE FROM fifo_matches WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM positions WHERE user_id = ?", (user_id,))
    db.execute("DELETE FROM unmatched_sells WHERE user_id = ?", (user_id,))
    executions = db.execute(
        "SELECT * FROM executions WHERE user_id = ? ORDER BY trade_date, trade_time, id", (user_id,)
    ).fetchall()
    queues: dict[str, list[dict]] = {}
    for row in executions:
        queue = queues.setdefault(row["stock_code"], [])
        fees = total_fees(row)
        if row["action"] == "BUY":
            queue.append({"row": row, "remaining": row["quantity"], "unit_cost": (row["deal_amount"] + fees) / row["quantity"]})
            continue

        remaining = row["quantity"]
        sell_unit_income = (row["deal_amount"] - fees) / row["quantity"]
        while remaining > 0.000001 and queue:
            buy = queue[0]
            quantity = min(remaining, buy["remaining"])
            buy_cost = quantity * buy["unit_cost"]
            sell_income = quantity * sell_unit_income
            profit = sell_income - buy_cost
            holding_days = (datetime.strptime(row["trade_date"], "%Y-%m-%d") - datetime.strptime(buy["row"]["trade_date"], "%Y-%m-%d")).days
            db.execute(
                """INSERT INTO fifo_matches
                (user_id, stock_code, stock_name, buy_execution_id, sell_execution_id, buy_date, sell_date,
                 quantity, buy_cost, sell_income, profit, profit_rate, holding_days)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (user_id, row["stock_code"], row["stock_name"], buy["row"]["id"], row["id"], buy["row"]["trade_date"],
                 row["trade_date"], quantity, buy_cost, sell_income, profit, profit / buy_cost if buy_cost else 0, holding_days),
            )
            remaining -= quantity
            buy["remaining"] -= quantity
            if buy["remaining"] <= 0.000001:
                queue.pop(0)
        if remaining > 0.000001:
            db.execute(
                "INSERT INTO unmatched_sells (user_id, sell_execution_id, stock_code, stock_name, sell_date, quantity, sell_income) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (user_id, row["id"], row["stock_code"], row["stock_name"], row["trade_date"], remaining, remaining * sell_unit_income),
            )

    for stock_code, queue in queues.items():
        for buy in queue:
            row = buy["row"]
            db.execute(
                "INSERT INTO positions (user_id, stock_code, stock_name, buy_execution_id, buy_date, quantity, unit_cost, total_cost) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (user_id, stock_code, row["stock_name"], row["id"], row["trade_date"], buy["remaining"], buy["unit_cost"], buy["remaining"] * buy["unit_cost"]),
            )
    rebuild_trade_episodes(db, user_id)


def analysis_data(db: sqlite3.Connection, user_id: int | None = None, stock_sort: str = "profit", performance_year: int | None = None) -> dict:
    user_id = user_id or current_user_id()
    stock_sort = stock_sort if stock_sort in STOCK_SORTS else "profit"
    summary = db.execute(
        """SELECT COUNT(*) trades, COALESCE(SUM(profit), 0) profit,
        COALESCE(AVG(profit), 0) avg_profit,
        COALESCE(AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END), 0) win_rate,
        COALESCE(AVG(holding_days), 0) avg_holding_days
        FROM fifo_matches WHERE user_id = ?""", (user_id,)
    ).fetchone()
    execution_summary = db.execute(
        """SELECT COUNT(*) executions, COUNT(DISTINCT stock_code) stocks,
        COALESCE(SUM(deal_amount), 0) turnover,
        COALESCE(SUM(commission + stamp_tax + other_fee + extra_fee + transfer_fee), 0) fees
        FROM executions WHERE user_id = ?""", (user_id,)
    ).fetchone()
    by_stock = db.execute(
        f"""WITH episode_profits AS (
            SELECT ep.id, ep.stock_code, ep.stock_name,
            COALESCE(SUM(f.profit), 0) profit,
            COALESCE(SUM(f.buy_cost), 0) buy_cost,
            CAST(julianday(SUBSTR(ep.closed_at, 1, 10)) -
                julianday(SUBSTR(ep.opened_at, 1, 10)) AS INTEGER) holding_days
            FROM trade_episodes ep
            JOIN trade_episode_executions link ON link.trade_episode_id = ep.id AND link.role IN ('ENTRY', 'ADD')
            LEFT JOIN fifo_matches f ON f.user_id = ep.user_id AND f.buy_execution_id = link.execution_id
            WHERE ep.user_id = ? AND ep.status = 'CLOSED'
            GROUP BY ep.id HAVING COUNT(f.id) > 0
        ), latest_executions AS (
            SELECT stock_code, MAX(trade_date || ' ' || trade_time) last_trade_at
            FROM executions WHERE user_id = ? GROUP BY stock_code
        )
        SELECT episodes.stock_code, MAX(episodes.stock_name) stock_name,
        COUNT(*) trades, COALESCE(SUM(episodes.profit), 0) profit,
        COALESCE(SUM(episodes.profit) / NULLIF(SUM(episodes.buy_cost), 0), 0) profit_rate,
        COALESCE(AVG(CASE WHEN episodes.profit > 0 THEN 1.0 ELSE 0 END), 0) win_rate,
        MAX(CASE WHEN episodes.profit > 0 THEN episodes.profit END) max_profit,
        MIN(CASE WHEN episodes.profit < 0 THEN episodes.profit END) max_loss,
        COALESCE(AVG(episodes.holding_days), 0) avg_holding_days,
        latest.last_trade_at
        FROM episode_profits episodes
        JOIN latest_executions latest ON latest.stock_code = episodes.stock_code
        GROUP BY episodes.stock_code, latest.last_trade_at
        ORDER BY {STOCK_SORTS[stock_sort]}""",
        (user_id, user_id),
    ).fetchall()
    monthly = db.execute(
        """SELECT SUBSTR(sell_date, 1, 7) month, COUNT(*) trades, SUM(profit) profit
        FROM fifo_matches WHERE user_id = ? GROUP BY month ORDER BY month""", (user_id,)
    ).fetchall()
    performance = performance_data(db, user_id, performance_year)
    recent = db.execute("SELECT * FROM fifo_matches WHERE user_id = ? ORDER BY sell_date DESC, id DESC LIMIT 12", (user_id,)).fetchall()
    review_progress = review_progress_data(db, user_id)
    return {
        "summary": summary,
        "execution_summary": execution_summary,
        "by_stock": by_stock,
        "monthly": monthly,
        "performance": performance,
        "recent": recent,
        "review_progress": review_progress,
        "trading_system": trading_system_stats(db, user_id),
    }


def performance_data(db: sqlite3.Connection, user_id: int, performance_year: int | None = None) -> dict:
    """Build the day/month/year views from realized FIFO matches."""
    daily_rows = db.execute(
        """SELECT sell_date period, COUNT(*) trades, SUM(profit) profit
        FROM fifo_matches WHERE user_id = ? GROUP BY sell_date ORDER BY sell_date""", (user_id,)
    ).fetchall()
    monthly_rows = db.execute(
        """SELECT SUBSTR(sell_date, 1, 7) period, COUNT(*) trades,
        SUM(profit) profit, AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END) win_rate
        FROM fifo_matches WHERE user_id = ? GROUP BY period ORDER BY period""", (user_id,)
    ).fetchall()
    yearly_rows = db.execute(
        """SELECT SUBSTR(sell_date, 1, 4) period, COUNT(*) trades,
        SUM(profit) profit, AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END) win_rate
        FROM fifo_matches WHERE user_id = ? GROUP BY period ORDER BY period""", (user_id,)
    ).fetchall()

    def rows_with_scale(rows: list) -> list[dict]:
        values = [abs(float(row["profit"] or 0)) for row in rows]
        scale = max(values or [1])
        return [{**dict(row), "profit": float(row["profit"] or 0), "scale": scale} for row in rows]

    daily = rows_with_scale(daily_rows)
    monthly = rows_with_scale(monthly_rows)
    yearly = rows_with_scale(yearly_rows)
    available_years = sorted({int(row["period"]) for row in yearly})
    selected_year = performance_year if performance_year in available_years else (available_years[-1] if available_years else date.today().year)
    year_monthly = {row["period"]: row for row in monthly if row["period"].startswith(f"{selected_year:04d}-")}
    month_grid = []
    for month_number in range(1, 13):
        period = f"{selected_year:04d}-{month_number:02d}"
        row = year_monthly.get(period)
        month_grid.append({
            "period": period,
            "label": f"{month_number}月",
            "profit": row["profit"] if row else None,
            "trades": row["trades"] if row else 0,
            "win_rate": row["win_rate"] if row else None,
        })
    year_rows = [row for row in yearly if int(row["period"]) == selected_year]
    selected_year_profit = sum(row["profit"] for row in year_monthly.values())
    selected_year_trades = sum(row["trades"] for row in year_monthly.values())
    selected_year_wins = sum(row["trades"] * row["win_rate"] for row in year_monthly.values())
    selected_year_positive_months = sum(1 for row in month_grid if row["profit"] is not None and row["profit"] > 0)
    latest_period = next((row["period"] for row in reversed(monthly) if row["period"].startswith(f"{selected_year:04d}-")), f"{selected_year:04d}-01")
    try:
        calendar_year, selected_month = (int(value) for value in latest_period.split("-"))
    except ValueError:
        calendar_year, selected_month = selected_year, date.today().month
    month_daily = {row["period"]: row for row in daily if row["period"].startswith(latest_period)}
    month_days = list(month_daily.values())
    weeks = []
    # The UI labels weeks Sunday-first, so the calendar cells must use the same order.
    for week in calendar.Calendar(firstweekday=calendar.SUNDAY).monthdayscalendar(calendar_year, selected_month):
        week_rows = []
        for day_number in week:
            day_key = f"{calendar_year:04d}-{selected_month:02d}-{day_number:02d}" if day_number else ""
            row = month_daily.get(day_key)
            week_rows.append({
                "day": day_number,
                "period": day_key,
                "profit": row["profit"] if row else None,
                "trades": row["trades"] if row else 0,
            })
        weeks.append(week_rows)

    active_rows = monthly
    total_profit = sum(row["profit"] for row in active_rows)
    wins = sum(1 for row in active_rows if row["profit"] > 0)
    losses = sum(1 for row in active_rows if row["profit"] < 0)
    best = max(active_rows, key=lambda row: row["profit"], default=None)
    worst = min(active_rows, key=lambda row: row["profit"], default=None)
    return {
        "daily": daily,
        "monthly": monthly,
        "yearly": yearly,
        "selected_month": latest_period,
        "selected_year": selected_year,
        "available_years": available_years,
        "month_grid": month_grid,
        "selected_year_profit": selected_year_profit,
        "selected_year_trades": selected_year_trades,
        "selected_year_wins": selected_year_wins,
        "selected_year_positive_months": selected_year_positive_months,
        "month_name": f"{calendar_year}年{selected_month}月",
        "calendar_weeks": weeks,
        "day_stats": {
            "profit": sum(row["profit"] for row in month_days),
            "trading_days": len(month_days),
            "best": max(month_days, key=lambda row: row["profit"], default=None),
            "worst": min(month_days, key=lambda row: row["profit"], default=None),
        },
        "stats": {"profit": total_profit, "wins": wins, "losses": losses, "best": best, "worst": worst},
    }


def portfolio_analysis_data(db: sqlite3.Connection, user_id: int) -> dict:
    today = date.today()
    position_rows = db.execute(
        """SELECT p.stock_code, MAX(p.stock_name) stock_name, MIN(p.buy_date) first_buy_date,
        SUM(p.quantity) quantity, SUM(p.total_cost) total_cost,
        SUM(p.total_cost) / NULLIF(SUM(p.quantity), 0) avg_cost,
        (SELECT ep.id FROM trade_episodes ep
            WHERE ep.user_id = p.user_id AND ep.stock_code = p.stock_code AND ep.status = 'OPEN'
            ORDER BY ep.opened_at DESC, ep.id DESC LIMIT 1) open_episode_id
        FROM positions p WHERE p.user_id = ? GROUP BY p.stock_code ORDER BY total_cost DESC""",
        (user_id,),
    ).fetchall()

    positions = []
    for row in position_rows:
        item = dict(row)
        item["quantity"] = float(item["quantity"] or 0)
        item["total_cost"] = float(item["total_cost"] or 0)
        item["avg_cost"] = float(item["avg_cost"] or 0)
        item["holding_days"] = max(0, (today - datetime.strptime(item["first_buy_date"], "%Y-%m-%d").date()).days)
        prices = db.execute(
            """SELECT trade_date, close, fetched_at FROM daily_prices
            WHERE stock_code = ? AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 2""",
            (item["stock_code"],),
        ).fetchall()
        price = prices[0] if prices else None
        item["latest_price"] = float(price["close"]) if price is not None else None
        item["previous_close"] = float(prices[1]["close"]) if len(prices) > 1 else None
        item["change_rate"] = (
            item["latest_price"] / item["previous_close"] - 1
            if item["latest_price"] is not None and item["previous_close"]
            else None
        )
        item["price_date"] = price["trade_date"] if price is not None else None
        item["price_fetched_at"] = price["fetched_at"] if price is not None else None
        if item["latest_price"] is not None:
            item["market_value"] = item["quantity"] * item["latest_price"]
            item["unrealized_profit"] = item["market_value"] - item["total_cost"]
            item["unrealized_rate"] = item["unrealized_profit"] / item["total_cost"] if item["total_cost"] else 0.0
        else:
            item["market_value"] = None
            item["unrealized_profit"] = None
            item["unrealized_rate"] = None

        review = None
        if item["open_episode_id"] is not None:
            review = db.execute(
                """SELECT review_status, trade_reason, expected_target_price, stop_loss_price,
                expected_holding_days FROM trade_reviews
                WHERE user_id = ? AND trade_episode_id = ?""",
                (user_id, item["open_episode_id"]),
            ).fetchone()
        item["review"] = dict(review) if review is not None else None
        positions.append(item)

    total_cost = sum(item["total_cost"] for item in positions)
    priced_positions = [item for item in positions if item["market_value"] is not None]
    prices_complete = bool(positions) and len(priced_positions) == len(positions)
    total_market_value = sum(item["market_value"] for item in priced_positions) if prices_complete else None
    total_unrealized_profit = total_market_value - total_cost if total_market_value is not None else None
    total_unrealized_rate = total_unrealized_profit / total_cost if total_unrealized_profit is not None and total_cost else None
    positions.sort(key=lambda item: item["total_cost"] or 0, reverse=True)
    visible_positions = positions

    latest_price_date = max((item["price_date"] for item in priced_positions), default=None)
    latest_fetched_at = max((item["price_fetched_at"] for item in priced_positions), default=None)
    return {
        "summary": {
            "count": len(positions), "total_cost": total_cost,
            "total_market_value": total_market_value, "unrealized_profit": total_unrealized_profit,
            "unrealized_rate": total_unrealized_rate, "prices_complete": prices_complete,
            "priced_count": len(priced_positions), "price_date": latest_price_date,
            "price_fetched_at": latest_fetched_at,
        },
        "positions": visible_positions, "all_positions": positions,
    }


WATCHLIST_SORTS = {
    "priority": "priority DESC, updated_at DESC, stock_code",
    "change": "change_rate DESC, priority DESC, stock_code",
    "recent": "created_at DESC, stock_code",
    "updated": "updated_at DESC, stock_code",
}


def watchlist_data(db: sqlite3.Connection, user_id: int, query: str = "", priority: str = "all", sort: str = "priority", direction: str = "desc") -> dict:
    sort = sort if sort in WATCHLIST_SORTS else "priority"
    priority = priority if priority in ("all", "high", "medium", "low") else "all"
    direction = direction if direction in ("asc", "desc") else "desc"
    rows = db.execute(
        "SELECT * FROM watchlist_stocks WHERE user_id = ? ORDER BY priority DESC, updated_at DESC, id DESC",
        (user_id,),
    ).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        prices = db.execute(
            """SELECT trade_date, close, fetched_at FROM daily_prices WHERE stock_code = ?
            AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 2""",
            (item["stock_code"],),
        ).fetchall()
        item["latest_price"] = float(prices[0]["close"]) if prices else None
        item["previous_close"] = float(prices[1]["close"]) if len(prices) > 1 else None
        item["price_date"] = prices[0]["trade_date"] if prices else None
        item["price_fetched_at"] = prices[0]["fetched_at"] if prices else None
        item["change_amount"] = item["latest_price"] - item["previous_close"] if item["previous_close"] else None
        item["change_rate"] = item["change_amount"] / item["previous_close"] if item["previous_close"] else None
        item["is_held"] = db.execute(
            "SELECT 1 FROM positions WHERE user_id = ? AND stock_code = ? LIMIT 1", (user_id, item["stock_code"])
        ).fetchone() is not None
        rebound = db.execute(
            """SELECT last_matched FROM alert_rule_states
            WHERE user_id = ? AND alert_type_id = (SELECT id FROM alert_types WHERE code = ?)
            AND stock_code = ? AND trade_date = ?""",
            (user_id, INTRADAY_REBOUND_CODE, item["stock_code"], item["price_date"]),
        ).fetchone()
        item["intraday_rebound"] = bool(rebound["last_matched"]) if rebound else False
        items.append(item)
    if query:
        lowered = query.lower()
        items = [item for item in items if lowered in item["stock_code"] or lowered in item["stock_name"].lower()]
    priority_values = {"high": 3, "medium": 2, "low": 1}
    if priority != "all":
        items = [item for item in items if item["priority"] == priority_values[priority]]
    reverse = direction == "desc"
    if sort == "change":
        items.sort(key=lambda item: (item["change_rate"] is not None, item["change_rate"] or 0), reverse=reverse)
        if direction == "asc":
            items.sort(key=lambda item: item["change_rate"] is None)
    elif sort == "recent":
        items.sort(key=lambda item: (item["created_at"], item["stock_code"]), reverse=True)
    elif sort == "updated":
        items.sort(key=lambda item: (item["updated_at"], item["stock_code"]), reverse=True)
    else:
        items.sort(key=lambda item: (item["priority"], item["updated_at"], item["stock_code"]), reverse=reverse)
    for index, item in enumerate(items, start=1):
        item["rank"] = index if item["change_rate"] is not None else None
    priced = [item for item in items if item["latest_price"] is not None]
    return {
        "items": items, "query": query, "priority": priority, "sort": sort, "direction": direction,
        "summary": {
            "total": len(items), "high": sum(item["priority"] == 3 for item in items),
            "gainers": sum(item["change_rate"] is not None and item["change_rate"] > 0 for item in items),
            "losers": sum(item["change_rate"] is not None and item["change_rate"] < 0 for item in items),
            "priced": len(priced), "latest_price_date": max((item["price_date"] for item in priced), default=None),
        },
    }


def three_day_dip_params(raw: dict | None = None) -> dict:
    params = dict(THREE_DAY_DIP_DEFAULT_PARAMS)
    if raw:
        params.update(raw)
    params["decline_days"] = max(1, min(5, int(params["decline_days"])))
    ratio_keys = (
        "min_decline_ratio", "max_signal_low_above_prior_ratio", "max_volume_ratio", "exhaustion_min_repair_ratio",
        "exhaustion_max_body_range_ratio", "exhaustion_min_change_ratio",
        "exhaustion_max_change_ratio", "shadow_min_repair_ratio",
        "shadow_max_body_range_ratio", "shadow_min_lower_shadow_body_ratio",
        "shadow_min_change_ratio", "shadow_max_change_ratio", "reversal_min_decline_ratio",
        "reversal_min_repair_ratio", "reversal_min_recovery_range_ratio",
        "reversal_min_change_ratio", "reversal_max_volume_ratio",
    )
    for key in ratio_keys:
        params[key] = max(-1.0, min(2.0, float(params[key])))
    for key in (
        "require_bearish_candles", "require_declining_closes",
        "intraday_candidate_enabled", "close_confirmation_enabled",
    ):
        params[key] = bool(params[key])
    return params


def evaluate_three_day_dip(
    prices: list[dict], params: dict | None = None, enforce_volume: bool = False,
) -> dict:
    params = three_day_dip_params(params)
    required = params["decline_days"] + 2
    if len(prices) < required:
        return {"matched": False, "reason": f"至少需要 {required} 个交易日行情"}
    candles = [dict(row) for row in prices[-required:]]
    decline_candles = candles[1:-1]
    signal = candles[-1]
    bearish_ok = not params["require_bearish_candles"] or all(
        float(item["close"]) < float(item["open"]) for item in decline_candles
    )
    close_chain = candles[:-1]
    declining_ok = not params["require_declining_closes"] or all(
        float(close_chain[index]["close"]) < float(close_chain[index - 1]["close"])
        for index in range(1, len(close_chain))
    )
    prior_lows = [float(item["low"]) for item in decline_candles]
    prior_low = min(prior_lows)
    signal_low = float(signal["low"])
    new_low_ok = signal_low < prior_low
    signal_low_above_prior_ratio = signal_low / prior_low - 1 if prior_low else 0.0
    low_position_ok = signal_low_above_prior_ratio <= params["max_signal_low_above_prior_ratio"]
    price_range = max(0.0, float(signal["high"]) - float(signal["low"]))
    recovery_amount = max(0.0, float(signal["close"]) - float(signal["low"]))
    recovery_range_ratio = recovery_amount / price_range if price_range else 0.0
    baseline_close = float(candles[0]["close"])
    previous_close = float(decline_candles[-1]["close"])
    decline_ratio = max(0.0, 1 - previous_close / baseline_close) if baseline_close else 0.0
    repair_ratio = recovery_amount / previous_close if previous_close else 0.0
    body_amount = abs(float(signal["close"]) - float(signal["open"]))
    body_range_ratio = body_amount / price_range if price_range else 1.0
    lower_shadow_amount = max(0.0, min(float(signal["open"]), float(signal["close"])) - float(signal["low"]))
    lower_shadow_body_ratio = lower_shadow_amount / body_amount if body_amount else None
    signal_change_ratio = float(signal["close"]) / previous_close - 1 if previous_close else 0.0
    previous_volume = float(decline_candles[-1].get("volume") or 0)
    signal_volume = float(signal.get("volume") or 0)
    volume_ratio = signal_volume / previous_volume if previous_volume and signal_volume else None
    common_ok = bearish_ok and declining_ok and low_position_ok and decline_ratio >= params["min_decline_ratio"]
    exhaustion_volume_ok = not enforce_volume or (
        volume_ratio is not None and volume_ratio <= params["max_volume_ratio"]
    )
    reversal_volume_ok = not enforce_volume or (
        volume_ratio is not None and volume_ratio <= params["reversal_max_volume_ratio"]
    )
    exhaustion_ok = (
        common_ok and repair_ratio >= params["exhaustion_min_repair_ratio"]
        and body_range_ratio <= params["exhaustion_max_body_range_ratio"]
        and params["exhaustion_min_change_ratio"] <= signal_change_ratio <= params["exhaustion_max_change_ratio"]
        and exhaustion_volume_ok
    )
    shadow_ok = (
        common_ok and repair_ratio >= params["shadow_min_repair_ratio"]
        and body_range_ratio <= params["shadow_max_body_range_ratio"]
        and (lower_shadow_body_ratio is None or lower_shadow_body_ratio >= params["shadow_min_lower_shadow_body_ratio"])
        and params["shadow_min_change_ratio"] <= signal_change_ratio <= params["shadow_max_change_ratio"]
        and exhaustion_volume_ok
    )
    reversal_ok = (
        common_ok and decline_ratio >= params["reversal_min_decline_ratio"]
        and repair_ratio >= params["reversal_min_repair_ratio"]
        and recovery_range_ratio >= params["reversal_min_recovery_range_ratio"]
        and signal_change_ratio >= params["reversal_min_change_ratio"]
        and reversal_volume_ok
    )
    pattern_type = (
        "STRONG_REVERSAL" if reversal_ok else "EXHAUSTION" if exhaustion_ok
        else "SHADOW_STOP" if shadow_ok else None
    )
    return {
        "matched": pattern_type is not None,
        "pattern_type": pattern_type,
        "bearish_ok": bearish_ok,
        "declining_ok": declining_ok,
        "new_low_ok": new_low_ok,
        "low_position_ok": low_position_ok,
        "signal_low_above_prior_ratio": signal_low_above_prior_ratio,
        "common_ok": common_ok,
        "exhaustion_ok": exhaustion_ok,
        "shadow_ok": shadow_ok,
        "reversal_ok": reversal_ok,
        "volume_ok": reversal_volume_ok if pattern_type == "STRONG_REVERSAL" else exhaustion_volume_ok,
        "decline_ratio": decline_ratio,
        "recovery_amount": recovery_amount,
        "price_range": price_range,
        "recovery_range_ratio": recovery_range_ratio,
        "repair_ratio": repair_ratio,
        "body_range_ratio": body_range_ratio,
        "lower_shadow_body_ratio": lower_shadow_body_ratio,
        "signal_change_ratio": signal_change_ratio,
        "volume_ratio": volume_ratio,
        "signal": signal,
    }


def intraday_rebound_params(raw: dict | None = None) -> dict:
    params = dict(INTRADAY_REBOUND_DEFAULT_PARAMS)
    if raw:
        params.update(raw)
    params["lookback_minutes"] = max(20, min(240, int(params["lookback_minutes"])))
    params["min_trough_age_minutes"] = max(3, min(30, int(params["min_trough_age_minutes"])))
    for key in ("min_drop_ratio", "min_rebound_ratio"):
        params[key] = max(0.001, min(0.2, float(params[key])))
    params["min_volume_multiple"] = max(0.0, min(10.0, float(params["min_volume_multiple"])))
    params["enabled"] = bool(params["enabled"])
    return params


def evaluate_intraday_rebound(samples: list[dict], params: dict | None = None) -> dict:
    params = intraday_rebound_params(params)
    samples = [dict(sample) for sample in samples[-params["lookback_minutes"]:]]
    if len(samples) < params["min_trough_age_minutes"] + 4:
        return {"matched": False, "status": "等待分钟行情", "reason": "分钟行情不足"}
    prices = [float(sample["price"]) for sample in samples]
    volumes = [float(sample.get("volume") or 0) for sample in samples]
    latest_index = len(samples) - 1
    candidate_indexes = sorted(
        range(2, latest_index - params["min_trough_age_minutes"] + 1), key=lambda index: prices[index]
    )
    for trough_index in candidate_indexes:
        peak_price = max(prices[:trough_index])
        peak_index = prices.index(peak_price)
        trough_price = prices[trough_index]
        drop_ratio = 1 - trough_price / peak_price if peak_price else 0.0
        if drop_ratio < params["min_drop_ratio"]:
            continue
        prior_prices = prices[trough_index + 1:latest_index]
        if len(prior_prices) < 2:
            continue
        first_high = max(prior_prices)
        first_high_index = trough_index + 1 + prior_prices.index(first_high)
        pullback_prices = prices[first_high_index + 1:latest_index]
        if not pullback_prices:
            continue
        higher_low = min(pullback_prices)
        recovery_ratio = prices[-1] / trough_price - 1 if trough_price else 0.0
        breakout_price = max(prior_prices)
        volume_deltas = [max(0.0, volumes[index] - volumes[index - 1]) for index in range(1, len(volumes))]
        recent_volumes = [value for value in volume_deltas[max(0, len(volume_deltas) - 6):-1] if value > 0]
        baseline_volume = sorted(recent_volumes)[len(recent_volumes) // 2] if recent_volumes else 0.0
        latest_volume = volume_deltas[-1] if volume_deltas else 0.0
        volume_multiple = latest_volume / baseline_volume if baseline_volume else None
        structure_ok = higher_low > trough_price and prices[-1] > breakout_price
        volume_ok = volume_multiple is not None and volume_multiple >= params["min_volume_multiple"]
        matched = (
            structure_ok and recovery_ratio >= params["min_rebound_ratio"]
            and (params["min_volume_multiple"] == 0 or volume_ok)
        )
        return {
            "matched": matched,
            "status": "日内反弹候选" if matched else "等待放量突破",
            "peak_price": peak_price,
            "peak_minute": samples[peak_index]["quote_minute"],
            "trough_price": trough_price,
            "trough_minute": samples[trough_index]["quote_minute"],
            "higher_low": higher_low,
            "breakout_price": breakout_price,
            "drop_ratio": drop_ratio,
            "recovery_ratio": recovery_ratio,
            "volume_multiple": volume_multiple,
            "structure_ok": structure_ok,
            "volume_ok": volume_ok,
        }
    return {"matched": False, "status": "未形成急跌结构", "reason": "未找到满足条件的低点"}


def alert_type_data(db: sqlite3.Connection) -> tuple[sqlite3.Row, dict]:
    alert_type = db.execute("SELECT * FROM alert_types WHERE code = ?", (THREE_DAY_DIP_CODE,)).fetchone()
    params = three_day_dip_params(json.loads(alert_type["params_json"]))
    return alert_type, params


def intraday_rebound_alert_type_data(db: sqlite3.Connection) -> tuple[sqlite3.Row, dict]:
    alert_type = db.execute("SELECT * FROM alert_types WHERE code = ?", (INTRADAY_REBOUND_CODE,)).fetchone()
    params = intraday_rebound_params(json.loads(alert_type["params_json"]))
    return alert_type, params


def backtest_three_day_dip(
    db: sqlite3.Connection, stock: sqlite3.Row, start_date: str, end_date: str, params: dict,
) -> dict:
    prices = [dict(row) for row in db.execute(
        """SELECT trade_date, open, high, low, close, volume FROM daily_prices
        WHERE stock_code = ? AND trade_date <= ? AND open IS NOT NULL AND high IS NOT NULL
        AND low IS NOT NULL AND close IS NOT NULL ORDER BY trade_date""",
        (stock["stock_code"], end_date),
    ).fetchall()]
    required = params["decline_days"] + 2
    signal_indexes = [
        index for index, price in enumerate(prices)
        if start_date <= price["trade_date"] <= end_date and index + 1 >= required
    ]
    hits = []
    for index in signal_indexes:
        candles = prices[index - required + 1:index + 1]
        result = evaluate_three_day_dip(candles, params, enforce_volume=True)
        if not result["matched"]:
            continue
        hits.append({
            "signal_date": result["signal"]["trade_date"],
            "bearish_ok": result["bearish_ok"],
            "declining_ok": result["declining_ok"],
            "new_low_ok": result["new_low_ok"],
            "pattern_type": result["pattern_type"],
            "decline_ratio": result["decline_ratio"],
            "recovery_range_ratio": result["recovery_range_ratio"],
            "repair_ratio": result["repair_ratio"],
            "body_range_ratio": result["body_range_ratio"],
            "signal_change_ratio": result["signal_change_ratio"],
            "volume_ratio": result["volume_ratio"],
            "candles": candles,
        })
    dates_in_range = [price["trade_date"] for price in prices if start_date <= price["trade_date"] <= end_date]
    return {
        "stock_code": stock["stock_code"],
        "stock_name": stock["stock_name"],
        "start_date": start_date,
        "end_date": end_date,
        "scanned_days": len(signal_indexes),
        "available_days": len(dates_in_range),
        "hits": hits,
        "insufficient": not signal_indexes,
        "first_price_date": prices[0]["trade_date"] if prices else None,
        "last_price_date": prices[-1]["trade_date"] if prices else None,
    }


def three_day_dip_metrics(result: dict) -> dict:
    return {
        key: result.get(key) for key in (
            "pattern_type", "decline_ratio", "signal_low_above_prior_ratio", "recovery_range_ratio", "repair_ratio",
            "body_range_ratio", "lower_shadow_body_ratio", "signal_change_ratio", "volume_ratio",
        )
    }


def update_alert_signal_outcome(db: sqlite3.Connection, sample_id: int) -> None:
    sample = db.execute(
        "SELECT stock_code, signal_date, candles_json FROM alert_signal_samples WHERE id = ?", (sample_id,)
    ).fetchone()
    if sample is None:
        return
    candles = json.loads(sample["candles_json"])
    signal_close = float(candles[-1]["close"])
    future = db.execute(
        """SELECT trade_date, high, low, close FROM daily_prices
        WHERE stock_code = ? AND trade_date > ? AND high IS NOT NULL AND low IS NOT NULL
        AND close IS NOT NULL ORDER BY trade_date LIMIT 10""",
        (sample["stock_code"], sample["signal_date"]),
    ).fetchall()
    values: dict[str, float | None] = {}
    for horizon in (1, 3, 5, 10):
        window = future[:horizon]
        values[f"day_{horizon}_close_return"] = (
            float(window[-1]["close"]) / signal_close - 1 if len(window) >= horizon else None
        )
        values[f"day_{horizon}_max_return"] = (
            max(float(row["high"]) for row in window) / signal_close - 1 if window else None
        )
        values[f"day_{horizon}_max_drawdown"] = (
            min(float(row["low"]) for row in window) / signal_close - 1 if window else None
        )
    now = datetime.now().isoformat(timespec="seconds")
    columns = [
        f"day_{horizon}_{metric}" for horizon in (1, 3, 5, 10)
        for metric in ("close_return", "max_return", "max_drawdown")
    ]
    db.execute(
        f"""INSERT INTO alert_signal_outcomes
        (sample_id, {', '.join(columns)}, evaluated_through_date, updated_at)
        VALUES (?, {', '.join('?' for _ in columns)}, ?, ?)
        ON CONFLICT(sample_id) DO UPDATE SET
        {', '.join(f'{column} = excluded.{column}' for column in columns)},
        evaluated_through_date = excluded.evaluated_through_date, updated_at = excluded.updated_at""",
        (sample_id, *(values[column] for column in columns), future[-1]["trade_date"] if future else None, now),
    )


def refresh_alert_signal_outcomes(db: sqlite3.Connection, stock_codes: list[str] | None = None) -> int:
    if stock_codes:
        placeholders = ",".join("?" for _ in stock_codes)
        rows = db.execute(
            f"SELECT id FROM alert_signal_samples WHERE stock_code IN ({placeholders})", tuple(stock_codes)
        ).fetchall()
    else:
        rows = db.execute("SELECT id FROM alert_signal_samples").fetchall()
    for row in rows:
        update_alert_signal_outcome(db, int(row["id"]))
    return len(rows)


def upsert_alert_signal_sample(
    db: sqlite3.Connection, user_id: int, alert_type_id: int, stock_code: str, stock_name: str,
    candles: list[dict], result: dict, source: str = "AUTO", review_status: str = "PENDING",
    note: str = "", pattern_type: str | None = None,
) -> int:
    now = datetime.now().isoformat(timespec="seconds")
    signal_date = candles[-1]["trade_date"]
    pattern_type = pattern_type or result.get("pattern_type") or "EXHAUSTION"
    cursor = db.execute(
        """INSERT INTO alert_signal_samples
        (user_id, alert_type_id, stock_code, stock_name, signal_date, pattern_type, source,
         review_status, rule_version, metrics_json, candles_json, note, created_at, updated_at, reviewed_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id, alert_type_id, stock_code, signal_date) DO UPDATE SET
        stock_name = excluded.stock_name, pattern_type = excluded.pattern_type,
        source = CASE WHEN excluded.source = 'MANUAL' THEN 'MANUAL' ELSE alert_signal_samples.source END,
        review_status = CASE WHEN excluded.source = 'MANUAL' THEN excluded.review_status ELSE alert_signal_samples.review_status END,
        rule_version = excluded.rule_version, metrics_json = excluded.metrics_json,
        candles_json = excluded.candles_json, updated_at = excluded.updated_at,
        reviewed_at = CASE WHEN excluded.source = 'MANUAL' THEN excluded.reviewed_at ELSE alert_signal_samples.reviewed_at END,
        note = CASE WHEN excluded.source = 'MANUAL' OR alert_signal_samples.note = '' THEN excluded.note ELSE alert_signal_samples.note END
        RETURNING id""",
        (
            user_id, alert_type_id, stock_code, stock_name, signal_date, pattern_type, source,
            review_status, THREE_DAY_DIP_RULE_VERSION,
            json.dumps(three_day_dip_metrics(result), ensure_ascii=False),
            json.dumps(candles, ensure_ascii=False), note, now, now,
            now if review_status != "PENDING" else None,
        ),
    )
    sample_id = int(cursor.fetchone()["id"])
    update_alert_signal_outcome(db, sample_id)
    return sample_id


def alert_signal_pool_data(db: sqlite3.Connection, user_id: int) -> dict:
    rows = [dict(row) for row in db.execute(
        """SELECT s.*, o.day_1_close_return, o.day_3_close_return, o.day_5_close_return,
        o.day_10_close_return, o.day_3_max_return, o.day_3_max_drawdown
        FROM alert_signal_samples s LEFT JOIN alert_signal_outcomes o ON o.sample_id = s.id
        WHERE s.user_id = ? ORDER BY s.signal_date DESC, s.id DESC""", (user_id,)
    ).fetchall()]
    for row in rows:
        row["metrics"] = json.loads(row.pop("metrics_json"))
        row["candles"] = json.loads(row.pop("candles_json"))
    reviewed = [row for row in rows if row["review_status"] != "PENDING"]
    confirmed = [row for row in rows if row["review_status"] == "CONFIRMED"]
    def performance(days: int) -> dict:
        key = f"day_{days}_close_return"
        values = [float(row[key]) for row in confirmed if row.get(key) is not None]
        return {
            "samples": len(values), "average": sum(values) / len(values) if values else None,
            "win_rate": sum(value > 0 for value in values) / len(values) if values else None,
        }
    return {
        "rows": rows,
        "summary": {
            "total": len(rows), "pending": sum(row["review_status"] == "PENDING" for row in rows),
            "confirmed": len(confirmed), "rejected": sum(row["review_status"] == "REJECTED" for row in rows),
            "accuracy": len(confirmed) / len(reviewed) if reviewed else None,
        },
        "performance": {days: performance(days) for days in (1, 3, 5)},
    }


def notification_rows(db: sqlite3.Connection, user_id: int, limit: int = 20) -> tuple[int, list[dict]]:
    unread = int(db.execute(
        "SELECT COUNT(*) FROM notifications WHERE user_id = ? AND read_at IS NULL", (user_id,)
    ).fetchone()[0])
    rows = [dict(row) for row in db.execute(
        """SELECT id, stock_code, stock_name, stage, title, content, quote_time, created_at, read_at
        FROM notifications WHERE user_id = ? ORDER BY created_at DESC, id DESC LIMIT ?""",
        (user_id, limit),
    ).fetchall()]
    return unread, rows


def notification_page_data(db: sqlite3.Connection, user_id: int, page: int, per_page: int = 20) -> dict:
    unread = int(db.execute(
        "SELECT COUNT(*) FROM notifications WHERE user_id = ? AND read_at IS NULL", (user_id,)
    ).fetchone()[0])
    total = int(db.execute(
        "SELECT COUNT(*) FROM notifications WHERE user_id = ?", (user_id,)
    ).fetchone()[0])
    total_pages = max(1, math.ceil(total / per_page))
    page = min(max(1, page), total_pages)
    offset = (page - 1) * per_page
    rows = [dict(row) for row in db.execute(
        """SELECT id, stock_code, stock_name, stage, title, content, quote_time, created_at, read_at
        FROM notifications WHERE user_id = ? ORDER BY created_at DESC, id DESC LIMIT ? OFFSET ?""",
        (user_id, per_page, offset),
    ).fetchall()]
    return {
        "unread": unread, "total": total, "per_page": per_page,
        "page": page, "total_pages": total_pages, "rows": rows,
    }


def review_progress_data(db: sqlite3.Connection, user_id: int) -> dict:
    row = db.execute(
        """SELECT COUNT(*) total,
        COALESCE(SUM(CASE WHEN r.review_status = 'COMPLETED' THEN 1 ELSE 0 END), 0) completed
        FROM trade_episodes ep
        LEFT JOIN trade_reviews r ON r.trade_episode_id = ep.id AND r.user_id = ep.user_id
        WHERE ep.user_id = ? AND ep.status = 'CLOSED'""",
        (user_id,),
    ).fetchone()
    total = int(row["total"])
    completed = int(row["completed"])
    pending = max(0, total - completed)
    rate = completed / total if total else 0.0
    return {
        "total": total,
        "completed": completed,
        "pending": pending,
        "rate": rate,
    }


def trading_system_stats(db: sqlite3.Connection, user_id: int) -> dict:
    rows = db.execute(
        """SELECT ep.closed_at,
        COALESCE((SELECT SUM(f.profit) FROM fifo_matches f
            JOIN trade_episode_executions link ON link.execution_id = f.buy_execution_id
            WHERE link.trade_episode_id = ep.id), 0) profit
        FROM trade_episodes ep
        WHERE ep.user_id = ? AND ep.status = 'CLOSED'
        ORDER BY ep.closed_at, ep.id""",
        (user_id,),
    ).fetchall()
    profits = [float(row["profit"]) for row in rows]
    wins = [p for p in profits if p > 0]
    losses = [p for p in profits if p < 0]
    avg_win = sum(wins) / len(wins) if wins else None
    avg_loss = sum(losses) / len(losses) if losses else None
    if avg_win is not None and avg_loss is not None and avg_loss != 0:
        profit_factor = avg_win / abs(avg_loss)
    else:
        profit_factor = None

    max_win_streak = max_loss_streak = 0
    cur_win = cur_loss = 0
    for profit in profits:
        if profit > 0:
            cur_win += 1
            cur_loss = 0
        elif profit < 0:
            cur_loss += 1
            cur_win = 0
        else:
            cur_win = cur_loss = 0
        max_win_streak = max(max_win_streak, cur_win)
        max_loss_streak = max(max_loss_streak, cur_loss)

    return {
        "total": len(profits),
        "avg_win": avg_win,
        "avg_loss": avg_loss,
        "profit_factor": profit_factor,
        "max_win_streak": max_win_streak,
        "max_loss_streak": max_loss_streak,
    }


def strategy_type_stats(db: sqlite3.Connection, user_id: int) -> list[dict]:
    rows = db.execute(
        """WITH episode_stats AS (
            SELECT ep.id,
            COALESCE((SELECT SUM(f.profit) FROM fifo_matches f
                JOIN trade_episode_executions link ON link.execution_id = f.buy_execution_id
                WHERE link.trade_episode_id = ep.id), 0) profit,
            COALESCE((SELECT SUM(f.buy_cost) FROM fifo_matches f
                JOIN trade_episode_executions link ON link.execution_id = f.buy_execution_id
                WHERE link.trade_episode_id = ep.id), 0) buy_cost,
            CAST(julianday(SUBSTR(ep.closed_at, 1, 10)) - julianday(SUBSTR(ep.opened_at, 1, 10)) AS INTEGER) holding_days
            FROM trade_episodes ep
            WHERE ep.user_id = ? AND ep.status = 'CLOSED'
        )
        SELECT rrt.reason_type,
        COUNT(*) trades,
        COALESCE(SUM(CASE WHEN es.profit > 0 THEN 1 ELSE 0 END), 0) wins,
        COALESCE(SUM(es.profit), 0) total_profit,
        COALESCE(AVG(es.profit), 0) avg_profit,
        COALESCE(AVG(CASE WHEN es.profit > 0 THEN es.profit END), 0) avg_win,
        COALESCE(AVG(CASE WHEN es.profit < 0 THEN es.profit END), 0) avg_loss,
        COALESCE(AVG(es.holding_days), 0) avg_holding,
        COALESCE(AVG(CASE WHEN es.buy_cost > 0 THEN es.profit / es.buy_cost END), 0) avg_return
        FROM trade_review_reason_types rrt
        JOIN trade_reviews r ON r.id = rrt.trade_review_id AND r.user_id = ?
        JOIN episode_stats es ON es.id = r.trade_episode_id
        GROUP BY rrt.reason_type
        ORDER BY total_profit DESC, trades DESC""",
        (user_id, user_id),
    ).fetchall()
    stats = []
    for row in rows:
        trades = int(row["trades"])
        wins = int(row["wins"])
        avg_win = float(row["avg_win"])
        avg_loss = float(row["avg_loss"])
        profit_factor = (avg_win / abs(avg_loss)) if avg_win and avg_loss and avg_loss != 0 else None
        stats.append({
            "reason_type": row["reason_type"],
            "trades": trades,
            "wins": wins,
            "win_rate": wins / trades if trades else 0.0,
            "avg_return": float(row["avg_return"]),
            "avg_profit": float(row["avg_profit"]),
            "avg_win": avg_win,
            "avg_loss": avg_loss,
            "profit_factor": profit_factor,
            "avg_holding": float(row["avg_holding"]),
            "total_profit": float(row["total_profit"]),
        })
    return stats


def mistake_stats(db: sqlite3.Connection, user_id: int) -> dict:
    total_reviewed = db.execute(
        """SELECT COUNT(*) FROM trade_reviews r
        JOIN trade_episodes ep ON ep.id = r.trade_episode_id AND ep.user_id = r.user_id
        WHERE r.user_id = ? AND r.review_status = 'COMPLETED' AND ep.status = 'CLOSED'""",
        (user_id,),
    ).fetchone()[0]
    rows = db.execute(
        """WITH reviewed AS (
            SELECT r.main_problem,
            COALESCE((SELECT SUM(f.profit) FROM fifo_matches f
                JOIN trade_episode_executions link ON link.execution_id = f.buy_execution_id
                WHERE link.trade_episode_id = ep.id), 0) profit
            FROM trade_reviews r
            JOIN trade_episodes ep ON ep.id = r.trade_episode_id AND ep.user_id = r.user_id
            WHERE r.user_id = ? AND r.review_status = 'COMPLETED' AND ep.status = 'CLOSED'
              AND r.main_problem IS NOT NULL
        )
        SELECT main_problem, COUNT(*) count,
        COALESCE(SUM(profit), 0) total_profit,
        COALESCE(SUM(CASE WHEN profit < 0 THEN profit ELSE 0 END), 0) total_loss
        FROM reviewed GROUP BY main_problem ORDER BY count DESC, total_loss ASC""",
        (user_id,),
    ).fetchall()
    total_reviewed = int(total_reviewed)
    items = []
    for row in rows:
        count = int(row["count"])
        items.append({
            "main_problem": row["main_problem"],
            "count": count,
            "total_profit": float(row["total_profit"] or 0),
            "total_loss": float(row["total_loss"] or 0),
            "rate": count / total_reviewed if total_reviewed else 0.0,
        })
    return {"total_reviewed": total_reviewed, "items": items}


def review_episodes(
    db: sqlite3.Connection,
    user_id: int,
    status_filter: str,
    start_date: str | None = None,
    end_date: str | None = None,
) -> list[sqlite3.Row]:
    if status_filter == "completed":
        status_condition = "r.review_status = 'COMPLETED'"
    elif status_filter == "all":
        status_condition = "1=1"
    else:
        status_condition = "COALESCE(r.review_status, 'PENDING') != 'COMPLETED'"
    where = ["ep.user_id = ?", "ep.status = 'CLOSED'", f"({status_condition})"]
    parameters: list[object] = [user_id]
    if start_date:
        where.append("date(ep.closed_at) >= ?")
        parameters.append(start_date)
    if end_date:
        where.append("date(ep.closed_at) <= ?")
        parameters.append(end_date)
    return db.execute(
        f"""SELECT ep.id, ep.stock_code, ep.stock_name, ep.opened_at, ep.closed_at, ep.status,
        r.review_status, r.updated_at, r.judgement_result, r.main_problem,
        r.review_note, r.next_action, r.completed_at, r.trade_reason, r.sell_reason,
        r.confidence_level, r.expected_profit_percent, r.expected_target_price,
        r.stop_loss_price, r.expected_holding_days,
        (SELECT GROUP_CONCAT(rrt.reason_type, ',') FROM trade_review_reason_types rrt
            WHERE rrt.trade_review_id = r.id) reason_types,
        COALESCE(SUM(f.profit), 0) profit,
        COALESCE(SUM(f.buy_cost), 0) buy_cost,
        CAST(julianday(SUBSTR(ep.closed_at, 1, 10)) - julianday(SUBSTR(ep.opened_at, 1, 10)) AS INTEGER) holding_days
        FROM trade_episodes ep
        LEFT JOIN trade_reviews r ON r.trade_episode_id = ep.id AND r.user_id = ep.user_id
        LEFT JOIN trade_episode_executions link ON link.trade_episode_id = ep.id AND link.role IN ('ENTRY', 'ADD')
        LEFT JOIN fifo_matches f ON f.user_id = ep.user_id AND f.buy_execution_id = link.execution_id
        WHERE {' AND '.join(where)}
        GROUP BY ep.id ORDER BY ep.closed_at DESC, ep.id DESC""",
        parameters,
    ).fetchall()


def price_sync_start_date(db: sqlite3.Connection, stock_code: str, today: date | None = None) -> str:
    today = today or date.today()
    latest = db.execute(
        "SELECT MAX(trade_date) FROM daily_prices WHERE stock_code = ?",
        (stock_code,),
    ).fetchone()[0]
    if latest:
        try:
            return (date.fromisoformat(latest) - timedelta(days=PRICE_SYNC_OVERLAP_DAYS)).isoformat()
        except ValueError:
            pass
    return (today - timedelta(days=PRICE_SYNC_INITIAL_DAYS)).isoformat()


def market_symbol(stock_code: str) -> str:
    if stock_code.startswith(("sh", "sz", "bj")):
        return stock_code
    if stock_code.startswith("920") or stock_code.startswith(("4", "8")):
        return f"bj{stock_code}"
    if stock_code.startswith(("5", "6", "9")):
        return f"sh{stock_code}"
    return f"sz{stock_code}"


def fetch_realtime_prices(stock_codes: list[str]) -> dict[str, dict]:
    if not stock_codes:
        return {}
    symbols = ",".join(market_symbol(stock_code) for stock_code in stock_codes)
    request_data = Request(
        f"https://qt.gtimg.cn/q={symbols}",
        headers={"User-Agent": "Mozilla/5.0", "Referer": "https://gu.qq.com/"},
    )
    with urlopen(request_data, timeout=15, context=SSL_CONTEXT) as response:
        text = response.read().decode("gbk", errors="replace")

    requested_symbols = {market_symbol(stock_code): stock_code for stock_code in stock_codes}
    quotes = {}
    for raw_record in text.split(";"):
        if '="' not in raw_record:
            continue
        variable, raw_value = raw_record.split('="', 1)
        variable = variable.strip()
        requested_code = requested_symbols.get(variable.removeprefix("v_"))
        if requested_code is None:
            continue
        fields = raw_value.rstrip('"\r\n').split("~")
        if len(fields) < 38 or len(fields[30]) < 14:
            continue
        try:
            timestamp = fields[30]
            trade_date = f"{timestamp[:4]}-{timestamp[4:6]}-{timestamp[6:8]}"
            quote_time = f"{trade_date}T{timestamp[8:10]}:{timestamp[10:12]}:{timestamp[12:14]}"
            quotes[requested_code] = {
                "trade_date": trade_date,
                "quote_time": quote_time,
                "open": float(fields[5]),
                "high": float(fields[33]),
                "low": float(fields[34]),
                "close": float(fields[3]),
                "previous_close": float(fields[4]),
                "volume": float(fields[36]),
                "amount": float(fields[35].split("/")[2]),
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
            }
        except (IndexError, ValueError):
            continue
    return quotes


MARKET_INDEXES = {
    "sh000001": "上证指数",
    "sz399001": "深证成指",
    "sz399006": "创业板指",
    "sh000688": "科创50",
    "sh000016": "上证50",
    "sh000300": "沪深300",
    "sh000852": "中证1000",
}


def fetch_market_indexes() -> list[dict]:
    quotes = fetch_realtime_prices(list(MARKET_INDEXES))
    indexes = []
    for code, name in MARKET_INDEXES.items():
        quote = quotes.get(code)
        if quote is None:
            continue
        previous_close = float(quote.get("previous_close") or 0)
        close = float(quote["close"])
        change = close - previous_close if previous_close else None
        indexes.append({
            "code": code,
            "name": name,
            "price": close,
            "change": change,
            "change_rate": change / previous_close if change is not None and previous_close else None,
            "trade_date": quote["trade_date"],
            "fetched_at": quote["fetched_at"],
        })
    return indexes


def sync_market_indexes() -> tuple[list[dict], str | None]:
    try:
        indexes = fetch_market_indexes()
        if not indexes:
            raise RuntimeError("未返回指数行情")
        updated_at = datetime.now().isoformat(timespec="seconds")
        with MARKET_INDEX_STATE_LOCK:
            MARKET_INDEX_STATE.update({"data": indexes, "updated_at": updated_at, "error": None})
        return indexes, None
    except Exception as error:
        message = str(error)
        with MARKET_INDEX_STATE_LOCK:
            MARKET_INDEX_STATE["error"] = message
        return [], message


def market_index_status() -> dict:
    with MARKET_INDEX_STATE_LOCK:
        return {"data": list(MARKET_INDEX_STATE["data"]), "updated_at": MARKET_INDEX_STATE["updated_at"], "error": MARKET_INDEX_STATE["error"]}


def save_realtime_prices(db: sqlite3.Connection, quotes: dict[str, dict]) -> int:
    for stock_code, quote in quotes.items():
        db.execute(
            """INSERT INTO daily_prices
            (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'tencent-realtime', ?)
            ON CONFLICT(stock_code, trade_date) DO UPDATE SET
            open = excluded.open, high = excluded.high, low = excluded.low,
            close = excluded.close, volume = excluded.volume, amount = excluded.amount,
            source = excluded.source, fetched_at = excluded.fetched_at""",
            (stock_code, quote["trade_date"], quote["open"], quote["high"], quote["low"],
             quote["close"], quote["volume"], quote["amount"], quote["fetched_at"]),
        )
    return len(quotes)


def save_intraday_quotes(db: sqlite3.Connection, quotes: dict[str, dict]) -> int:
    saved = 0
    for stock_code, quote in quotes.items():
        quote_time = quote.get("quote_time", quote["fetched_at"])
        try:
            moment = datetime.fromisoformat(quote_time)
            price = float(quote["close"])
        except (KeyError, TypeError, ValueError):
            continue
        if price <= 0 or not is_auto_sync_time(moment):
            continue
        quote_minute = moment.replace(second=0, microsecond=0).isoformat(timespec="minutes")
        db.execute(
            """INSERT INTO intraday_quotes
            (stock_code, trade_date, quote_minute, price, day_high, day_low, previous_close,
             volume, amount, source, fetched_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'tencent-realtime', ?)
            ON CONFLICT(stock_code, quote_minute) DO UPDATE SET
            price = excluded.price, day_high = excluded.day_high, day_low = excluded.day_low,
            previous_close = excluded.previous_close, volume = excluded.volume,
            amount = excluded.amount, source = excluded.source, fetched_at = excluded.fetched_at""",
            (
                stock_code, quote["trade_date"], quote_minute, price, quote.get("high"), quote.get("low"),
                quote.get("previous_close"), quote.get("volume"), quote.get("amount"), quote["fetched_at"],
            ),
        )
        saved += 1
    return saved


def is_closing_quote(quote: dict) -> bool:
    try:
        moment = datetime.fromisoformat(quote["fetched_at"])
    except (KeyError, TypeError, ValueError):
        return False
    return moment.hour >= 15


def create_alert_notification(
    db: sqlite3.Connection, alert_type_id: int, watch: sqlite3.Row, quote: dict,
    stage: str, result: dict,
) -> bool:
    created_at = datetime.now().isoformat(timespec="seconds")
    stage_label = "盘中候选" if stage == "CANDIDATE" else "收盘确认"
    title = f"{watch['stock_name']}触发三日低吸·{stage_label}"
    pattern_label = {
        "STRONG_REVERSAL": "强修复型", "SHADOW_STOP": "下影止跌型",
        "EXHAUSTION": "衰竭止跌型",
    }[result["pattern_type"]]
    content = (
        f"{pattern_label}：前两日累计下跌 {result['decline_ratio'] * 100:.1f}%，"
        f"今日最低 {float(quote['low']):.2f}，当前价 {float(quote['close']):.2f}，"
        f"低点修复 {result['repair_ratio'] * 100:.1f}%。"
    )
    details = {
        "low": float(quote["low"]),
        "price": float(quote["close"]),
        "recovery_range_ratio": result["recovery_range_ratio"],
        "pattern_type": result["pattern_type"],
        "decline_ratio": result["decline_ratio"],
        "repair_ratio": result["repair_ratio"],
    }
    dedupe_key = f"{watch['user_id']}:{THREE_DAY_DIP_CODE}:{watch['stock_code']}:{quote['trade_date']}:{stage}"
    cursor = db.execute(
        """INSERT OR IGNORE INTO notifications
        (user_id, alert_type_id, stock_code, stock_name, stage, title, content,
         details_json, quote_time, created_at, dedupe_key)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            watch["user_id"], alert_type_id, watch["stock_code"], watch["stock_name"], stage,
            title, content, json.dumps(details, ensure_ascii=False), quote["fetched_at"], created_at, dedupe_key,
        ),
    )
    return cursor.rowcount > 0


def create_intraday_rebound_notification(
    db: sqlite3.Connection, alert_type_id: int, watch: sqlite3.Row, quote: dict, result: dict,
) -> bool:
    created_at = datetime.now().isoformat(timespec="seconds")
    content = (
        f"低点 {result['trough_price']:.3f}（{result['trough_minute'][11:16]}），"
        f"突破 {result['breakout_price']:.3f}，当前 {float(quote['close']):.3f}；"
        f"低点反弹 {result['recovery_ratio'] * 100:.1f}%"
    )
    details = {
        "trough_price": result["trough_price"],
        "trough_minute": result["trough_minute"],
        "higher_low": result["higher_low"],
        "breakout_price": result["breakout_price"],
        "drop_ratio": result["drop_ratio"],
        "recovery_ratio": result["recovery_ratio"],
        "volume_multiple": result["volume_multiple"],
        "invalidation_price": result["higher_low"],
    }
    dedupe_key = f"{watch['user_id']}:{INTRADAY_REBOUND_CODE}:{watch['stock_code']}:{quote['trade_date']}:CANDIDATE"
    cursor = db.execute(
        """INSERT OR IGNORE INTO notifications
        (user_id, alert_type_id, stock_code, stock_name, stage, title, content,
         details_json, quote_time, created_at, dedupe_key)
        VALUES (?, ?, ?, ?, 'CANDIDATE', ?, ?, ?, ?, ?, ?)""",
        (
            watch["user_id"], alert_type_id, watch["stock_code"], watch["stock_name"],
            f"{watch['stock_name']}触发日内反弹·候选", content,
            json.dumps(details, ensure_ascii=False), quote.get("quote_time", quote["fetched_at"]), created_at, dedupe_key,
        ),
    )
    return cursor.rowcount > 0


def evaluate_realtime_alerts(db: sqlite3.Connection, quotes: dict[str, dict]) -> int:
    if not quotes:
        return 0
    alert_type = db.execute(
        "SELECT * FROM alert_types WHERE code = ? AND enabled = 1", (THREE_DAY_DIP_CODE,)
    ).fetchone()
    if alert_type is None:
        return 0
    params = three_day_dip_params(json.loads(alert_type["params_json"]))
    placeholders = ",".join("?" for _ in quotes)
    watches = db.execute(
        f"""SELECT user_id, stock_code, stock_name FROM watchlist_stocks
        WHERE stock_code IN ({placeholders}) ORDER BY user_id, stock_code""",
        tuple(quotes),
    ).fetchall()
    created = 0
    for watch in watches:
        quote = quotes[watch["stock_code"]]
        prices = [dict(row) for row in reversed(db.execute(
            """SELECT trade_date, open, high, low, close, volume FROM daily_prices
            WHERE stock_code = ? AND trade_date <= ? AND open IS NOT NULL AND high IS NOT NULL
            AND low IS NOT NULL AND close IS NOT NULL ORDER BY trade_date DESC LIMIT ?""",
            (watch["stock_code"], quote["trade_date"], params["decline_days"] + 2),
        ).fetchall())]
        closing = is_closing_quote(quote)
        result = evaluate_three_day_dip(prices, params, enforce_volume=closing)
        state = db.execute(
            """SELECT * FROM alert_rule_states WHERE user_id = ? AND alert_type_id = ?
            AND stock_code = ? AND trade_date = ?""",
            (watch["user_id"], alert_type["id"], watch["stock_code"], quote["trade_date"]),
        ).fetchone()
        candidate_at = state["candidate_triggered_at"] if state else None
        confirmation_at = state["confirmation_triggered_at"] if state else None
        if result["matched"] and params["intraday_candidate_enabled"] and not candidate_at:
            if create_alert_notification(db, alert_type["id"], watch, quote, "CANDIDATE", result):
                created += 1
            candidate_at = quote["fetched_at"]
        if (
            result["matched"] and params["close_confirmation_enabled"]
            and result["pattern_type"] != "SHADOW_STOP" and closing and not confirmation_at
        ):
            if create_alert_notification(db, alert_type["id"], watch, quote, "CONFIRMED", result):
                created += 1
            confirmation_at = quote["fetched_at"]
        if result["matched"] and closing:
            upsert_alert_signal_sample(
                db, watch["user_id"], alert_type["id"], watch["stock_code"], watch["stock_name"],
                prices[-(params["decline_days"] + 2):], result,
            )
        db.execute(
            """INSERT INTO alert_rule_states
            (user_id, alert_type_id, stock_code, trade_date, candidate_triggered_at,
             confirmation_triggered_at, last_evaluated_at, last_matched)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(user_id, alert_type_id, stock_code, trade_date) DO UPDATE SET
            candidate_triggered_at = excluded.candidate_triggered_at,
            confirmation_triggered_at = excluded.confirmation_triggered_at,
            last_evaluated_at = excluded.last_evaluated_at, last_matched = excluded.last_matched""",
            (
                watch["user_id"], alert_type["id"], watch["stock_code"], quote["trade_date"],
                candidate_at, confirmation_at, quote["fetched_at"], int(result["matched"]),
            ),
        )
    return created


def evaluate_intraday_rebound_alerts(db: sqlite3.Connection, quotes: dict[str, dict]) -> int:
    if not quotes:
        return 0
    alert_type = db.execute(
        "SELECT * FROM alert_types WHERE code = ? AND enabled = 1", (INTRADAY_REBOUND_CODE,)
    ).fetchone()
    if alert_type is None:
        return 0
    params = intraday_rebound_params(json.loads(alert_type["params_json"]))
    if not params["enabled"]:
        return 0
    placeholders = ",".join("?" for _ in quotes)
    watches = db.execute(
        f"""SELECT user_id, stock_code, stock_name FROM watchlist_stocks
        WHERE stock_code IN ({placeholders}) ORDER BY user_id, stock_code""",
        tuple(quotes),
    ).fetchall()
    results = {}
    created = 0
    for watch in watches:
        quote = quotes[watch["stock_code"]]
        if watch["stock_code"] not in results:
            samples = [dict(row) for row in db.execute(
                """SELECT quote_minute, price, volume FROM intraday_quotes
                WHERE stock_code = ? AND trade_date = ? ORDER BY quote_minute DESC LIMIT ?""",
                (watch["stock_code"], quote["trade_date"], params["lookback_minutes"]),
            ).fetchall()][::-1]
            results[watch["stock_code"]] = evaluate_intraday_rebound(samples, params)
        result = results[watch["stock_code"]]
        state = db.execute(
            """SELECT * FROM alert_rule_states WHERE user_id = ? AND alert_type_id = ?
            AND stock_code = ? AND trade_date = ?""",
            (watch["user_id"], alert_type["id"], watch["stock_code"], quote["trade_date"]),
        ).fetchone()
        candidate_at = state["candidate_triggered_at"] if state else None
        if result["matched"] and not candidate_at:
            if create_intraday_rebound_notification(db, alert_type["id"], watch, quote, result):
                created += 1
            candidate_at = quote.get("quote_time", quote["fetched_at"])
        db.execute(
            """INSERT INTO alert_rule_states
            (user_id, alert_type_id, stock_code, trade_date, candidate_triggered_at,
             confirmation_triggered_at, last_evaluated_at, last_matched)
            VALUES (?, ?, ?, ?, ?, NULL, ?, ?)
            ON CONFLICT(user_id, alert_type_id, stock_code, trade_date) DO UPDATE SET
            candidate_triggered_at = excluded.candidate_triggered_at,
            last_evaluated_at = excluded.last_evaluated_at, last_matched = excluded.last_matched""",
            (
                watch["user_id"], alert_type["id"], watch["stock_code"], quote["trade_date"],
                candidate_at, quote.get("quote_time", quote["fetched_at"]), int(result["matched"]),
            ),
        )
    return created


def fetch_daily_prices(stock_code: str, start_date: str, end_date: str):
    import akshare as ak

    symbol = market_symbol(stock_code)

    frame = None
    source = None
    errors = []
    compact_start = start_date.replace("-", "")
    compact_end = end_date.replace("-", "")
    providers = (
        ("akshare-tencent", lambda: ak.stock_zh_a_hist_tx(
            symbol=symbol, start_date=compact_start, end_date=compact_end, adjust="qfq", timeout=15,
        )),
        ("akshare-sina", lambda: ak.stock_zh_a_daily(
            symbol=symbol, start_date=compact_start, end_date=compact_end, adjust="qfq",
        )),
        ("akshare-eastmoney", lambda: ak.stock_zh_a_hist(
            symbol=stock_code, period="daily", start_date=compact_start, end_date=compact_end,
            adjust="qfq", timeout=15,
        )),
        ("akshare-eastmoney", lambda: ak.fund_etf_hist_em(
            symbol=stock_code, period="daily", start_date=compact_start, end_date=compact_end, adjust="qfq",
        )),
    )
    for provider_source, provider in providers:
        try:
            candidate = provider()
            if candidate is not None and not getattr(candidate, "empty", True):
                frame = candidate
                source = provider_source
                break
        except Exception as error:
            errors.append(f"{provider_source}: {error}")
    if frame is None and errors:
        raise RuntimeError("AkShare 行情接口请求失败：" + "；".join(errors))
    if frame is None:
        return None, None
    return frame, source


def fetch_hot_industry():
    import akshare as ak

    frame = ak.stock_sector_spot("新浪行业")
    return normalize_board_frame(frame)


def fetch_hot_concept():
    import akshare as ak

    frame = ak.stock_sector_spot("概念")
    return normalize_board_frame(frame)


def normalize_board_frame(frame):
    if frame is None or getattr(frame, "empty", True):
        return frame
    frame = frame.sort_values("涨跌幅", ascending=False).reset_index(drop=True)
    return frame.rename(columns={
        "板块": "板块名称", "label": "板块代码", "平均价格": "最新价",
        "公司家数": "公司家数", "总成交额": "成交额",
        "股票名称": "领涨股票", "个股-涨跌幅": "领涨股票-涨跌幅",
    })


def build_hot_rank_records(rank_rows: list[dict], quotes: dict[str, dict]) -> list[dict]:
    records = []
    for item in sorted(rank_rows, key=lambda entry: entry["rank"]):
        quote = quotes.get(item["code"])
        if quote is None:
            records.append({
                "排名": item["rank"], "代码": item["code"], "股票名称": item["name"],
                "最新价": None, "涨跌额": None, "涨跌幅": None,
            })
            continue
        close = float(quote["close"])
        previous_close = float(quote.get("previous_close") or 0)
        change = close - previous_close if previous_close else None
        change_rate = change / previous_close * 100 if change is not None and previous_close else None
        records.append({
            "排名": item["rank"], "代码": item["code"], "股票名称": item["name"],
            "最新价": round(close, 4),
            "涨跌额": round(change, 4) if change is not None else None,
            "涨跌幅": round(change_rate, 4) if change_rate is not None else None,
        })
    return records


def fetch_hot_rank():
    payload = {
        "appId": "appId01",
        "globalId": "786e4c21-70dc-435a-93bb-38",
        "marketType": "",
        "pageNo": 1,
        "pageSize": 20,
    }
    request = Request(
        "https://emappdata.eastmoney.com/stockrank/getAllCurrentList",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
    )
    with urlopen(request, timeout=15, context=SSL_CONTEXT) as response:
        raw_rows = json.loads(response.read().decode("utf-8"))["data"]
    rank_rows = [
        {"rank": int(row["rk"]), "code": str(row["sc"])[-6:], "name": str(row["sc"]).lower()}
        for row in raw_rows
    ]
    quotes = fetch_realtime_prices([item["code"] for item in rank_rows])
    name_rows = fetch_tencent_names([item["code"] for item in rank_rows])
    for item in rank_rows:
        if item["code"] in name_rows:
            item["name"] = name_rows[item["code"]]
    import pandas as pd

    return pd.DataFrame(build_hot_rank_records(rank_rows, quotes))


def fetch_tencent_names(stock_codes: list[str]) -> dict[str, str]:
    symbols = ",".join(market_symbol(stock_code) for stock_code in stock_codes)
    request = Request(
        f"https://qt.gtimg.cn/q={symbols}",
        headers={"User-Agent": "Mozilla/5.0", "Referer": "https://gu.qq.com/"},
    )
    with urlopen(request, timeout=15, context=SSL_CONTEXT) as response:
        text = response.read().decode("gbk", errors="replace")
    requested_symbols = {market_symbol(stock_code): stock_code for stock_code in stock_codes}
    names = {}
    for raw_record in text.split(";"):
        if '="' not in raw_record:
            continue
        variable, raw_value = raw_record.split('="', 1)
        variable = variable.strip()
        requested_code = requested_symbols.get(variable.removeprefix("v_"))
        if requested_code is None:
            continue
        fields = raw_value.rstrip('"\r\n').split("~")
        if len(fields) >= 2:
            names[requested_code] = fields[1]
    return names



def df_records(frame, limit: int = 20) -> list[dict]:
    if frame is None or getattr(frame, "empty", True):
        return []
    records = []
    for _, row in frame.head(limit).iterrows():
        record = {}
        for key, value in row.items():
            if value is None:
                record[key] = None
            elif isinstance(value, bool):
                record[key] = value
            else:
                try:
                    number = float(value)
                except (TypeError, ValueError):
                    record[key] = value
                else:
                    if math.isnan(number) or math.isinf(number):
                        record[key] = None
                    elif number.is_integer():
                        record[key] = int(number)
                    else:
                        record[key] = round(number, 4)
        records.append(record)
    return records


def hot_sectors_snapshot_locked() -> dict:
    return {
        "modules": {
            key: {"records": list(item["records"]), "error": item["error"]}
            for key, item in HOT_SECTORS_CACHE["modules"].items()
        },
        "updated_at": HOT_SECTORS_CACHE["updated_at"],
    }


def hot_sectors_data(force: bool = False) -> dict:
    now = time.time()
    with HOT_SECTORS_CACHE_LOCK:
        if (not force and HOT_SECTORS_CACHE["updated_at"] is not None
                and now - HOT_SECTORS_CACHE["updated_at"] < HOT_SECTORS_CACHE_TTL_SECONDS):
            return hot_sectors_snapshot_locked()
    providers = (
        ("industry", fetch_hot_industry, "行业板块"),
        ("concept", fetch_hot_concept, "概念板块"),
        ("hot_rank", fetch_hot_rank, "人气榜"),
    )
    modules: dict[str, dict] = {}
    for key, provider, label in providers:
        try:
            modules[key] = {"records": df_records(provider(), 20), "error": None}
        except Exception as error:
            with HOT_SECTORS_CACHE_LOCK:
                previous = HOT_SECTORS_CACHE["modules"].get(key)
            modules[key] = {
                "records": previous["records"] if previous else [],
                "error": f"{label}：{error}",
            }
    with HOT_SECTORS_CACHE_LOCK:
        HOT_SECTORS_CACHE["modules"] = modules
        HOT_SECTORS_CACHE["updated_at"] = now
        return hot_sectors_snapshot_locked()


def save_daily_prices(db: sqlite3.Connection, stock_code: str, frame, source: str, start_date: str, end_date: str) -> int:

    inserted = 0
    fetched_at = datetime.now().isoformat(timespec="seconds")
    seen_dates: set[str] = set()
    for _, row in frame.iterrows():
        raw_date = row.get("日期", row.get("date", ""))
        trade_date = str(raw_date).strip()[:10]
        if not trade_date or trade_date in seen_dates or trade_date < start_date or trade_date > end_date:
            continue
        def num(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return None
        seen_dates.add(trade_date)
        db.execute(
            """INSERT INTO daily_prices
            (stock_code, trade_date, open, high, low, close, volume, amount, source, fetched_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(stock_code, trade_date) DO UPDATE SET
            open = excluded.open, high = excluded.high, low = excluded.low,
            close = excluded.close, volume = excluded.volume, amount = excluded.amount,
            source = excluded.source, fetched_at = excluded.fetched_at""",
            (stock_code, trade_date, num(row.get("开盘", row.get("open"))),
             num(row.get("最高", row.get("high"))), num(row.get("最低", row.get("low"))),
             num(row.get("收盘", row.get("close"))), num(row.get("成交量", row.get("volume"))),
             num(row.get("成交额", row.get("amount"))),
             source, fetched_at),
        )
        inserted += 1
    return inserted


def sync_daily_prices(db: sqlite3.Connection, stock_code: str, start_date: str, end_date: str) -> int:
    frame, source = fetch_daily_prices(stock_code, start_date, end_date)
    if frame is None:
        return 0
    return save_daily_prices(db, stock_code, frame, source, start_date, end_date)


def sync_realtime_codes(stock_codes: list[str]) -> tuple[int, list[str]]:
    stock_codes = list(dict.fromkeys(stock_codes))
    if not REALTIME_SYNC_LOCK.acquire(blocking=False):
        return 0, ["实时行情同步正在进行"]
    try:
        try:
            quotes = fetch_realtime_prices(stock_codes)
        except Exception as error:
            return 0, [f"实时行情接口：{error}"]
        missing_codes = [stock_code for stock_code in stock_codes if stock_code not in quotes]
        with db_connect() as db:
            synced = save_realtime_prices(db, quotes)
            save_intraday_quotes(db, quotes)
            evaluate_realtime_alerts(db, quotes)
            evaluate_intraday_rebound_alerts(db, quotes)
        return synced, [f"{stock_code}：未返回实时行情" for stock_code in missing_codes]
    finally:
        REALTIME_SYNC_LOCK.release()


def sync_history_codes(stock_codes: list[str], full: bool = False) -> tuple[int, list[str]]:
    stock_codes = list(dict.fromkeys(stock_codes))
    today = date.today()
    with db_connect() as db:
        requests = [
            (
                stock_code,
                (today - timedelta(days=PRICE_SYNC_INITIAL_DAYS)).isoformat()
                if full else price_sync_start_date(db, stock_code, today),
                today.isoformat(),
            )
            for stock_code in stock_codes
        ]

    fetched, failed = [], []
    with ThreadPoolExecutor(max_workers=min(PRICE_SYNC_WORKERS, len(requests))) as executor:
        futures = {
            executor.submit(fetch_daily_prices, stock_code, start_date, end_date): (stock_code, start_date, end_date)
            for stock_code, start_date, end_date in requests
        }
        for future in as_completed(futures):
            stock_code, start_date, end_date = futures[future]
            try:
                frame, source = future.result()
                if frame is not None:
                    fetched.append((stock_code, frame, source, start_date, end_date))
            except Exception as error:
                failed.append(f"{stock_code}：{error}")

    synced = 0
    with db_connect() as db:
        for stock_code, frame, source, start_date, end_date in fetched:
            synced += save_daily_prices(db, stock_code, frame, source, start_date, end_date)
        refresh_alert_signal_outcomes(db, [item[0] for item in fetched])
    return synced, failed


def is_auto_sync_time(moment: datetime) -> bool:
    if moment.weekday() >= 5:
        return False
    minute = moment.hour * 60 + moment.minute
    return 9 * 60 + 15 <= minute <= 11 * 60 + 30 or 13 * 60 <= minute <= 15 * 60


def auto_sync_status() -> dict:
    with AUTO_SYNC_STATE_LOCK:
        return dict(AUTO_SYNC_STATE)


def run_auto_realtime_sync(moment: datetime | None = None) -> bool:
    moment = moment or datetime.now()
    if not is_auto_sync_time(moment):
        return False
    with db_connect() as db:
        stock_codes = [
            row["stock_code"]
            for row in db.execute(
                """SELECT stock_code FROM watchlist_stocks
                UNION
                SELECT stock_code FROM positions
                ORDER BY stock_code"""
            ).fetchall()
        ]
    if not stock_codes:
        return False
    indexes, index_error = sync_market_indexes()
    with AUTO_SYNC_STATE_LOCK:
        AUTO_SYNC_STATE["indexes_last_success_at"] = datetime.now().isoformat(timespec="seconds") if indexes else AUTO_SYNC_STATE["indexes_last_success_at"]
        AUTO_SYNC_STATE["indexes_last_error"] = index_error

    attempted_at = moment.isoformat(timespec="seconds")
    with AUTO_SYNC_STATE_LOCK:
        AUTO_SYNC_STATE.update({"running": True, "last_attempt_at": attempted_at, "last_error": None})
    try:
        synced, failed = sync_realtime_codes(stock_codes)
    except Exception as error:
        synced, failed = 0, [str(error)]
    with AUTO_SYNC_STATE_LOCK:
        AUTO_SYNC_STATE["running"] = False
        AUTO_SYNC_STATE["last_synced"] = synced
        if failed:
            AUTO_SYNC_STATE["last_error"] = "；".join(failed)
        else:
            AUTO_SYNC_STATE["last_success_at"] = attempted_at
            AUTO_SYNC_STATE["last_error"] = None
    return not failed


def auto_sync_loop() -> None:
    while not AUTO_SYNC_STOP.is_set():
        wait_seconds = AUTO_REALTIME_SYNC_INTERVAL_SECONDS - time.time() % AUTO_REALTIME_SYNC_INTERVAL_SECONDS
        if AUTO_SYNC_STOP.wait(wait_seconds):
            break
        run_auto_realtime_sync(datetime.now())


def start_auto_sync_scheduler() -> threading.Thread:
    global AUTO_SYNC_THREAD
    with AUTO_SYNC_START_LOCK:
        if AUTO_SYNC_THREAD is not None and AUTO_SYNC_THREAD.is_alive():
            return AUTO_SYNC_THREAD
        AUTO_SYNC_STOP.clear()
        AUTO_SYNC_THREAD = threading.Thread(target=auto_sync_loop, name="watchlist-realtime-sync", daemon=True)
        AUTO_SYNC_THREAD.start()
        return AUTO_SYNC_THREAD


def calculate_excursion_metrics(db: sqlite3.Connection, user_id: int, episode) -> dict | None:
    buy_price = float(episode["buy_price"] or 0)
    quantity = float(episode["quantity"] or 0)
    buy_date = episode["buy_date"]
    end_date = episode["sell_date"] if episode["status"] == "CLOSED" and episode["sell_date"] else date.today().isoformat()
    if not buy_price or quantity <= 0:
        return None
    prices = db.execute(
        """SELECT MAX(high) highest, MIN(low) lowest FROM daily_prices
        WHERE stock_code = ? AND trade_date BETWEEN ? AND ? AND high IS NOT NULL AND low IS NOT NULL""",
        (episode["stock_code"], buy_date, end_date),
    ).fetchone()
    if prices["highest"] is None or prices["lowest"] is None:
        return None
    highest = float(prices["highest"])
    lowest = float(prices["lowest"])
    mfe = (highest - buy_price) / buy_price
    mae = (lowest - buy_price) / buy_price
    max_gain_amount = (highest - buy_price) * quantity
    profit = float(episode["profit"] or 0)
    if max_gain_amount <= 0:
        capture_rate = None
    elif profit <= 0:
        capture_rate = 0.0
    else:
        capture_rate = max(0.0, min(1.0, profit / max_gain_amount))
    calculated_at = datetime.now().isoformat(timespec="seconds")
    db.execute(
        """INSERT INTO trade_excursion_metrics
        (trade_episode_id, user_id, mfe, mae, highest_price, lowest_price, max_gain_amount,
         capture_rate, buy_price, sell_price, quantity, data_start_date, data_end_date,
         price_source, calculated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'akshare', ?)
        ON CONFLICT(trade_episode_id) DO UPDATE SET
        mfe = excluded.mfe, mae = excluded.mae, highest_price = excluded.highest_price,
        lowest_price = excluded.lowest_price, max_gain_amount = excluded.max_gain_amount,
        capture_rate = excluded.capture_rate, buy_price = excluded.buy_price,
        sell_price = excluded.sell_price, quantity = excluded.quantity,
        data_start_date = excluded.data_start_date, data_end_date = excluded.data_end_date,
        price_source = excluded.price_source, calculated_at = excluded.calculated_at""",
        (episode["id"], user_id, mfe, mae, highest, lowest, max_gain_amount, capture_rate,
         buy_price, float(episode["sell_price"] or 0) if episode["status"] == "CLOSED" else None,
         quantity, buy_date, end_date, calculated_at),
    )
    return {
        "mfe": mfe, "mae": mae, "highest_price": highest, "lowest_price": lowest,
        "max_gain_amount": max_gain_amount, "capture_rate": capture_rate,
        "buy_price": buy_price, "sell_price": float(episode["sell_price"] or 0),
        "quantity": quantity, "data_start_date": buy_date, "data_end_date": end_date,
    }


def excursion_metrics_for(db: sqlite3.Connection, user_id: int, episode_id: int):
    return db.execute(
        "SELECT * FROM trade_excursion_metrics WHERE trade_episode_id = ? AND user_id = ?",
        (episode_id, user_id),
    ).fetchone()


def trade_diagnosis(db: sqlite3.Connection, review, metrics, episode) -> list[dict]:
    diagnoses: list[dict] = []
    profit = float(episode["profit"] or 0)
    closed = episode["status"] == "CLOSED"
    sell_price = float(episode["sell_price"] or 0) if closed and episode["sell_price"] else None
    buy_price = float(episode["buy_price"] or 0)

    if metrics is None:
        diagnoses.append({
            "code": "NO_PRICE", "severity": "info", "title": "行情数据不足",
            "message": "暂无持仓期历史行情，无法评估 MFE/MAE 与卖出时机。",
        })
        return diagnoses

    capture = metrics["capture_rate"]
    highest = float(metrics["highest_price"] or 0)
    lowest = float(metrics["lowest_price"] or 0)

    if profit > 0 and capture is not None and capture < 0.5:
        diagnoses.append({
            "code": "EARLY_EXIT", "severity": "warning", "title": "可能卖出较早",
            "message": f"方向判断可能正确，但只捕获了约 {capture * 100:.0f}% 的可实现收益。",
        })

    if closed and sell_price and highest and sell_price >= highest * 0.95:
        diagnoses.append({
            "code": "GOOD_EXIT", "severity": "good", "title": "卖出执行较好",
            "message": "卖出位置接近持仓期阶段高点。",
        })

    stop_loss = float(review["stop_loss_price"] or 0) if review and review["stop_loss_price"] else None
    if stop_loss and sell_price and sell_price < stop_loss and profit < 0:
        diagnoses.append({
            "code": "STOP_VIOLATION", "severity": "bad", "title": "止损执行偏差",
            "message": f"实际卖出价 {sell_price:.2f} 低于计划止损价 {stop_loss:.2f}。",
        })

    target = float(review["expected_target_price"] or 0) if review and review["expected_target_price"] else None
    if target and sell_price and sell_price >= target:
        diagnoses.append({
            "code": "TARGET_MET", "severity": "good", "title": "达到目标价",
            "message": f"实际卖出价达到计划目标价 {target:.2f}。",
        })

    if closed and profit < 0 and sell_price and episode["sell_date"]:
        exit_high = db.execute(
            """SELECT MAX(high) h FROM daily_prices
            WHERE stock_code = ? AND trade_date > ? AND trade_date <= date(?, '+15 days') AND high IS NOT NULL""",
            (episode["stock_code"], episode["sell_date"], episode["sell_date"]),
        ).fetchone()
        if exit_high["h"] is not None and float(exit_high["h"]) >= sell_price * 1.05:
            diagnoses.append({
                "code": "MISSED_REBOUND", "severity": "warning", "title": "卖出后快速上涨",
                "message": "最终亏损，但卖出后 15 日内股价明显上涨，方向可能判断正确，买入时机或执行存在问题。",
            })

    if review is None or (not review["stop_loss_price"] and not review["expected_target_price"]):
        diagnoses.append({
            "code": "NO_PLAN", "severity": "info", "title": "计划数据不足",
            "message": "未记录计划止损或目标价，无法评估计划执行情况。",
        })

    if not diagnoses:
        diagnoses.append({
            "code": "BALANCED", "severity": "neutral", "title": "无明显偏差",
            "message": "从已有数据看，未发现明显可优化的执行问题。",
        })

    return diagnoses


def calculate_discipline_data(db: sqlite3.Connection, user_id: int | None = None) -> dict:
    """计算情绪纪律相关指标"""
    user_id = user_id or current_user_id()
    
    # 1. 连续盈亏分析（核心指标）
    streak_analysis = db.execute("""
        WITH consecutive AS (
            SELECT 
                sell_date, 
                profit,
                stock_code,
                stock_name,
                CASE WHEN profit > 0 THEN 1 ELSE 0 END as is_win,
                LAG(profit, 1) OVER (ORDER BY sell_date, id) as prev_profit,
                LAG(sell_date, 1) OVER (ORDER BY sell_date, id) as prev_date
            FROM fifo_matches WHERE user_id = ?
        )
        SELECT 
            CASE 
                WHEN prev_profit > 0 THEN 'after_win'
                WHEN prev_profit <= 0 THEN 'after_loss'
            END as prev_state,
            COUNT(*) as trades,
            SUM(CASE WHEN is_win = 1 THEN 1 ELSE 0 END) as wins,
            AVG(CASE WHEN is_win = 1 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(profit) as avg_profit,
            SUM(profit) as total_profit
        FROM consecutive
        WHERE prev_profit IS NOT NULL
        GROUP BY prev_state
    """, (user_id,)).fetchall()
    
    # 2. 冲动交易识别（同一股票短时间内反复操作）
    impulse_trades = db.execute("""
        SELECT 
            f1.stock_code,
            f1.stock_name,
            COUNT(*) as trade_count,
            SUM(f1.profit + f2.profit) as total_loss,
            AVG(f1.holding_days) as avg_holding
        FROM fifo_matches f1
        JOIN fifo_matches f2 ON f1.user_id = f2.user_id AND f1.stock_code = f2.stock_code
        WHERE f1.user_id = ? AND f1.sell_date = f2.buy_date
          AND f1.profit < 0
          AND f2.profit < 0
        GROUP BY f1.stock_code, f1.stock_name
        HAVING COUNT(*) >= 2
        ORDER BY total_loss ASC
        LIMIT 10
    """, (user_id,)).fetchall()
    
    # 3. 持仓时长与盈亏关系（T+0, T+1, T+2...）
    holding_analysis = db.execute("""
        SELECT 
            holding_days,
            COUNT(*) as count,
            SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END) as wins,
            AVG(profit) as avg_profit,
            SUM(profit) as total_profit,
            AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END) as win_rate
        FROM fifo_matches WHERE user_id = ?
        GROUP BY holding_days
        HAVING COUNT(*) >= 5
        ORDER BY holding_days
    """, (user_id,)).fetchall()
    
    # 4. 连胜连败趋势数据（用于图表）
    streak_trend = db.execute("""
        WITH numbered AS (
            SELECT 
                sell_date,
                profit,
                SUM(profit) OVER (ORDER BY sell_date, id) as cumulative_profit,
                CASE WHEN profit > 0 THEN 1 ELSE 0 END as is_win
            FROM fifo_matches WHERE user_id = ?
        )
        SELECT 
            sell_date,
            cumulative_profit,
            profit,
            is_win
        FROM numbered
        ORDER BY sell_date
    """, (user_id,)).fetchall()
    
    # 5. 最危险的操作模式（亏损后立即同一只票再买）
    revenge_trades = db.execute("""
        WITH ranked AS (
            SELECT 
                stock_code,
                stock_name,
                sell_date,
                profit,
                LEAD(buy_date, 1) OVER (PARTITION BY stock_code ORDER BY sell_date) as next_buy_date
            FROM fifo_matches WHERE user_id = ?
        )
        SELECT 
            stock_code,
            stock_name,
            COUNT(*) as revenge_count,
            AVG(profit) as avg_loss
        FROM ranked
        WHERE profit < 0 
          AND next_buy_date IS NOT NULL
          AND julianday(next_buy_date) - julianday(sell_date) <= 1
        GROUP BY stock_code, stock_name
        ORDER BY revenge_count DESC
        LIMIT 10
    """, (user_id,)).fetchall()
    
    return {
        "streak": streak_analysis,
        "impulse": impulse_trades,
        "holding": holding_analysis,
        "trend": streak_trend,
        "revenge": revenge_trades
    }


def calculate_recent_data(db: sqlite3.Connection, days: int = 7, user_id: int | None = None) -> dict:
    """计算最近 N 天的表现数据"""
    user_id = user_id or current_user_id()

    # 近期汇总
    recent_summary = db.execute(
        """SELECT COUNT(*) as trades,
        COALESCE(SUM(profit), 0) as total_profit,
        COALESCE(AVG(profit), 0) as avg_profit,
        COALESCE(AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END), 0) as win_rate,
        COALESCE(AVG(holding_days), 0) as avg_hold,
        COALESCE(SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END), 0) as wins
        FROM fifo_matches
        WHERE user_id = ? AND sell_date >= date('now', ? )""",
        (user_id, f"-{days} days"),
    ).fetchone()

    # 历史全局汇总（用于对比）
    all_summary = db.execute(
        """SELECT COUNT(*) as trades,
        COALESCE(AVG(profit), 0) as avg_profit,
        COALESCE(AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END), 0) as win_rate,
        COALESCE(AVG(holding_days), 0) as avg_hold
        FROM fifo_matches WHERE user_id = ?""", (user_id,)
    ).fetchone()

    # 每日盈亏明细
    daily = db.execute(
        """SELECT sell_date,
        COUNT(*) as trades,
        COALESCE(SUM(profit), 0) as profit,
        COALESCE(SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END), 0) as wins
        FROM fifo_matches
        WHERE user_id = ? AND sell_date >= date('now', ? )
        GROUP BY sell_date ORDER BY sell_date""",
        (user_id, f"-{days} days"),
    ).fetchall()

    # 本期个股排行（全部，前端按盈亏排序后取前5/后5）
    stocks = db.execute(
        """SELECT stock_code, MAX(stock_name) as stock_name,
        COUNT(*) as trades,
        COALESCE(SUM(profit), 0) as profit,
        COALESCE(AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END), 0) as win_rate,
        COALESCE(AVG(holding_days), 0) as avg_hold
        FROM fifo_matches
        WHERE user_id = ? AND sell_date >= date('now', ? )
        GROUP BY stock_code
        ORDER BY profit DESC""",
        (user_id, f"-{days} days"),
    ).fetchall()

    # 最差单日 + 当日明细
    worst_day = db.execute(
        """SELECT sell_date, SUM(profit) as loss, COUNT(*) as trades,
        SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END) as wins
        FROM fifo_matches
        WHERE user_id = ? AND sell_date >= date('now', ? )
        GROUP BY sell_date ORDER BY loss ASC LIMIT 1""",
        (user_id, f"-{days} days"),
    ).fetchone()

    worst_day_trades = []
    if worst_day:
        worst_day_trades = db.execute(
            """SELECT f.stock_code, MAX(f.stock_name) as stock_name,
            COUNT(*) as matches, SUM(f.profit) as profit,
            MIN(f.buy_date) as buy_date, f.sell_date
            FROM fifo_matches f
            WHERE f.user_id = ? AND f.sell_date = ?
            GROUP BY f.stock_code, f.sell_date
            ORDER BY profit ASC""",
            (user_id, worst_day["sell_date"]),
        ).fetchall()

    return {
        "summary": recent_summary,
        "all_summary": all_summary,
        "daily": daily,
        "stocks": stocks,
        "best_stocks": list(stocks[:5]),
        "worst_stocks": list(reversed(stocks[-5:])) if len(stocks) >= 5 else list(reversed(stocks)),
        "worst_day": worst_day,
        "worst_day_trades": worst_day_trades,
    }


def calculate_timing_data(db: sqlite3.Connection, user_id: int | None = None) -> dict:
    """计算时机把握分析数据：买入时段、卖出时段、星期效应"""
    user_id = user_id or current_user_id()

    # 1. 买入小时胜率
    buy_by_hour = db.execute("""
        SELECT SUBSTR(e.trade_time, 1, 2) as hour,
            COUNT(*) as trades,
            SUM(CASE WHEN f.profit > 0 THEN 1 ELSE 0 END) as wins,
            AVG(CASE WHEN f.profit > 0 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(f.profit) as avg_profit,
            SUM(f.profit) as total_profit
        FROM fifo_matches f
        JOIN executions e ON f.user_id = e.user_id AND f.buy_execution_id = e.id
        WHERE f.user_id = ?
        GROUP BY hour ORDER BY hour
    """, (user_id,)).fetchall()

    # 2. 卖出小时胜率
    sell_by_hour = db.execute("""
        SELECT SUBSTR(e.trade_time, 1, 2) as hour,
            COUNT(*) as trades,
            SUM(CASE WHEN f.profit > 0 THEN 1 ELSE 0 END) as wins,
            AVG(CASE WHEN f.profit > 0 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(f.profit) as avg_profit,
            SUM(f.profit) as total_profit
        FROM fifo_matches f
        JOIN executions e ON f.user_id = e.user_id AND f.sell_execution_id = e.id
        WHERE f.user_id = ?
        GROUP BY hour ORDER BY hour
    """, (user_id,)).fetchall()

    # 3. 星期效应（按卖出日期的星期）
    weekday_perf = db.execute("""
        SELECT
            CAST(strftime('%w', sell_date) AS INTEGER) as dow,
            CASE CAST(strftime('%w', sell_date) AS INTEGER)
                WHEN 1 THEN '周一' WHEN 2 THEN '周二' WHEN 3 THEN '周三'
                WHEN 4 THEN '周四' WHEN 5 THEN '周五'
            END as label,
            COUNT(*) as trades,
            SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END) as wins,
            AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(profit) as avg_profit,
            SUM(profit) as total_profit
        FROM fifo_matches
        WHERE user_id = ? AND CAST(strftime('%w', sell_date) AS INTEGER) BETWEEN 1 AND 5
        GROUP BY dow ORDER BY dow
    """, (user_id,)).fetchall()

    # 4. 买入星期效应
    buy_weekday = db.execute("""
        SELECT
            CAST(strftime('%w', e.trade_date) AS INTEGER) as dow,
            CASE CAST(strftime('%w', e.trade_date) AS INTEGER)
                WHEN 1 THEN '周一' WHEN 2 THEN '周二' WHEN 3 THEN '周三'
                WHEN 4 THEN '周四' WHEN 5 THEN '周五'
            END as label,
            COUNT(*) as trades,
            AVG(CASE WHEN f.profit > 0 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(f.profit) as avg_profit
        FROM fifo_matches f
        JOIN executions e ON f.user_id = e.user_id AND f.buy_execution_id = e.id
        WHERE f.user_id = ? AND CAST(strftime('%w', e.trade_date) AS INTEGER) BETWEEN 1 AND 5
        GROUP BY dow ORDER BY dow
    """, (user_id,)).fetchall()

    # 5. 开盘 vs 尾盘买入对比（9:25-9:35 为开盘，14:30-15:00 为尾盘）
    session_buy = db.execute("""
        SELECT
            CASE
                WHEN SUBSTR(e.trade_time,1,5) BETWEEN '09:25' AND '09:35' THEN '开盘'
                WHEN SUBSTR(e.trade_time,1,5) BETWEEN '09:36' AND '11:30' THEN '上午盘'
                WHEN SUBSTR(e.trade_time,1,5) BETWEEN '13:00' AND '14:29' THEN '下午盘'
                WHEN SUBSTR(e.trade_time,1,5) >= '14:30' THEN '尾盘'
                ELSE '其他'
            END as session,
            COUNT(*) as trades,
            SUM(CASE WHEN f.profit > 0 THEN 1 ELSE 0 END) as wins,
            AVG(CASE WHEN f.profit > 0 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(f.profit) as avg_profit,
            SUM(f.profit) as total_profit
        FROM fifo_matches f
        JOIN executions e ON f.user_id = e.user_id AND f.buy_execution_id = e.id
        WHERE f.user_id = ?
        GROUP BY session
        ORDER BY CASE session
            WHEN '开盘' THEN 1 WHEN '上午盘' THEN 2
            WHEN '下午盘' THEN 3 WHEN '尾盘' THEN 4 ELSE 5
        END
    """, (user_id,)).fetchall()

    # 6. 开盘 vs 尾盘卖出对比
    session_sell = db.execute("""
        SELECT
            CASE
                WHEN SUBSTR(e.trade_time,1,5) BETWEEN '09:25' AND '09:35' THEN '开盘'
                WHEN SUBSTR(e.trade_time,1,5) BETWEEN '09:36' AND '11:30' THEN '上午盘'
                WHEN SUBSTR(e.trade_time,1,5) BETWEEN '13:00' AND '14:29' THEN '下午盘'
                WHEN SUBSTR(e.trade_time,1,5) >= '14:30' THEN '尾盘'
                ELSE '其他'
            END as session,
            COUNT(*) as trades,
            SUM(CASE WHEN f.profit > 0 THEN 1 ELSE 0 END) as wins,
            AVG(CASE WHEN f.profit > 0 THEN 1.0 ELSE 0 END) as win_rate,
            AVG(f.profit) as avg_profit,
            SUM(f.profit) as total_profit
        FROM fifo_matches f
        JOIN executions e ON f.user_id = e.user_id AND f.sell_execution_id = e.id
        WHERE f.user_id = ?
        GROUP BY session
        ORDER BY CASE session
            WHEN '开盘' THEN 1 WHEN '上午盘' THEN 2
            WHEN '下午盘' THEN 3 WHEN '尾盘' THEN 4 ELSE 5
        END
    """, (user_id,)).fetchall()

    return {
        "buy_by_hour": buy_by_hour,
        "sell_by_hour": sell_by_hour,
        "weekday": weekday_perf,
        "buy_weekday": buy_weekday,
        "session_buy": session_buy,
        "session_sell": session_sell,
    }


STRATEGY_SESSIONS = {
    "open": "开盘 09:25-09:35",
    "morning": "上午盘 09:36-11:30",
    "afternoon": "下午盘 13:00-14:29",
    "close": "尾盘 14:30-15:00",
}
STRATEGY_WEEKDAYS = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五"}


def strategy_filters(args) -> dict:
    """Validate strategy query parameters and return normalized filters."""
    buy_session = args.get("buy_session", "all")
    sell_session = args.get("sell_session", "all")
    if buy_session not in {"all", *STRATEGY_SESSIONS}:
        buy_session = "all"
    if sell_session not in {"all", *STRATEGY_SESSIONS}:
        sell_session = "all"

    def bounded_int(name: str, default: int, minimum: int, maximum: int) -> int:
        try:
            return min(max(int(args.get(name, default)), minimum), maximum)
        except (TypeError, ValueError):
            return default

    min_days = bounded_int("min_days", 0, 0, 3650)
    max_days = bounded_int("max_days", 30, 0, 3650)
    if min_days > max_days:
        min_days, max_days = max_days, min_days
    buy_weekday = bounded_int("buy_weekday", 0, 0, 5)
    sell_weekday = bounded_int("sell_weekday", 0, 0, 5)
    period = bounded_int("period", 0, 0, 3650)
    if period not in (0, 30, 60, 90, 180, 365):
        period = 0
    return {
        "buy_session": buy_session,
        "sell_session": sell_session,
        "min_days": min_days,
        "max_days": max_days,
        "buy_weekday": buy_weekday,
        "sell_weekday": sell_weekday,
        "period": period,
    }


def calculate_strategy_data(db: sqlite3.Connection, filters: dict, user_id: int | None = None) -> dict:
    """Backtest a timing strategy against historical FIFO matches."""
    user_id = user_id or current_user_id()
    session_ranges = {
        "open": ("09:25", "09:35"),
        "morning": ("09:36", "11:30"),
        "afternoon": ("13:00", "14:29"),
        "close": ("14:30", "15:00"),
    }
    where = ["f.user_id = ?", "f.holding_days BETWEEN ? AND ?"]
    params: list[object] = [user_id, filters["min_days"], filters["max_days"]]
    for key, alias in (("buy_session", "buy"), ("sell_session", "sell")):
        if filters[key] != "all":
            start, end = session_ranges[filters[key]]
            where.append(f"SUBSTR({alias}.trade_time, 1, 5) BETWEEN ? AND ?")
            params.extend((start, end))
    if filters["buy_weekday"]:
        where.append("CAST(strftime('%w', buy.trade_date) AS INTEGER) = ?")
        params.append(filters["buy_weekday"])
    if filters["sell_weekday"]:
        where.append("CAST(strftime('%w', sell.trade_date) AS INTEGER) = ?")
        params.append(filters["sell_weekday"])
    if filters["period"]:
        where.append("f.sell_date >= date('now', ?)")
        params.append(f"-{filters['period']} days")

    from_sql = """FROM fifo_matches f
        JOIN executions buy ON f.user_id = buy.user_id AND f.buy_execution_id = buy.id
        JOIN executions sell ON f.user_id = sell.user_id AND f.sell_execution_id = sell.id"""
    summary = db.execute(
        f"""SELECT COUNT(*) trades, COALESCE(SUM(f.profit), 0) profit,
        COALESCE(AVG(f.profit), 0) avg_profit,
        COALESCE(AVG(CASE WHEN f.profit > 0 THEN 1.0 ELSE 0 END), 0) win_rate,
        COALESCE(AVG(CASE WHEN f.profit > 0 THEN f.profit END), 0) avg_win,
        COALESCE(AVG(CASE WHEN f.profit <= 0 THEN f.profit END), 0) avg_loss,
        COALESCE(SUM(CASE WHEN f.profit > 0 THEN f.profit ELSE 0 END), 0) gross_profit,
        COALESCE(SUM(CASE WHEN f.profit <= 0 THEN -f.profit ELSE 0 END), 0) gross_loss,
        COALESCE(AVG(f.holding_days), 0) avg_holding
        {from_sql} WHERE {' AND '.join(where)}""",
        params,
    ).fetchone()
    baseline = db.execute(
        """SELECT COUNT(*) trades, COALESCE(SUM(profit), 0) profit,
        COALESCE(AVG(profit), 0) avg_profit,
        COALESCE(AVG(CASE WHEN profit > 0 THEN 1.0 ELSE 0 END), 0) win_rate
        FROM fifo_matches WHERE user_id = ?""", (user_id,)
    ).fetchone()
    trades = db.execute(
        f"""SELECT f.stock_code, f.stock_name, f.buy_date, f.sell_date,
        buy.trade_time buy_time, sell.trade_time sell_time, f.holding_days,
        f.quantity, f.profit, f.profit_rate
        {from_sql} WHERE {' AND '.join(where)}
        ORDER BY f.sell_date DESC, sell.trade_time DESC, f.id DESC LIMIT 100""",
        params,
    ).fetchall()

    profit_factor = None
    if summary["gross_loss"] > 0:
        profit_factor = summary["gross_profit"] / summary["gross_loss"]
    elif summary["gross_profit"] > 0:
        profit_factor = float("inf")

    labels = []
    if filters["period"]:
        labels.append(f"近 {filters['period']} 天")
    if filters["buy_session"] != "all":
        labels.append(f"{STRATEGY_SESSIONS[filters['buy_session']]}买入")
    if filters["sell_session"] != "all":
        labels.append(f"{STRATEGY_SESSIONS[filters['sell_session']]}卖出")
    labels.append(f"持仓 T+{filters['min_days']} 至 T+{filters['max_days']}")
    if filters["buy_weekday"]:
        labels.append(f"{STRATEGY_WEEKDAYS[filters['buy_weekday']]}买入")
    if filters["sell_weekday"]:
        labels.append(f"{STRATEGY_WEEKDAYS[filters['sell_weekday']]}卖出")
    return {
        "summary": summary,
        "baseline": baseline,
        "trades": trades,
        "profit_factor": profit_factor,
        "labels": labels,
        "sample_warning": summary["trades"] < 30,
    }


BUY_COLOR = "#b95d4c"
SELL_COLOR = "#22654d"
GRID_COLOR = "#e4e0d6"


def place_label(labels, x, y, text, color):
    text_width = max(54.0, len(text) * 6.4)
    candidates = [y - 10, y + 21]
    for k in range(1, 14):
        candidates.append(y - (10 + k * 24))
        candidates.append(y + (21 + k * 24))
    for candidate_y in candidates:
        if all(abs(px - x) > text_width or abs(py - candidate_y) > 15 for px, py, _, _ in labels):
            labels.append((x, candidate_y, text, color))
            return
    labels.append((x, y - 10, text, color))


def build_chart(points, matches) -> str | None:
    n = len(points)
    if n == 0:
        return None
    prices = [float(point["deal_price"]) for point in points]
    low, high = min(prices), max(prices)
    if high <= low:
        pad = max(high * 0.05, 0.5)
        low, high = low - pad, high + pad
    else:
        pad = (high - low) * 0.15
        low, high = low - pad, high + pad

    width, height = 640, 250
    pad_left, pad_right, pad_top, pad_bottom = 52, 18, 18, 30
    plot_w = width - pad_left - pad_right
    plot_h = height - pad_top - pad_bottom

    def xy(index, price):
        span = max(1, n - 1)
        x = pad_left + (index / span) * plot_w
        y = pad_top + (1 - (price - low) / (high - low)) * plot_h
        return round(x, 1), round(y, 1)

    grid_lines = []
    tick_count = 4
    for k in range(tick_count + 1):
        value = low + (high - low) * k / tick_count
        y = pad_top + (1 - k / tick_count) * plot_h
        grid_lines.append(f'<line x1="{pad_left}" y1="{y:.1f}" x2="{width - pad_right}" y2="{y:.1f}" stroke="{GRID_COLOR}" stroke-width="1"/><text x="{pad_left - 8}" y="{y + 4:.1f}" text-anchor="end" font-size="11" fill="#718078">{value:.2f}</text>')

    sell_profit_by_execution: dict[int, float] = {}
    for match in matches:
        sell_profit_by_execution[match["sell_execution_id"]] = sell_profit_by_execution.get(match["sell_execution_id"], 0.0) + float(match["profit"])

    x_ticks: list[int] = []
    last_date = None
    for index, point in enumerate(points):
        if point["trade_date"] != last_date:
            x_ticks.append(index)
            last_date = point["trade_date"]
    if len(x_ticks) > 6:
        step = (len(x_ticks) - 1) / 5
        x_ticks = [x_ticks[round(step * k)] for k in range(6)]
    if len(x_ticks) == 1 and n > 1:
        x_ticks = [0, n - 1]
    x_labels = "".join(
        f'<text x="{xy(index, prices[index])[0]}" y="{height - 10}" text-anchor="middle" font-size="11" fill="#718078">{points[index]["trade_date"]}</text>'
        for index in x_ticks
    )

    polyline = " ".join(f"{xy(index, prices[index])[0]},{xy(index, prices[index])[1]}" for index in range(n))
    line = f'<polyline points="{polyline}" fill="none" stroke="#8ba79a" stroke-width="2"/>'

    def place_label(labels, x, y, text, color):
        text_width = max(54.0, len(text) * 6.4)
        for candidate_y in (y - 10, y + 21, y - 25, y + 34, y - 40):
            if all(abs(px - x) > text_width or abs(py - candidate_y) > 15 for px, py, _, _ in labels):
                labels.append((x, candidate_y, text, color))
                return
        labels.append((x, y - 10, text, color))

    points_svg = []
    profit_labels = []
    for index, point in enumerate(points):
        x, y = xy(index, float(point["deal_price"]))
        action = point["action"]
        color = BUY_COLOR if action == "BUY" else SELL_COLOR
        tooltip = f"{point['trade_date']} {point['trade_time']} {'买入' if action == 'BUY' else '卖出'} {point['quantity']:g} 股 @ {point['deal_price']:.2f}"
        points_svg.append(f'<circle cx="{x}" cy="{y}" r="5" fill="{color}" stroke="#fff" stroke-width="1.5"><title>{tooltip}</title></circle>')
        if action == "SELL" and point["id"] in sell_profit_by_execution:
            profit = sell_profit_by_execution[point["id"]]
            label_color = BUY_COLOR if profit >= 0 else SELL_COLOR
            sign = "+" if profit >= 0 else ""
            place_label(profit_labels, x, y, f"{sign}{profit:,.2f}", label_color)

    profit_labels_svg = "".join(
        f'<text x="{x}" y="{y}" text-anchor="middle" font-size="11" font-weight="bold" fill="{color}">{text}</text>'
        for x, y, text, color in profit_labels
    )

    svg = (
        f'<svg viewBox="0 0 {width} {height}" role="img" aria-label="成交价格走势">'
        f'<rect width="{width}" height="{height}" fill="#fffdf8"/>'
        + "".join(grid_lines)
        + x_labels
        + line
        + "".join(points_svg)
        + profit_labels_svg
        + "</svg>"
    )
    return svg


@app.template_filter("money")
def money(value: object) -> str:
    return f"{float(value or 0):,.2f}"


@app.template_filter("percent")
def percent(value: object) -> str:
    return f"{float(value or 0) * 100:.1f}%"


def user_name_from_form(field: str = "name") -> str | None:
    name = request.form.get(field, "").strip()
    if not name:
        flash("用户名称不能为空", "error")
        return None
    if len(name) > MAX_USER_NAME_LENGTH:
        flash(f"用户名称不能超过 {MAX_USER_NAME_LENGTH} 个字符", "error")
        return None
    return name


def user_redirect():
    target = request.form.get("next", "")
    if not target.startswith("/") or target.startswith("//"):
        target = url_for("analysis_stocks")
    return redirect(target)


@app.post("/users/switch")
def switch_user():
    try:
        user_id = int(request.form.get("user_id", 0))
    except (TypeError, ValueError):
        user_id = 0
    with db_connect() as db:
        user = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if user is None:
        flash("用户不存在，已保留当前用户", "error")
    else:
        session["user_id"] = user["id"]
    return user_redirect()


@app.post("/users/create")
def create_user():
    name = user_name_from_form()
    if name is not None:
        try:
            with db_connect() as db:
                cursor = db.execute(
                    "INSERT INTO users (name, created_at) VALUES (?, ?)",
                    (name, datetime.now().isoformat(timespec="seconds")),
                )
                session["user_id"] = cursor.lastrowid
            flash(f"已创建并切换到用户：{name}", "success")
        except sqlite3.IntegrityError:
            flash("用户名称已存在", "error")
    return user_redirect()


@app.post("/users/<int:id>/rename")
def rename_user(id: int):
    user_id = id
    name = user_name_from_form()
    if name is not None:
        try:
            with db_connect() as db:
                cursor = db.execute("UPDATE users SET name = ? WHERE id = ?", (name, user_id))
                if cursor.rowcount == 0:
                    flash("用户不存在", "error")
                else:
                    flash(f"用户已重命名为：{name}", "success")
        except sqlite3.IntegrityError:
            flash("用户名称已存在", "error")
    return user_redirect()


@app.post("/users/<int:id>/delete")
def delete_user(id: int):
    user_id = id
    confirmation = request.form.get("confirm_name", "").strip()
    with db_connect() as db:
        user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if user is None:
            flash("用户不存在", "error")
        elif db.execute("SELECT COUNT(*) FROM users").fetchone()[0] <= 1:
            flash("不能删除最后一个用户", "error")
        elif confirmation != user["name"]:
            flash("确认名称不匹配，未删除用户", "error")
        else:
            db.execute("DELETE FROM users WHERE id = ?", (user_id,))
            replacement = db.execute("SELECT id FROM users ORDER BY id LIMIT 1").fetchone()
            if session.get("user_id") == user_id:
                session["user_id"] = replacement["id"]
            flash(f"已删除用户：{user['name']}", "success")
    return user_redirect()


@app.get("/")
def dashboard():
    return redirect(url_for("analysis_portfolio"))


@app.get("/analysis/stocks")
def analysis_stocks():
    code = request.args.get("code", "").strip()
    stock_query = request.args.get("q", "").strip()
    stock_sort = request.args.get("sort", "recent")
    if stock_sort not in STOCK_SORTS:
        stock_sort = "recent"
    user_id = current_user_id()
    with db_connect() as db:
        data = analysis_data(db, user_id, stock_sort)
        stocks = data["by_stock"]
        position_rows = db.execute(
            """SELECT p.stock_code, MAX(p.stock_name) stock_name, MIN(p.buy_date) first_buy_date,
            SUM(p.quantity) quantity, SUM(p.total_cost) total_cost,
            SUM(p.total_cost) / NULLIF(SUM(p.quantity), 0) avg_cost,
            (SELECT ep.id FROM trade_episodes ep
                WHERE ep.user_id = p.user_id AND ep.stock_code = p.stock_code AND ep.status = 'OPEN'
                ORDER BY ep.opened_at DESC, ep.id DESC LIMIT 1) open_episode_id
            FROM positions p WHERE p.user_id = ? GROUP BY p.stock_code""",
            (user_id,),
        ).fetchall()
        positions_by_code = {row["stock_code"]: dict(row) for row in position_rows}
        if stock_query:
            query_lower = stock_query.lower()
            query_digits = "".join(character for character in stock_query if character.isdigit())
            stocks = [
                stock for stock in stocks
                if query_lower in stock["stock_name"].lower()
                or query_lower in stock["stock_code"].lower()
                or (query_digits and query_digits in stock["stock_code"])
            ]
        detail = None
        if code and not any(stock["stock_code"] == code for stock in stocks) and code not in positions_by_code:
            code = ""
        if not code and stocks:
            code = stocks[0]["stock_code"]
        if code:
            name = next((stock["stock_name"] for stock in stocks if stock["stock_code"] == code), "")
            if not name:
                name = db.execute("SELECT MAX(stock_name) stock_name FROM executions WHERE user_id = ? AND stock_code = ?", (user_id, code)).fetchone()["stock_name"] or ""
            current_position = positions_by_code.get(code)
            if current_position is not None:
                current_position["quantity"] = float(current_position["quantity"] or 0)
                current_position["total_cost"] = float(current_position["total_cost"] or 0)
                current_position["avg_cost"] = float(current_position["avg_cost"] or 0)
                current_position["holding_days"] = max(
                    0, (date.today() - datetime.strptime(current_position["first_buy_date"], "%Y-%m-%d").date()).days,
                )
                latest_price = db.execute(
                    """SELECT trade_date, close FROM daily_prices WHERE stock_code = ? AND close IS NOT NULL
                    ORDER BY trade_date DESC LIMIT 1""",
                    (code,),
                ).fetchone()
                current_position["latest_price"] = float(latest_price["close"]) if latest_price else None
                current_position["price_date"] = latest_price["trade_date"] if latest_price else None
                current_position["market_value"] = (
                    current_position["quantity"] * current_position["latest_price"]
                    if current_position["latest_price"] is not None else None
                )
                current_position["unrealized_profit"] = (
                    current_position["market_value"] - current_position["total_cost"]
                    if current_position["market_value"] is not None else None
                )
                current_position["unrealized_rate"] = (
                    current_position["unrealized_profit"] / current_position["total_cost"]
                    if current_position["unrealized_profit"] is not None and current_position["total_cost"] else None
                )
            episode_rows = [row for row in trade_episode_rows(db, user_id, code) if row["status"] == "CLOSED"]
            episodes = []
            for index, row in enumerate(
                sorted(episode_rows, key=lambda item: (item["closed_at"], item["id"]), reverse=True),
                start=1,
            ):
                item = dict(row)
                item["index"] = index
                item["buy_at_display"] = datetime.fromisoformat(item["opened_at"]).strftime("%m-%d %H:%M")
                item["sell_at_display"] = datetime.fromisoformat(item["closed_at"]).strftime("%m-%d %H:%M")
                exact_days = max(0.0, float(item["holding_days_exact"] or 0))
                item["holding_display"] = "<1天" if exact_days < 1 else f"{int(exact_days)}天"
                review = db.execute(
                    "SELECT * FROM trade_reviews WHERE trade_episode_id = ? AND user_id = ?",
                    (item["id"], user_id),
                ).fetchone()
                if review is not None:
                    reason_codes = [
                        rt["reason_type"]
                        for rt in db.execute(
                            "SELECT reason_type FROM trade_review_reason_types WHERE trade_review_id = ? ORDER BY reason_type",
                            (review["id"],),
                        )
                    ]
                    item["review_data"] = {
                        "stock_code": item["stock_code"],
                        "stock_name": item["stock_name"],
                        "completed_at": (review["completed_at"] or "")[:10],
                        "judgement_label": REVIEW_JUDGEMENT_LABELS.get(review["judgement_result"], "") or "未填写",
                        "main_problem_label": REVIEW_MAIN_PROBLEM_LABELS.get(review["main_problem"], "") or "未填写",
                        "confidence_level": review["confidence_level"],
                        "trade_reason": review["trade_reason"],
                        "sell_reason": review["sell_reason"],
                        "review_note": review["review_note"],
                        "next_action": review["next_action"],
                        "expected_profit_percent": review["expected_profit_percent"],
                        "expected_target_price": review["expected_target_price"],
                        "stop_loss_price": review["stop_loss_price"],
                        "expected_holding_days": review["expected_holding_days"],
                        "reason_labels": [
                            REVIEW_REASON_TYPE_LABELS[rt] for rt in reason_codes if rt in REVIEW_REASON_TYPE_LABELS
                        ],
                    }
                episodes.append(item)
            wins = [episode for episode in episodes if episode["profit"] > 0]
            losses = [episode for episode in episodes if episode["profit"] < 0]
            avg_win = sum(float(episode["profit"]) for episode in wins) / len(wins) if wins else 0.0
            avg_loss = sum(float(episode["profit"]) for episode in losses) / len(losses) if losses else None
            summary = {
                "trades": len(episodes),
                "win_rate": len(wins) / len(episodes) if episodes else 0.0,
                "profit": sum(float(episode["profit"]) for episode in episodes),
                "profit_factor": avg_win / abs(avg_loss) if avg_loss else None,
            }
            detail = {"code": code, "name": name, "episodes": episodes, "summary": summary, "position": current_position}
        return render_template("index.html", page="stocks", data=data, stocks=stocks, selected=code,
                               detail=detail, stock_sort=stock_sort, stock_query=stock_query)


@app.get("/analysis/trades/<int:episode_id>")
def analysis_trade_detail(episode_id: int):
    stock_query = request.args.get("q", "").strip()
    stock_sort = request.args.get("sort", "recent")
    if stock_sort not in STOCK_SORTS:
        stock_sort = "recent"
    with db_connect() as db:
        detail = trade_episode_detail(db, current_user_id(), episode_id)
    if detail is None:
        abort(404)
    return render_template(
        "index.html",
        page="trade_detail",
        trade_detail=detail,
        stock_sort=stock_sort,
        stock_query=stock_query,
        review_reason_types=REVIEW_REASON_TYPES,
        review_reason_labels=REVIEW_REASON_TYPE_LABELS,
        review_judgement_labels=REVIEW_JUDGEMENT_LABELS,
        review_main_problem_labels=REVIEW_MAIN_PROBLEM_LABELS,
        review_judgement_results=REVIEW_JUDGEMENT_RESULTS,
        review_main_problems=REVIEW_MAIN_PROBLEMS,
    )


@app.get("/analysis/stocks/kline/<stock_code>")
def stocks_kline(stock_code: str):
    try:
        stock_code = normalize_code(stock_code)
    except ValueError:
        abort(404)
    with db_connect() as db:
        stock = db.execute(
            "SELECT MAX(stock_name) stock_name FROM executions WHERE user_id = ? AND stock_code = ?",
            (current_user_id(), stock_code),
        ).fetchone()
        if not stock or not stock["stock_name"]:
            abort(404)
        def load_chart():
            return [
                {
                    "date": row["trade_date"], "open": float(row["open"]), "close": float(row["close"]),
                    "low": float(row["low"]), "high": float(row["high"]), "volume": kline_volume_lots(row), "amount": float(row["amount"] or 0),
                }
                for row in db.execute(
                    """SELECT trade_date, open, close, low, high, volume, amount, source FROM daily_prices
                    WHERE stock_code = ? AND open IS NOT NULL AND close IS NOT NULL
                    AND low IS NOT NULL AND high IS NOT NULL ORDER BY trade_date""",
                    (stock_code,),
                ).fetchall()
            ]

        chart = load_chart()
        sync_attempted = False
        sync_error = None
        if not chart:
            sync_attempted = True
            try:
                _, failed = sync_history_codes([stock_code], full=True)
                if failed:
                    sync_error = failed[0]
            except Exception as error:
                sync_error = f"{stock_code}：{error}"
    if sync_attempted:
        with db_connect() as db:
            chart = [
                {
                    "date": row["trade_date"], "open": float(row["open"]), "close": float(row["close"]),
                    "low": float(row["low"]), "high": float(row["high"]), "volume": kline_volume_lots(row), "amount": float(row["amount"] or 0),
                }
                for row in db.execute(
                    """SELECT trade_date, open, close, low, high, volume, amount, source FROM daily_prices
                    WHERE stock_code = ? AND open IS NOT NULL AND close IS NOT NULL
                    AND low IS NOT NULL AND high IS NOT NULL ORDER BY trade_date""",
                    (stock_code,),
                ).fetchall()
            ]
    with db_connect() as db:
        bs_points = kline_bs_points(db, current_user_id(), stock_code)
    return jsonify({
        "stock_code": stock_code, "stock_name": stock["stock_name"], "data": chart,
        "bs_points": bs_points, "sync_attempted": sync_attempted, "sync_error": sync_error,
    })


def review_optional_number(form, name: str) -> float | None:
    value = (form.get(name) or "").strip()
    if value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{name} 格式不正确")
    return number


@app.post("/analysis/trades/<int:episode_id>/review")
def save_trade_review(episode_id: int):
    user_id = current_user_id()
    sort = request.form.get("sort", "recent")
    if sort not in STOCK_SORTS:
        sort = "recent"
    query = request.form.get("q", "").strip()
    try:
        with db_connect() as db:
            episode = db.execute(
                "SELECT * FROM trade_episodes WHERE id = ? AND user_id = ?",
                (episode_id, user_id),
            ).fetchone()
            if episode is None:
                abort(404)
            action = request.form.get("action", "draft")
            review_status = "COMPLETED" if action == "complete" else "PENDING"
            if review_status == "COMPLETED" and episode["status"] != "CLOSED":
                review_status = "PENDING"
            if review_status not in REVIEW_STATUSES:
                raise ValueError("复盘状态无效")
            trade_reason = request.form.get("trade_reason", "").strip()
            sell_reason = request.form.get("sell_reason", "").strip()
            review_note = request.form.get("review_note", "").strip()
            next_action = request.form.get("next_action", "").strip()
            expected_profit_percent = review_optional_number(request.form, "expected_profit_percent")
            expected_target_price = review_optional_number(request.form, "expected_target_price")
            stop_loss_price = review_optional_number(request.form, "stop_loss_price")
            expected_holding_days = review_optional_number(request.form, "expected_holding_days")
            confidence_level = review_optional_number(request.form, "confidence_level")
            if confidence_level is not None:
                confidence_level = int(confidence_level)
                if confidence_level < 1 or confidence_level > 5:
                    raise ValueError("主观信心必须是 1 至 5")
            if expected_holding_days is not None:
                expected_holding_days = int(expected_holding_days)
                if expected_holding_days < 0:
                    raise ValueError("预期持有天数不能为负")
            if expected_profit_percent is not None and expected_profit_percent < 0:
                raise ValueError("预期上涨空间不能为负")
            if expected_target_price is not None and expected_target_price <= 0:
                raise ValueError("目标价格必须大于 0")
            if stop_loss_price is not None and stop_loss_price <= 0:
                raise ValueError("止损价格必须大于 0")
            judgement = request.form.get("judgement_result", "").strip() or None
            main_problem = request.form.get("main_problem", "").strip() or None
            if judgement is not None and judgement not in REVIEW_JUDGEMENT_RESULTS:
                raise ValueError("判断结果无效")
            if main_problem is not None and main_problem not in REVIEW_MAIN_PROBLEMS:
                raise ValueError("主要问题无效")
            reason_types = request.form.getlist("reason_type")
            for reason_type in reason_types:
                if reason_type not in REVIEW_REASON_TYPES:
                    raise ValueError("交易逻辑类型无效")
            if review_status == "COMPLETED" and judgement is None:
                raise ValueError("标记已复盘前请选择判断结果")

            now = datetime.now().isoformat(timespec="seconds")
            existing = db.execute(
                "SELECT id FROM trade_reviews WHERE trade_episode_id = ? AND user_id = ?",
                (episode_id, user_id),
            ).fetchone()
            completed_at = now if review_status == "COMPLETED" else None
            if existing is not None:
                db.execute(
                    """UPDATE trade_reviews SET review_status = ?, trade_reason = ?,
                    expected_profit_percent = ?, expected_target_price = ?, stop_loss_price = ?,
                    expected_holding_days = ?, confidence_level = ?, sell_reason = ?,
                    judgement_result = ?, main_problem = ?, review_note = ?, next_action = ?,
                    updated_at = ?, completed_at = ? WHERE id = ? AND user_id = ?""",
                    (review_status, trade_reason, expected_profit_percent, expected_target_price,
                     stop_loss_price, expected_holding_days, confidence_level, sell_reason,
                     judgement, main_problem, review_note, next_action, now, completed_at,
                     existing["id"], user_id),
                )
                review_id = existing["id"]
            else:
                cursor = db.execute(
                    """INSERT INTO trade_reviews
                    (user_id, trade_episode_id, review_status, trade_reason,
                     expected_profit_percent, expected_target_price, stop_loss_price,
                     expected_holding_days, confidence_level, sell_reason,
                     judgement_result, main_problem, review_note, next_action,
                     original_opening_execution_id, stock_code_snapshot,
                     opened_at_snapshot, closed_at_snapshot, created_at, updated_at, completed_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (user_id, episode_id, review_status, trade_reason,
                     expected_profit_percent, expected_target_price, stop_loss_price,
                     expected_holding_days, confidence_level, sell_reason,
                     judgement, main_problem, review_note, next_action,
                     episode["opening_execution_id"], episode["stock_code"],
                     episode["opened_at"], episode["closed_at"], now, now, completed_at),
                )
                review_id = cursor.lastrowid
            db.execute("DELETE FROM trade_review_reason_types WHERE trade_review_id = ?", (review_id,))
            for reason_type in reason_types:
                db.execute(
                    "INSERT INTO trade_review_reason_types (trade_review_id, reason_type) VALUES (?, ?)",
                    (review_id, reason_type),
                )
        if review_status == "COMPLETED":
            flash("复盘已完成并保存", "success")
        else:
            flash("复盘草稿已保存", "success")
    except ValueError as error:
        flash(f"保存失败：{error}", "error")
    return redirect(url_for("analysis_trade_detail", episode_id=episode_id, sort=sort, q=query or None))


@app.post("/analysis/trades/<int:episode_id>/sync-prices")
def sync_trade_prices(episode_id: int):
    user_id = current_user_id()
    sort = request.form.get("sort", "recent")
    if sort not in STOCK_SORTS:
        sort = "recent"
    query = request.form.get("q", "").strip()
    with db_connect() as db:
        episodes = [row for row in trade_episode_rows(db, user_id) if row["id"] == episode_id]
        if not episodes:
            abort(404)
        episode = episodes[0]
        end_date = episode["sell_date"] if episode["status"] == "CLOSED" and episode["sell_date"] else date.today().isoformat()
        try:
            inserted = sync_daily_prices(db, episode["stock_code"], episode["buy_date"], end_date)
            metrics = calculate_excursion_metrics(db, user_id, episode)
        except Exception as error:
            flash(f"行情同步失败：{error}", "error")
            inserted = None
            metrics = None
        if inserted is not None:
            if metrics is None:
                flash(f"已同步 {inserted} 条行情，但该区间暂无可用最高/最低价数据。", "warning")
            else:
                flash(f"已同步 {inserted} 条行情并计算 MFE/MAE。", "success")
    return redirect(url_for("analysis_trade_detail", episode_id=episode_id, sort=sort, q=query or None))


@app.get("/analysis/summary")
def analysis_summary():
    performance_period = request.args.get("period", "month")
    if performance_period not in {"day", "month", "year"}:
        performance_period = "month"
    year_value = request.args.get("year", "")
    try:
        performance_year = int(year_value) if year_value else None
    except ValueError:
        performance_year = None
    with db_connect() as db:
        return render_template("index.html", page="summary", data=analysis_data(db, performance_year=performance_year), performance_period=performance_period)


@app.get("/analysis/watchlist")
def analysis_watchlist():
    query = request.args.get("q", "").strip()
    priority = request.args.get("priority", "all")
    sort = request.args.get("sort", "priority")
    direction = request.args.get("direction", "desc")
    if direction not in ("asc", "desc"):
        direction = "desc"
    with db_connect() as db:
        watchlist = watchlist_data(db, current_user_id(), query, priority, sort, direction)
    return render_template("index.html", page="watchlist", watchlist=watchlist, auto_sync=auto_sync_status())


@app.get("/analysis/watchlist/quotes")
def watchlist_quotes():
    query = request.args.get("q", "").strip()
    priority = request.args.get("priority", "all")
    sort = request.args.get("sort", "priority")
    direction = request.args.get("direction", "desc")
    with db_connect() as db:
        watchlist = watchlist_data(db, current_user_id(), query, priority, sort, direction)
    response = jsonify({
        "items": [
            {key: item[key] for key in (
                "stock_code", "latest_price", "price_date", "price_fetched_at", "change_amount",
                "change_rate", "intraday_rebound",
            )}
            for item in watchlist["items"]
        ],
    })
    response.headers["Cache-Control"] = "no-store"
    return response


@app.post("/analysis/watchlist")
def save_watchlist_stock():
    created = False
    try:
        stock_code = normalize_code(request.form.get("stock_code", ""))
        stock_name = request.form.get("stock_name", "").strip()
        priority = int(request.form.get("priority", "2"))
        note = request.form.get("note", "").strip()
        if not stock_name:
            raise ValueError("证券名称不能为空")
        if len(stock_name) > 50 or len(note) > 500:
            raise ValueError("证券名称或关注理由长度超限")
        if priority not in (1, 2, 3):
            raise ValueError("优先级无效")
        now = datetime.now().isoformat(timespec="seconds")
        with db_connect() as db:
            existing = db.execute("SELECT id FROM watchlist_stocks WHERE user_id = ? AND stock_code = ?", (current_user_id(), stock_code)).fetchone()
            if existing:
                db.execute("UPDATE watchlist_stocks SET stock_name=?, priority=?, note=?, updated_at=? WHERE id=? AND user_id=?", (stock_name, priority, note, now, existing["id"], current_user_id()))
                flash(f"已更新重点观察：{stock_name}", "success")
            else:
                db.execute("INSERT INTO watchlist_stocks (user_id, stock_code, stock_name, priority, note, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)", (current_user_id(), stock_code, stock_name, priority, note, now, now))
                created = True
        if created:
            synced, failed = sync_history_codes([stock_code], full=True)
            if failed:
                flash(f"已加入重点观察：{stock_name}；历史行情同步失败，可稍后重试。", "warning")
            else:
                flash(f"已加入重点观察：{stock_name}，已同步 {synced} 条历史行情。", "success")
    except (TypeError, ValueError) as error:
        flash(str(error), "error")
    return redirect(url_for("analysis_watchlist"))


@app.post("/analysis/watchlist/<int:watch_id>/delete")
def delete_watchlist_stock(watch_id: int):
    with db_connect() as db:
        item = db.execute("SELECT stock_name FROM watchlist_stocks WHERE id = ? AND user_id = ?", (watch_id, current_user_id())).fetchone()
        if item is None:
            abort(404)
        db.execute("DELETE FROM watchlist_stocks WHERE id = ? AND user_id = ?", (watch_id, current_user_id()))
    flash(f"已移出重点观察：{item['stock_name']}", "success")
    return redirect(url_for("analysis_watchlist"))


def watchlist_sync_codes():
    with db_connect() as db:
        items = db.execute("SELECT stock_code FROM watchlist_stocks WHERE user_id = ?", (current_user_id(),)).fetchall()
    return [item["stock_code"] for item in items]


def watchlist_sync_response(message: str, category: str):
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"message": message, "category": category})
    flash(message, category)
    return redirect(url_for("analysis_watchlist"))


def portfolio_sync_response(message: str, category: str):
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"message": message, "category": category})
    flash(message, category)
    return redirect(url_for("analysis_portfolio"))


def kline_bs_points(db: sqlite3.Connection, user_id: int, stock_code: str) -> list[dict]:
    return [
        {
            "date": row["trade_date"],
            "time": row["trade_time"],
            "action": row["action"],
            "price": float(row["deal_price"]),
            "quantity": float(row["quantity"]),
        }
        for row in db.execute(
            """SELECT trade_date, trade_time, action, deal_price, quantity
            FROM executions WHERE user_id = ? AND stock_code = ?
            ORDER BY trade_date, trade_time, id""",
            (user_id, stock_code),
        ).fetchall()
    ]


def kline_volume_lots(row: sqlite3.Row) -> float:
    return float(row["volume"] or 0) / 100


@app.post("/analysis/watchlist/sync-prices")
def sync_watchlist_prices():
    stock_codes = watchlist_sync_codes()
    if not stock_codes:
        return watchlist_sync_response("重点观察列表为空，无需同步实时行情。", "warning")
    synced, failed = sync_realtime_codes(stock_codes)
    message = f"已同步 {synced} 条实时行情。" if not failed else f"实时行情部分同步失败：{'；'.join(failed)}"
    category = "success" if not failed else "warning"
    return watchlist_sync_response(message, category)


@app.post("/analysis/watchlist/sync-history")
def sync_watchlist_history():
    stock_codes = watchlist_sync_codes()
    if not stock_codes:
        return watchlist_sync_response("重点观察列表为空，无需同步历史行情。", "warning")
    synced, failed = sync_history_codes(stock_codes)
    message = f"已同步 {synced} 条历史行情。" if not failed else f"历史行情部分同步失败：{'；'.join(failed)}"
    category = "success" if not failed else "warning"
    return watchlist_sync_response(message, category)


@app.get("/analysis/watchlist/kline/<stock_code>")
def watchlist_kline(stock_code: str):
    try:
        stock_code = normalize_code(stock_code)
    except ValueError:
        abort(404)
    with db_connect() as db:
        item = db.execute(
            "SELECT stock_name FROM watchlist_stocks WHERE user_id = ? AND stock_code = ?",
            (current_user_id(), stock_code),
        ).fetchone()
        if item is None:
            abort(404)
    with db_connect() as db:
        bs_points = kline_bs_points(db, current_user_id(), stock_code)
        data = [
            {
                "date": row["trade_date"], "open": float(row["open"]), "close": float(row["close"]),
                "low": float(row["low"]), "high": float(row["high"]), "volume": kline_volume_lots(row), "amount": float(row["amount"] or 0),
            }
            for row in db.execute(
                """SELECT trade_date, open, close, low, high, volume, amount, source FROM daily_prices
                WHERE stock_code = ? AND open IS NOT NULL AND close IS NOT NULL
                AND low IS NOT NULL AND high IS NOT NULL ORDER BY trade_date""",
                (stock_code,),
            ).fetchall()
        ]
    return jsonify({"stock_code": stock_code, "stock_name": item["stock_name"], "data": data, "bs_points": bs_points})


@app.get("/analysis/portfolio")
def analysis_portfolio():
    with db_connect() as db:
        portfolio = portfolio_analysis_data(db, current_user_id())
    return render_template("index.html", page="portfolio", portfolio=portfolio, auto_sync=auto_sync_status())


@app.get("/analysis/portfolio/quotes")
def portfolio_quotes():
    with db_connect() as db:
        portfolio = portfolio_analysis_data(db, current_user_id())
    response = jsonify({
        "summary": portfolio["summary"],
        "positions": [
            {key: item[key] for key in ("stock_code", "latest_price", "price_date", "price_fetched_at", "change_rate", "unrealized_profit", "unrealized_rate")}
            for item in portfolio["positions"]
        ],
    })
    response.headers["Cache-Control"] = "no-store"
    return response


@app.get("/analysis/portfolio/indexes")
def portfolio_indexes():
    indexes, error = sync_market_indexes()
    status = {"data": indexes, "updated_at": datetime.now().isoformat(timespec="seconds") if indexes else None, "error": error}
    return jsonify(status)


@app.get("/analysis/hot-sectors")
def analysis_hot_sectors():
    return render_template("index.html", page="hot_sectors")


@app.get("/api/hot-sectors")
def hot_sectors_api():
    force = request.args.get("force", "") == "1"
    payload = hot_sectors_data(force=force)
    response = jsonify(payload)
    response.headers["Cache-Control"] = "no-store"
    return response


@app.get("/analysis/portfolio/kline/<stock_code>")
def portfolio_kline(stock_code: str):
    try:
        stock_code = normalize_code(stock_code)
    except ValueError:
        abort(404)
    with db_connect() as db:
        position = db.execute(
            """SELECT p.stock_code, MAX(p.stock_name) stock_name, SUM(p.quantity) quantity,
            SUM(p.total_cost) total_cost, SUM(p.total_cost) / NULLIF(SUM(p.quantity), 0) avg_cost
            FROM positions p WHERE p.user_id = ? AND p.stock_code = ? GROUP BY p.stock_code""",
            (current_user_id(), stock_code),
        ).fetchone()
        if position is None:
            abort(404)
        chart = [
            {
                "date": row["trade_date"], "open": float(row["open"]), "close": float(row["close"]),
                "low": float(row["low"]), "high": float(row["high"]), "volume": kline_volume_lots(row), "amount": float(row["amount"] or 0),
            }
            for row in db.execute(
                """SELECT trade_date, open, close, low, high, volume, amount, source FROM daily_prices
                WHERE stock_code = ? AND open IS NOT NULL AND close IS NOT NULL
                AND low IS NOT NULL AND high IS NOT NULL ORDER BY trade_date""",
                (stock_code,),
            ).fetchall()
        ]
        bs_points = kline_bs_points(db, current_user_id(), stock_code)
    return jsonify({
        "stock_code": stock_code, "stock_name": position["stock_name"],
        "avg_cost": float(position["avg_cost"] or 0), "data": chart, "bs_points": bs_points,
    })


@app.post("/analysis/portfolio/sync-prices")
def sync_portfolio_prices():
    user_id = current_user_id()
    with db_connect() as db:
        positions = db.execute(
            """SELECT stock_code, MIN(buy_date) first_buy_date FROM positions
            WHERE user_id = ? GROUP BY stock_code ORDER BY stock_code""",
            (user_id,),
        ).fetchall()
        if not positions:
            flash("当前没有持仓，无需同步行情。", "warning")
            return redirect(url_for("analysis_portfolio"))
        synced = 0
        failed = []
        volatility_start = date.today() - timedelta(days=100)
        for position in positions:
            first_buy = datetime.strptime(position["first_buy_date"], "%Y-%m-%d").date()
            start_date = min(first_buy, volatility_start).isoformat()
            try:
                synced += sync_daily_prices(db, position["stock_code"], start_date, date.today().isoformat())
            except Exception as error:
                failed.append(f"{position['stock_code']}：{error}")
        if failed:
            flash(f"已处理 {len(positions) - len(failed)} 只持仓，{len(failed)} 只同步失败：{'；'.join(failed)}", "warning")
        else:
            flash(f"已为 {len(positions)} 只持仓同步 {synced} 条日线行情。", "success")
    return redirect(url_for("analysis_portfolio"))


@app.post("/analysis/portfolio/sync-realtime")
def sync_portfolio_realtime():
    with db_connect() as db:
        stock_codes = [
            row["stock_code"]
            for row in db.execute(
                "SELECT DISTINCT stock_code FROM positions WHERE user_id = ? ORDER BY stock_code",
                (current_user_id(),),
            ).fetchall()
        ]
    if not stock_codes:
        return portfolio_sync_response("当前没有持仓，无需同步实时行情。", "warning")
    synced, failed = sync_realtime_codes(stock_codes)
    _, index_error = sync_market_indexes()
    if index_error:
        failed.append(f"指数行情：{index_error}")
    message = f"已同步 {synced} 条实时行情。" if not failed else f"实时行情部分同步失败：{'；'.join(failed)}"
    return portfolio_sync_response(message, "success" if not failed else "warning")


@app.post("/analysis/portfolio/sync-history")
def sync_portfolio_history():
    with db_connect() as db:
        stock_codes = [
            row["stock_code"]
            for row in db.execute(
                "SELECT DISTINCT stock_code FROM positions WHERE user_id = ? ORDER BY stock_code",
                (current_user_id(),),
            ).fetchall()
        ]
    if not stock_codes:
        return portfolio_sync_response("当前没有持仓，无需同步历史行情。", "warning")
    synced, failed = sync_history_codes(stock_codes)
    message = f"已同步 {synced} 条历史行情。" if not failed else f"历史行情部分同步失败：{'；'.join(failed)}"
    return portfolio_sync_response(message, "success" if not failed else "warning")


@app.get("/analysis/reviews")
def analysis_reviews():
    status_filter = request.args.get("status", "all")
    if status_filter not in ("pending", "completed", "all"):
        status_filter = "all"
    today = date.today()
    default_start = (today - timedelta(days=6)).isoformat()
    default_end = today.isoformat()

    def parse_date(value: str) -> str | None:
        value = value.strip()
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").strftime("%Y-%m-%d")
        except (TypeError, ValueError):
            return None

    start_date = parse_date(request.args.get("start_date", ""))
    end_date = parse_date(request.args.get("end_date", ""))
    if start_date is None and end_date is None:
        start_date, end_date = default_start, default_end
    else:
        start_date = start_date or default_start
        end_date = end_date or default_end
        if start_date > end_date:
            start_date, end_date = end_date, start_date
    with db_connect() as db:
        user_id = current_user_id()
        episodes = review_episodes(db, user_id, status_filter, start_date, end_date)
        progress = review_progress_data(db, user_id)
    episode_rows = []
    for row in episodes:
        item = dict(row)
        reason_codes = [code for code in (item.get("reason_types") or "").split(",") if code]
        item["reason_type_labels"] = [
            REVIEW_REASON_TYPE_LABELS[code] for code in reason_codes if code in REVIEW_REASON_TYPE_LABELS
        ]
        episode_rows.append(item)
    episodes = episode_rows
    wins = sum(1 for row in episodes if row["profit"] > 0)
    range_summary = {
        "trades": len(episodes),
        "wins": wins,
        "profit": sum(float(row["profit"]) for row in episodes),
    }
    quick_ranges = [
        {"days": 7, "start": (today - timedelta(days=6)).isoformat(), "end": today.isoformat()},
        {"days": 15, "start": (today - timedelta(days=14)).isoformat(), "end": today.isoformat()},
        {"days": 30, "start": (today - timedelta(days=29)).isoformat(), "end": today.isoformat()},
    ]
    return render_template(
        "index.html",
        page="reviews",
        review_episodes=episodes,
        review_progress=progress,
        status_filter=status_filter,
        start_date=start_date,
        end_date=end_date,
        today=today.isoformat(),
        range_summary=range_summary,
        quick_ranges=quick_ranges,
        review_judgement_labels=REVIEW_JUDGEMENT_LABELS,
        review_main_problem_labels=REVIEW_MAIN_PROBLEM_LABELS,
    )


@app.get("/analysis/discipline")
def analysis_discipline():
    with db_connect() as db:
        data = calculate_discipline_data(db)
        overall = analysis_data(db)
        mistakes = mistake_stats(db, current_user_id())
        return render_template(
            "index.html", 
            page="discipline", 
            discipline=data,
            summary=overall["summary"],
            mistake_stats=mistakes,
            mistake_labels=REVIEW_MAIN_PROBLEM_LABELS,
        )


@app.get("/analysis/recent")
def analysis_recent():
    try:
        days = int(request.args.get("days", 7))
        if days not in (7, 15, 30, 60):
            days = 7
    except (TypeError, ValueError):
        days = 7
    with db_connect() as db:
        data = calculate_recent_data(db, days)
        return render_template("index.html", page="recent", recent=data, days=days)


@app.get("/analysis/timing")
def analysis_timing():
    with db_connect() as db:
        data = calculate_timing_data(db)
        return render_template("index.html", page="timing", timing=data)


@app.get("/analysis/strategy")
def analysis_strategy():
    filters = strategy_filters(request.args)
    with db_connect() as db:
        data = calculate_strategy_data(db, filters)
        reason_stats = strategy_type_stats(db, current_user_id())
    return render_template(
        "index.html",
        page="strategy",
        strategy=data,
        strategy_filters=filters,
        strategy_sessions=STRATEGY_SESSIONS,
        strategy_weekdays=STRATEGY_WEEKDAYS,
        reason_stats=reason_stats,
        reason_labels=REVIEW_REASON_TYPE_LABELS,
    )


@app.get("/admin")
def admin():
    return redirect(url_for("statements"))


@app.get("/admin/alert-types")
def alert_types_management():
    query = request.query_string.decode("utf-8")
    target = url_for("three_day_dip_management")
    return redirect(f"{target}?{query}" if query else target)


@app.get("/admin/alert-types/three-day-dip")
def three_day_dip_management():
    backtest = None
    backtest_error = None
    selected_stock_code = request.args.get("stock_code", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    with db_connect() as db:
        alert_type, params = alert_type_data(db)
        watchlist = db.execute(
            """SELECT stock_code, stock_name FROM watchlist_stocks WHERE user_id = ?
            ORDER BY priority DESC, stock_name COLLATE NOCASE, stock_code""",
            (current_user_id(),),
        ).fetchall()
        if request.args.get("run_test") == "1":
            stock = db.execute(
                """SELECT stock_code, stock_name FROM watchlist_stocks
                WHERE user_id = ? AND stock_code = ?""",
                (current_user_id(), selected_stock_code),
            ).fetchone()
            try:
                start_date = normalize_date(start_date)
                end_date = normalize_date(end_date)
                if start_date > end_date:
                    raise ValueError("开始日期不能晚于截止日期")
                if stock is None:
                    raise ValueError("请选择当前用户重点观察中的股票")
                backtest = backtest_three_day_dip(db, stock, start_date, end_date, params)
            except (TypeError, ValueError) as error:
                backtest_error = str(error) if str(error) else "测试条件无效，请检查股票和日期。"
    return render_template(
        "index.html", page="three_day_dip", alert_type=alert_type, alert_params=params,
        alert_watchlist=watchlist,
        selected_stock_code=selected_stock_code, test_start_date=start_date, test_end_date=end_date,
        backtest=backtest, backtest_error=backtest_error,
    )


@app.post("/admin/alert-types/three-day-dip")
def save_three_day_dip_alert_type():
    try:
        params = three_day_dip_params({
            "decline_days": int(request.form.get("decline_days", "2")),
            "require_bearish_candles": "require_bearish_candles" in request.form,
            "require_declining_closes": "require_declining_closes" in request.form,
            "max_signal_low_above_prior_ratio": float(request.form.get("max_signal_low_above_prior_percent", "4")) / 100,
            "min_decline_ratio": float(request.form.get("min_decline_percent", "8")) / 100,
            "max_volume_ratio": float(request.form.get("max_volume_percent", "100")) / 100,
            "exhaustion_min_repair_ratio": float(request.form.get("exhaustion_min_repair_percent", "2")) / 100,
            "exhaustion_max_body_range_ratio": float(request.form.get("exhaustion_max_body_percent", "10")) / 100,
            "exhaustion_min_change_ratio": float(request.form.get("exhaustion_min_change_percent", "-1.5")) / 100,
            "exhaustion_max_change_ratio": float(request.form.get("exhaustion_max_change_percent", "3")) / 100,
            "shadow_min_repair_ratio": float(request.form.get("shadow_min_repair_percent", "1.5")) / 100,
            "shadow_max_body_range_ratio": float(request.form.get("shadow_max_body_percent", "20")) / 100,
            "shadow_min_lower_shadow_body_ratio": float(request.form.get("shadow_min_lower_shadow_body_ratio", "1")),
            "shadow_min_change_ratio": float(request.form.get("shadow_min_change_percent", "-1.5")) / 100,
            "shadow_max_change_ratio": float(request.form.get("shadow_max_change_percent", "3")) / 100,
            "reversal_min_decline_ratio": float(request.form.get("reversal_min_decline_percent", "10")) / 100,
            "reversal_min_repair_ratio": float(request.form.get("reversal_min_repair_percent", "6")) / 100,
            "reversal_min_recovery_range_ratio": float(request.form.get("reversal_min_recovery_percent", "60")) / 100,
            "reversal_min_change_ratio": float(request.form.get("reversal_min_change_percent", "4")) / 100,
            "reversal_max_volume_ratio": float(request.form.get("reversal_max_volume_percent", "100")) / 100,
            "intraday_candidate_enabled": "intraday_candidate_enabled" in request.form,
            "close_confirmation_enabled": "close_confirmation_enabled" in request.form,
        })
        enabled = int("enabled" in request.form)
        now = datetime.now().isoformat(timespec="seconds")
        with db_connect() as db:
            db.execute(
                """UPDATE alert_types SET params_json = ?, enabled = ?, updated_at = ? WHERE code = ?""",
                (json.dumps(params, ensure_ascii=False), enabled, now, THREE_DAY_DIP_CODE),
            )
        flash("已保存三日低吸提醒参数。", "success")
    except (TypeError, ValueError):
        flash("提醒参数格式无效，请检查数值。", "error")
    return redirect(url_for("three_day_dip_management"))


@app.get("/admin/alert-types/intraday-rebound")
def intraday_rebound_management():
    with db_connect() as db:
        alert_type, params = intraday_rebound_alert_type_data(db)
        states = db.execute(
            """SELECT watch.stock_code, watch.stock_name, watch.priority, state.trade_date,
            state.last_matched, state.last_evaluated_at, state.candidate_triggered_at,
            quote.price, quote.quote_minute
            FROM watchlist_stocks AS watch
            LEFT JOIN alert_rule_states AS state ON state.user_id = watch.user_id
            AND state.stock_code = watch.stock_code
            AND state.alert_type_id = ?
            AND state.trade_date = (SELECT MAX(trade_date) FROM intraday_quotes WHERE stock_code = watch.stock_code)
            LEFT JOIN intraday_quotes AS quote ON quote.stock_code = watch.stock_code
            AND quote.quote_minute = (
                SELECT MAX(quote_minute) FROM intraday_quotes WHERE stock_code = watch.stock_code
            )
            WHERE watch.user_id = ? ORDER BY watch.priority DESC, watch.stock_name COLLATE NOCASE, watch.stock_code""",
            (alert_type["id"], current_user_id()),
        ).fetchall()
        notifications = db.execute(
            """SELECT stock_code, stock_name, content, quote_time, created_at FROM notifications
            WHERE user_id = ? AND alert_type_id = ? ORDER BY created_at DESC LIMIT 20""",
            (current_user_id(), alert_type["id"]),
        ).fetchall()
    return render_template(
        "index.html", page="intraday_rebound", alert_type=alert_type, alert_params=params,
        intraday_states=states, intraday_notifications=notifications,
    )


@app.post("/admin/alert-types/intraday-rebound")
def save_intraday_rebound_alert_type():
    try:
        params = intraday_rebound_params({
            "lookback_minutes": int(request.form.get("lookback_minutes", "120")),
            "min_drop_ratio": float(request.form.get("min_drop_percent", "2")) / 100,
            "min_rebound_ratio": float(request.form.get("min_rebound_percent", "0.8")) / 100,
            "min_trough_age_minutes": int(request.form.get("min_trough_age_minutes", "5")),
            "min_volume_multiple": float(request.form.get("min_volume_multiple", "1.5")),
            "enabled": "enabled" in request.form,
        })
        now = datetime.now().isoformat(timespec="seconds")
        with db_connect() as db:
            db.execute(
                "UPDATE alert_types SET params_json = ?, enabled = ?, updated_at = ? WHERE code = ?",
                (json.dumps(params, ensure_ascii=False), int(params["enabled"]), now, INTRADAY_REBOUND_CODE),
            )
        flash("已保存日内反弹提醒参数。", "success")
    except (TypeError, ValueError):
        flash("提醒参数格式无效，请检查数值。", "error")
    return redirect(url_for("intraday_rebound_management"))


@app.get("/admin/alert-types/three-day-dip/pool")
def three_day_dip_pool():
    with db_connect() as db:
        pool = alert_signal_pool_data(db, current_user_id())
    return render_template("index.html", page="three_day_dip_pool", signal_pool=pool)


@app.post("/admin/alert-types/three-day-dip/pool")
def add_three_day_dip_sample():
    try:
        stock_code = normalize_code(request.form.get("stock_code", ""))
        stock_name = request.form.get("stock_name", "").strip()
        signal_date = normalize_date(request.form.get("signal_date", ""))
        note = request.form.get("note", "").strip()
        if not stock_name or len(stock_name) > 50 or len(note) > 500:
            raise ValueError("证券名称不能为空，备注不能超过 500 字")
        with db_connect() as db:
            prices = [dict(row) for row in reversed(db.execute(
                """SELECT trade_date, open, high, low, close, volume FROM daily_prices
                WHERE stock_code = ? AND trade_date <= ? AND open IS NOT NULL AND high IS NOT NULL
                AND low IS NOT NULL AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 4""",
                (stock_code, signal_date),
            ).fetchall())]
        if len(prices) < 4 or prices[-1]["trade_date"] != signal_date:
            start_date = (date.fromisoformat(signal_date) - timedelta(days=12)).isoformat()
            with db_connect() as db:
                sync_daily_prices(db, stock_code, start_date, signal_date)
                prices = [dict(row) for row in reversed(db.execute(
                    """SELECT trade_date, open, high, low, close, volume FROM daily_prices
                    WHERE stock_code = ? AND trade_date <= ? AND open IS NOT NULL AND high IS NOT NULL
                    AND low IS NOT NULL AND close IS NOT NULL ORDER BY trade_date DESC LIMIT 4""",
                    (stock_code, signal_date),
                ).fetchall())]
        if len(prices) < 4 or prices[-1]["trade_date"] != signal_date:
            raise ValueError("信号日前行情不足，无法建立样本")
        with db_connect() as db:
            alert_type, params = alert_type_data(db)
            result = evaluate_three_day_dip(prices, params, enforce_volume=True)
            pattern_type = request.form.get("pattern_type", "").strip() or result.get("pattern_type")
            if pattern_type not in ("EXHAUSTION", "SHADOW_STOP", "STRONG_REVERSAL"):
                raise ValueError("当前规则未识别形态，请手工选择形态类型")
            upsert_alert_signal_sample(
                db, current_user_id(), alert_type["id"], stock_code, stock_name, prices, result,
                source="MANUAL", review_status="CONFIRMED", note=note, pattern_type=pattern_type,
            )
        flash(f"已加入三日低吸池：{stock_name} {signal_date}", "success")
    except (RuntimeError, TypeError, ValueError) as error:
        flash(f"加入样本失败：{error}", "error")
    return redirect(url_for("three_day_dip_pool"))


@app.post("/admin/alert-types/three-day-dip/pool/<int:sample_id>/review")
def review_three_day_dip_sample(sample_id: int):
    status = request.form.get("review_status", "")
    note = request.form.get("note", "").strip()
    if status not in ("CONFIRMED", "REJECTED") or len(note) > 500:
        abort(400)
    now = datetime.now().isoformat(timespec="seconds")
    with db_connect() as db:
        cursor = db.execute(
            """UPDATE alert_signal_samples SET review_status = ?, note = ?, reviewed_at = ?, updated_at = ?
            WHERE id = ? AND user_id = ?""",
            (status, note, now, now, sample_id, current_user_id()),
        )
    if cursor.rowcount == 0:
        abort(404)
    flash("样本审核结果已保存。", "success")
    return redirect(url_for("three_day_dip_pool"))


@app.post("/admin/alert-types/three-day-dip/pool/scan")
def scan_three_day_dip_history():
    try:
        start_date = normalize_date(request.form.get("start_date", ""))
        end_date = normalize_date(request.form.get("end_date", ""))
        if start_date > end_date:
            raise ValueError("开始日期不能晚于截止日期")
        created = 0
        with db_connect() as db:
            alert_type, params = alert_type_data(db)
            watches = db.execute(
                "SELECT stock_code, stock_name FROM watchlist_stocks WHERE user_id = ?", (current_user_id(),)
            ).fetchall()
            for stock in watches:
                result = backtest_three_day_dip(db, stock, start_date, end_date, params)
                for hit in result["hits"]:
                    evaluated = evaluate_three_day_dip(hit["candles"], params, enforce_volume=True)
                    upsert_alert_signal_sample(
                        db, current_user_id(), alert_type["id"], stock["stock_code"], stock["stock_name"],
                        hit["candles"], evaluated,
                    )
                    created += 1
        flash(f"历史扫描完成，处理 {created} 个命中样本。", "success")
    except (TypeError, ValueError) as error:
        flash(f"历史扫描失败：{error}", "error")
    return redirect(url_for("three_day_dip_pool"))


@app.get("/analysis/notifications")
def notifications():
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    with db_connect() as db:
        data = notification_page_data(db, current_user_id(), page)
    return render_template("index.html", page="notifications", notification_data=data)


@app.get("/api/notifications")
def notifications_api():
    with db_connect() as db:
        unread, notifications = notification_rows(db, current_user_id(), 20)
    return jsonify({"unread_count": unread, "notifications": notifications})


@app.post("/api/notifications/<int:notification_id>/read")
def read_notification(notification_id: int):
    with db_connect() as db:
        cursor = db.execute(
            """UPDATE notifications SET read_at = COALESCE(read_at, ?)
            WHERE id = ? AND user_id = ?""",
            (datetime.now().isoformat(timespec="seconds"), notification_id, current_user_id()),
        )
    if cursor.rowcount == 0:
        abort(404)
    return jsonify({"ok": True})


@app.post("/api/notifications/read-all")
def read_all_notifications():
    with db_connect() as db:
        cursor = db.execute(
            "UPDATE notifications SET read_at = ? WHERE user_id = ? AND read_at IS NULL",
            (datetime.now().isoformat(timespec="seconds"), current_user_id()),
        )
    return jsonify({"ok": True, "updated": cursor.rowcount})


@app.get("/admin/account")
def account_management():
    with db_connect() as db:
        snapshots = db.execute(
            """SELECT * FROM account_snapshots WHERE user_id = ?
            ORDER BY snapshot_date DESC, id DESC LIMIT 20""",
            (current_user_id(),),
        ).fetchall()
    return render_template(
        "index.html", page="account", snapshots=snapshots,
        latest_snapshot=snapshots[0] if snapshots else None, today=date.today().isoformat(),
    )


@app.post("/admin/account")
def save_account_snapshot():
    try:
        snapshot_date = normalize_date(request.form.get("snapshot_date", ""))
        total_assets = manual_number("total_assets", "账户总资产", positive=True)
        cash_text = request.form.get("available_cash", "").strip()
        available_cash = None if not cash_text else manual_number("available_cash", "可用资金")
        note = request.form.get("note", "").strip()
        if len(note) > 200:
            raise ValueError("备注不能超过 200 个字符")
        if available_cash is not None and available_cash > total_assets:
            raise ValueError("可用资金不能大于账户总资产")
        now = datetime.now().isoformat(timespec="seconds")
        with db_connect() as db:
            db.execute(
                """INSERT INTO account_snapshots
                (user_id, snapshot_date, total_assets, available_cash, note, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(user_id, snapshot_date) DO UPDATE SET
                total_assets = excluded.total_assets, available_cash = excluded.available_cash,
                note = excluded.note, updated_at = excluded.updated_at""",
                (current_user_id(), snapshot_date, total_assets, available_cash, note, now, now),
            )
        flash(f"已保存 {snapshot_date} 的账户资金快照。", "success")
    except (TypeError, ValueError) as error:
        flash(str(error), "error")
    return redirect(url_for("account_management"))


@app.get("/admin/statements")
def statements():
    try:
        current_page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        current_page = 1
    with db_connect() as db:
        user_id = current_user_id()
        total = db.execute("SELECT COUNT(*) FROM executions WHERE user_id = ?", (user_id,)).fetchone()[0]
        total_pages = max(1, math.ceil(total / STATEMENTS_PER_PAGE))
        current_page = min(current_page, total_pages)
        offset = (current_page - 1) * STATEMENTS_PER_PAGE
        executions = db.execute(
            "SELECT * FROM executions WHERE user_id = ? ORDER BY trade_date DESC, trade_time DESC, id DESC LIMIT ? OFFSET ?",
            (user_id, STATEMENTS_PER_PAGE, offset),
        ).fetchall()
        counts = db.execute("SELECT COUNT(*) executions, COUNT(DISTINCT stock_code) stocks FROM executions WHERE user_id = ?", (user_id,)).fetchone()
        return render_template("index.html", page="statements", executions=executions, counts=counts,
                               current_page=current_page, total_pages=total_pages, total=total,
                               today=date.today().isoformat())


@app.post("/admin/manual")
def create_manual_execution():
    try:
        trade_date = normalize_date(request.form.get("trade_date", ""))
        trade_time = normalize_time(request.form.get("trade_time", ""))
        stock_code = normalize_code(request.form.get("stock_code", ""))
        stock_name = request.form.get("stock_name", "").strip() or "未知证券"
        action = normalize_action(request.form.get("action", ""), 1)
        quantity = manual_number("quantity", "成交数量", positive=True)
        deal_price = manual_number("deal_price", "成交价格", positive=True)
        commission = manual_number("commission", "佣金")
        stamp_tax = manual_number("stamp_tax", "印花税")
        transfer_fee = manual_number("transfer_fee", "过户费")
    except (ValueError, TypeError) as error:
        flash(f"补录失败：{error}", "error")
        return redirect(url_for("statements"))

    user_id = current_user_id()
    job_id = uuid.uuid4().hex
    created_at = datetime.now().isoformat(timespec="seconds")
    fingerprint = hashlib.sha256(f"MANUAL|{user_id}|{job_id}".encode("utf-8")).hexdigest()
    raw = {
        "trade_date": trade_date,
        "trade_time": trade_time,
        "stock_code": stock_code,
        "stock_name": stock_name,
        "action": action,
        "quantity": quantity,
        "deal_price": deal_price,
    }
    with db_connect() as db:
        db.execute(
            """INSERT INTO import_jobs
            (id, user_id, filename, status, total_rows, valid_rows, invalid_rows, duplicate_rows, created_at, imported_at)
            VALUES (?, ?, '当天成交补录', 'IMPORTED', 1, 1, 0, 0, ?, ?)""",
            (job_id, user_id, created_at, created_at),
        )
        db.execute(
            """INSERT INTO executions
            (user_id, import_job_id, fingerprint, trade_date, trade_time, stock_code, stock_name, action,
             raw_quantity, quantity, deal_id, deal_price, deal_amount, commission, stamp_tax,
             other_fee, extra_fee, transfer_fee, market, raw_json, created_at, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', ?, ?, ?, ?, 0, 0, ?, '', ?, ?, 'MANUAL')""",
            (user_id, job_id, fingerprint, trade_date, trade_time, stock_code, stock_name, action,
             quantity if action == "BUY" else -quantity, quantity, deal_price, quantity * deal_price,
             commission, stamp_tax, transfer_fee, json.dumps(raw, ensure_ascii=False), created_at),
        )
        rebuild_fifo(db, user_id)
    flash(f"已补录当天成交：{stock_code} {stock_name}，次日导入正式交割单时会自动对账。", "success")
    return redirect(url_for("statements"))


@app.post("/admin/manual/paste")
def paste_manual_executions():
    trade_date = request.form.get("trade_date", "").strip()
    text = request.form.get("trades", "").strip()
    if not trade_date or not text:
        flash("批量补录失败：请选择成交日期并粘贴成交文本", "error")
        return redirect(url_for("statements"))
    try:
        normalized_date = normalize_date(trade_date)
        pasted_rows = [row for row in csv.reader(io.StringIO(text), delimiter="\t") if any(cell.strip() for cell in row)]
        rows = [["成交日期", *pasted_rows[0]], *[[normalized_date, *row] for row in pasted_rows[1:]]]
        parsed, errors, ignored, _ = parse_rows(rows)
        if errors:
            raise ValueError("；".join(errors[:5]))
        if ignored:
            raise ValueError("粘贴内容包含非买卖记录")
        if not parsed:
            raise ValueError("没有识别到可补录的买卖成交")
    except (ValueError, TypeError, csv.Error) as error:
        flash(f"批量补录失败：{error}", "error")
        return redirect(url_for("statements"))

    user_id = current_user_id()
    job_id = uuid.uuid4().hex
    created_at = datetime.now().isoformat(timespec="seconds")
    inserted = 0
    with db_connect() as db:
        db.execute(
            """INSERT INTO import_jobs
            (id, user_id, filename, status, total_rows, valid_rows, invalid_rows, duplicate_rows, created_at, imported_at)
            VALUES (?, ?, '当天成交文本补录', 'IMPORTED', ?, ?, 0, 0, ?, ?)""",
            (job_id, user_id, len(parsed), len(parsed), created_at, created_at),
        )
        for row in parsed:
            identity = f"BROKER_TODAY|{user_id}|{row['trade_date']}|{row['deal_id']}|{row['stock_code']}|{row['action']}|{row['quantity']}|{row['deal_price']}"
            fingerprint = hashlib.sha256(identity.encode("utf-8")).hexdigest()
            cursor = db.execute(
                """INSERT INTO executions
                (user_id, import_job_id, fingerprint, trade_date, trade_time, stock_code, stock_name, action,
                 raw_quantity, quantity, deal_id, deal_price, deal_amount, commission, stamp_tax,
                 other_fee, extra_fee, transfer_fee, market, raw_json, created_at, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, 0, '', ?, ?, 'BROKER_TODAY')
                ON CONFLICT(user_id, fingerprint) DO NOTHING""",
                (user_id, job_id, fingerprint, row["trade_date"], row["trade_time"], row["stock_code"],
                 row["stock_name"], row["action"], row["raw_quantity"], row["quantity"], row["deal_id"],
                 row["deal_price"], row["deal_amount"], json.dumps(row["raw"], ensure_ascii=False), created_at),
            )
            inserted += cursor.rowcount
        db.execute("UPDATE import_jobs SET duplicate_rows = ? WHERE id = ?", (len(parsed) - inserted, job_id))
        rebuild_fifo(db, user_id)
    flash(f"当天成交补录完成：新增 {inserted} 条，跳过重复 {len(parsed) - inserted} 条。", "success")
    return redirect(url_for("statements"))


@app.get("/admin/import")
def import_view():
    with db_connect() as db:
        counts = db.execute("SELECT COUNT(*) executions, COUNT(DISTINCT stock_code) stocks FROM executions WHERE user_id = ?", (current_user_id(),)).fetchone()
    return render_template("index.html", page="import", counts=counts)


@app.post("/admin/preview")
def preview_import():
    uploaded = request.files.get("statement")
    if not uploaded or not uploaded.filename:
        flash("请选择交割单文件", "error")
        return redirect(url_for("import_view"))
    try:
        parsed, errors, ignored, total_rows = parse_rows(read_upload(uploaded))
        if not parsed:
            raise ValueError("没有识别到可导入的成交记录")
        job_id = uuid.uuid4().hex
        user_id = current_user_id()
        with db_connect() as db:
            existing = {row["fingerprint"] for row in db.execute("SELECT fingerprint FROM executions WHERE user_id = ?", (user_id,))}
            duplicate_rows = sum(row["fingerprint"] in existing for row in parsed)
            db.execute(
                """INSERT INTO import_jobs
                (id, user_id, filename, status, total_rows, valid_rows, invalid_rows, duplicate_rows, payload, errors, created_at)
                VALUES (?, ?, ?, 'PREVIEW', ?, ?, ?, ?, ?, ?, ?)""",
                (job_id, user_id, uploaded.filename, total_rows, len(parsed), len(errors), duplicate_rows,
                 json.dumps(parsed, ensure_ascii=False), json.dumps(errors, ensure_ascii=False), datetime.now().isoformat(timespec="microseconds")),
            )
            counts = db.execute("SELECT COUNT(*) executions, COUNT(DISTINCT stock_code) stocks FROM executions WHERE user_id = ?", (user_id,)).fetchone()
        return render_template(
            "index.html",
            page="import",
            preview=parsed[:100],
            errors=errors,
            job_id=job_id,
            filename=uploaded.filename,
            total_rows=total_rows,
            valid_rows=len(parsed),
            ignored=ignored,
            duplicate_rows=duplicate_rows,
            importable_rows=len(parsed) - duplicate_rows,
            counts=counts,
        )
    except (ValueError, TypeError, OSError) as error:
        flash(str(error), "error")
        return redirect(url_for("import_view"))


@app.post("/admin/import/<job_id>")
def confirm_import(job_id: str):
    user_id = current_user_id()
    with db_connect() as db:
        job = db.execute("SELECT * FROM import_jobs WHERE id = ? AND user_id = ?", (job_id, user_id)).fetchone()
        if not job or job["status"] != "PREVIEW" or not job["payload"]:
            abort(404)
        rows = json.loads(job["payload"])
        inserted = 0
        reconciled = 0
        for row in rows:
            existing = db.execute(
                "SELECT id FROM executions WHERE user_id = ? AND fingerprint = ?",
                (user_id, row["fingerprint"]),
            ).fetchone()
            if existing:
                continue
            manual = db.execute(
                """SELECT id FROM executions
                WHERE user_id = ? AND source = 'MANUAL' AND trade_date = ? AND stock_code = ? AND action = ?
                  AND ABS(quantity - ?) < 0.000001 AND ABS(deal_price - ?) < 0.000001
                ORDER BY id LIMIT 1""",
                (user_id, row["trade_date"], row["stock_code"], row["action"], row["quantity"], row["deal_price"]),
            ).fetchone()
            if manual:
                db.execute(
                    """UPDATE executions SET import_job_id = ?, fingerprint = ?, trade_time = ?, stock_name = ?,
                    raw_quantity = ?, deal_id = ?, deal_amount = ?, commission = ?, stamp_tax = ?, other_fee = ?,
                    extra_fee = ?, transfer_fee = ?, market = ?, raw_json = ?, source = 'STATEMENT'
                    WHERE id = ? AND user_id = ?""",
                    (job_id, row["fingerprint"], row["trade_time"], row["stock_name"], row["raw_quantity"],
                     row["deal_id"], row["deal_amount"], row["commission"], row["stamp_tax"], row["other_fee"],
                     row["extra_fee"], row["transfer_fee"], row["market"], json.dumps(row["raw"], ensure_ascii=False),
                     manual["id"], user_id),
                )
                reconciled += 1
                continue
            cursor = db.execute(
                """INSERT INTO executions
                (user_id, import_job_id, fingerprint, trade_date, trade_time, stock_code, stock_name, action,
                 raw_quantity, quantity, deal_id, deal_price, deal_amount, commission, stamp_tax,
                 other_fee, extra_fee, transfer_fee, market, raw_json, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(user_id, fingerprint) DO NOTHING""",
                (user_id, job_id, row["fingerprint"], row["trade_date"], row["trade_time"], row["stock_code"],
                 row["stock_name"], row["action"], row["raw_quantity"], row["quantity"], row["deal_id"],
                 row["deal_price"], row["deal_amount"], row["commission"], row["stamp_tax"], row["other_fee"],
                 row["extra_fee"], row["transfer_fee"], row["market"], json.dumps(row["raw"], ensure_ascii=False),
                 datetime.now().isoformat(timespec="seconds")),
            )
            inserted += cursor.rowcount
        rebuild_fifo(db, user_id)
        duplicate_rows = len(rows) - inserted - reconciled
        db.execute(
            "UPDATE import_jobs SET status = 'IMPORTED', payload = NULL, imported_at = ?, duplicate_rows = ? WHERE id = ? AND user_id = ?",
            (datetime.now().isoformat(timespec="seconds"), duplicate_rows, job_id, user_id),
        )
    flash(f"导入完成：新增 {inserted} 条，转正补录 {reconciled} 条，跳过重复 {duplicate_rows} 条，已重新计算 FIFO。", "success")
    return redirect(url_for("statements"))


@app.post("/admin/rebuild")
def rebuild():
    with db_connect() as db:
        rebuild_fifo(db, current_user_id())
    flash("FIFO 匹配和分析数据已重新计算", "success")
    return redirect(url_for("statements"))


@app.get("/admin/database/export")
def export_database():
    with tempfile.TemporaryDirectory() as directory:
        backup_path = Path(directory) / "stocknotes-backup.db"
        source = db_connect()
        backup = sqlite3.connect(backup_path)
        try:
            source.backup(backup)
            if backup.execute("PRAGMA foreign_key_check").fetchone() is not None:
                app.logger.error("Database backup failed foreign key validation")
                abort(500, description="数据库备份完整性检查失败")
        except sqlite3.Error:
            app.logger.exception("Failed to create database backup")
            abort(500, description="数据库备份生成失败")
        finally:
            backup.close()
            source.close()
        try:
            payload = io.BytesIO(backup_path.read_bytes())
        except OSError:
            app.logger.exception("Failed to read database backup")
            abort(500, description="数据库备份生成失败")

    filename = f"stocknotes-backup-{datetime.now():%Y%m%d-%H%M%S}.db"
    response = send_file(payload, mimetype="application/vnd.sqlite3", as_attachment=True, download_name=filename)
    response.headers["Cache-Control"] = "no-store"
    return response


init_db()

if __name__ == "__main__":
    start_auto_sync_scheduler()
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", "5001")), debug=True, use_reloader=False)
